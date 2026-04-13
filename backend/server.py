from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Query, Depends, Header, BackgroundTasks
from fastapi.responses import JSONResponse
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from io import BytesIO
import json
import asyncio
from concurrent.futures import ThreadPoolExecutor
from collections import defaultdict
import re
import gc
import math

# Thread pool for CPU-bound Excel parsing
_parse_executor = ThreadPoolExecutor(max_workers=1)

BATCH_INSERT_SIZE = 2000

from dotenv import load_dotenv
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env', override=False)

# Import shared modules
from database import db, client
from auth import (
    get_current_user, verify_password, get_password_hash,
    create_access_token, security
)
from models import (
    Transaction, OpeningStock, PhysicalStock, MasterItem, ItemMapping,
    PurchaseLedger, User, LoginRequest, CreateUserRequest, Token,
    ActionHistory, ResetRequest, StampAssignment, OrderCreate,
    HistoricalUploadRequest, ItemGroup
)
from services.helpers import (
    normalize_stamp, get_column_value, parse_labor_value,
    normalize_date, stamp_sort_key, save_action, auto_normalize_stamps
)
from services.stock_service import get_current_inventory, get_effective_physical_base_for_date, _flat_base_from_inventory
from services.group_utils import build_group_maps, build_group_ledger, resolve_to_leader

# Simple TTL cache for heavy inventory computations
import time

class InventoryCache:
    """In-memory cache with TTL for expensive inventory computations."""
    def __init__(self, ttl_seconds=30):
        self._cache = {}
        self._ttl = ttl_seconds

    def get(self, key):
        entry = self._cache.get(key)
        if entry and (time.time() - entry['ts']) < self._ttl:
            return entry['data']
        if entry:
            del self._cache[key]
        return None

    def set(self, key, data):
        self._cache[key] = {'data': data, 'ts': time.time()}

    def invalidate(self, key=None):
        if key:
            self._cache.pop(key, None)
        else:
            self._cache.clear()

_inv_cache = InventoryCache(ttl_seconds=30)

app = FastAPI()

@app.on_event("startup")
async def create_upload_indexes():
    """Create indexes for chunked upload collections and clean stale tasks"""
    await db.upload_sessions.create_index("upload_id", unique=True)
    await db.upload_chunks.create_index([("upload_id", 1), ("chunk_index", 1)], unique=True)
    # Indexes for fast aggregation on historical_transactions (200k+ docs)
    await db.historical_transactions.create_index([("historical_year", 1), ("type", 1)])
    await db.historical_transactions.create_index([("type", 1), ("item_name", 1)])
    await db.historical_transactions.create_index("batch_id")
    await db.transactions.create_index([("type", 1), ("item_name", 1)])

    # Performance indexes for inventory computation (critical at 17K+ transactions)
    await db.transactions.create_index("date")
    await db.transactions.create_index([("date", 1), ("type", 1)])
    await db.inventory_baselines.create_index("item_key")
    await db.inventory_baselines.create_index("baseline_date")
    await db.inventory_baselines.create_index([("item_key", 1), ("baseline_date", -1)])
    await db.physical_stock.create_index("verification_date")
    await db.physical_stock.create_index([("verification_date", 1), ("item_name", 1)])
    await db.stock_entries.create_index([("stamp", 1), ("status", 1)])
    await db.stock_entries.create_index([("stamp", 1), ("entry_day", 1)])
    await db.stock_entries.create_index([("stamp", 1), ("verification_date", 1), ("status", 1)])
    await db.stock_entries.create_index([("entered_by", 1), ("status", 1)])
    await db.stock_entries.create_index("entry_date")
    await db.polythene_adjustments.create_index("date")
    await db.polythene_adjustments.create_index("item_name")
    await db.master_items.create_index("stamp")
    await db.master_items.create_index("item_name", unique=True)
    await db.notifications.create_index([("target_user", 1), ("read", 1)])
    await db.notifications.create_index("timestamp")
    await db.physical_stock_update_sessions.create_index("verification_date")
    await db.activity_log.create_index("timestamp")
    await db.stamp_approvals.create_index([("stamp", 1), ("approval_day", 1)])

    # Mark any orphaned "processing" tasks as failed (from OOM/crash during previous run)
    stale = await db.upload_sessions.update_many(
        {"status": "processing"},
        {"$set": {"status": "error", "error": "Server restarted during processing. Please re-upload."}}
    )
    if stale.modified_count > 0:
        logger.info(f"Cleaned {stale.modified_count} stale upload tasks on startup")

# Health check endpoints
@app.get("/health")
async def health_check():
    return {"status": "healthy", "service": "stockbud-backend", "version": "2.1-individual-stock"}

@app.get("/api/health")
async def api_health_check():
    return {"status": "healthy", "service": "stockbud-backend", "version": "2.1-individual-stock"}

api_router = APIRouter(prefix="/api")

# Include modular routers
from routes.seasonal_analytics import router as seasonal_router
api_router.include_router(seasonal_router)

# ==================== EXCEL PARSING ====================

def _resolve_col(df_columns_set, possible_names):
    """Resolve which column name exists in the DataFrame - called once per column"""
    for name in possible_names:
        if name in df_columns_set:
            return name
    return None


def _safe_float(val, default=0.0):
    """Fast float conversion — handles comma-separated numbers (e.g., 1,705.433)"""
    if val is None or val == '':
        return default
    try:
        f = float(val)
        return default if math.isnan(f) else f
    except (ValueError, TypeError):
        # Try removing commas (Indian/international thousand separators)
        try:
            f = float(str(val).replace(',', ''))
            return default if math.isnan(f) else f
        except (ValueError, TypeError):
            return default


def _safe_str(val, default=''):
    """Fast string conversion"""
    if val is None:
        return default
    s = str(val).strip()
    return default if s in ('', 'nan', 'None') else s


def _safe_int(val, default=0):
    """Fast int conversion"""
    try:
        f = float(val) if val not in (None, '', 'nan') else default
        return default if math.isnan(f) else int(f)
    except (ValueError, TypeError):
        return default


def _read_excel_once(file_content: bytes):
    """Read Excel file ONCE and detect header row efficiently"""
    import pandas as pd
    df = pd.read_excel(BytesIO(file_content), header=None, dtype=str)
    return _detect_header_and_clean(df)


def _read_excel_from_path(file_path: str):
    """Read Excel file from disk path (memory-efficient for large files)"""
    import pandas as pd
    df = pd.read_excel(file_path, header=None, dtype=str, engine='openpyxl')
    return _detect_header_and_clean(df)


def _detect_header_and_clean(df):
    """Detect header row and clean up DataFrame"""
    header_row_idx = None
    search_limit = min(20, len(df))
    for idx in range(search_limit):
        row_str = ' '.join(str(val).lower() for val in df.iloc[idx] if val is not None and str(val) != 'nan')
        if 'item name' in row_str or 'particular' in row_str or 'party name' in row_str or 'lnarr' in row_str:
            header_row_idx = idx
            break
    
    if header_row_idx is not None:
        df.columns = df.iloc[header_row_idx].astype(str).str.strip()
        df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
    else:
        df.columns = df.iloc[0].astype(str).str.strip()
        df = df.iloc[1:].reset_index(drop=True)
    
    return df


def parse_excel_file(file_content, file_type: str) -> List[Dict]:
    """Parse Excel file using fast dict-based iteration. Weights in KG -> grams.
    file_content can be bytes OR a file path string (for memory-efficient large file processing)."""
    try:
        if isinstance(file_content, str):
            df = _read_excel_from_path(file_content)
        else:
            df = _read_excel_once(file_content)
        cols = set(df.columns)
        KG_TO_GRAMS = 1000

        # Convert DataFrame to list of dicts ONCE (much faster than iterrows)
        raw_rows = df.to_dict('records')

        if file_type == 'purchase':
            item_col = _resolve_col(cols, ['Item Name', 'Particular', 'item name'])
            type_col = _resolve_col(cols, ['Type', 'type'])
            tag_col = _resolve_col(cols, ['Tag.No.', 'Tag No', 'tag no'])
            wt_rs_col = _resolve_col(cols, ['Wt/Rs', 'Wt Rs'])
            total_col = _resolve_col(cols, ['Total', 'total'])
            tunch_col = _resolve_col(cols, ['Tunch', 'tunch'])
            wstg_col = _resolve_col(cols, ['Wstg', 'wstg'])
            date_col = _resolve_col(cols, ['Date', 'date'])
            refno_col = _resolve_col(cols, ['Refno', 'refno', 'Ref No'])
            party_col = _resolve_col(cols, ['Party Name', 'party name', 'Party'])
            stamp_col = _resolve_col(cols, ['Stamp', 'stamp'])
            gr_col = _resolve_col(cols, ['Gr.Wt.', 'Gr Wt', 'Gross Wt'])
            net_col = _resolve_col(cols, ['Net.Wt.', 'Net Wt'])
            fine_col = _resolve_col(cols, ['Fine', 'Sil.Fine', 'Sil Fine', 'Silver Fine'])
            dia_col = _resolve_col(cols, ['Dia.Wt.', 'Dia Wt'])
            stn_col = _resolve_col(cols, ['Stn.Wt.', 'Stn Wt'])
            rate_col = _resolve_col(cols, ['Rate', 'rate'])
            pc_col = _resolve_col(cols, ['Pc', 'pc', 'Pieces'])

            records = []
            for r in raw_rows:
                item_name = _safe_str(r.get(item_col) if item_col else None)
                if len(item_name) < 2:
                    continue
                trans_type = _safe_str(r.get(type_col) if type_col else None, 'P').upper()
                if trans_type.isdigit():
                    continue

                tag_no = _safe_str(r.get(tag_col) if tag_col else None)
                labor_val, labor_on = parse_labor_value(tag_no)
                wt_rs = r.get(wt_rs_col) if wt_rs_col else None
                if wt_rs and str(wt_rs).replace('.', '').isdigit():
                    labor_val = float(wt_rs)

                total_amount = _safe_float(r.get(total_col) if total_col else None)
                tunch_v = _safe_float(r.get(tunch_col) if tunch_col else None)
                wstg_v = _safe_float(r.get(wstg_col) if wstg_col else None)
                purchase_tunch = tunch_v + wstg_v

                records.append({
                    'date': normalize_date(r.get(date_col) if date_col else ''),
                    'type': 'purchase' if trans_type in ('P', 'PURCHASE') else 'purchase_return',
                    'refno': _safe_str(r.get(refno_col) if refno_col else None),
                    'party_name': _safe_str(r.get(party_col) if party_col else None),
                    'item_name': item_name,
                    'stamp': normalize_stamp(r.get(stamp_col) if stamp_col else ''),
                    'tag_no': tag_no,
                    'gr_wt': _safe_float(r.get(gr_col) if gr_col else None) * KG_TO_GRAMS,
                    'net_wt': _safe_float(r.get(net_col) if net_col else None) * KG_TO_GRAMS,
                    'fine': _safe_float(r.get(fine_col) if fine_col else None) * KG_TO_GRAMS,
                    'labor': labor_val,
                    'labor_on': labor_on,
                    'dia_wt': _safe_float(r.get(dia_col) if dia_col else None) * KG_TO_GRAMS,
                    'stn_wt': _safe_float(r.get(stn_col) if stn_col else None) * KG_TO_GRAMS,
                    'tunch': str(purchase_tunch),
                    'rate': _safe_float(r.get(rate_col) if rate_col else None),
                    'total_pc': _safe_int(r.get(pc_col) if pc_col else None),
                    'total_amount': total_amount,
                })
            return records

        elif file_type == 'sale':
            item_col = _resolve_col(cols, ['Item Name', 'Particular', 'item name'])
            type_col = _resolve_col(cols, ['Type', 'type'])
            tag_col = _resolve_col(cols, ['Lbr. On Tag.No.', 'Tag.No.', 'Tag No'])
            on_col = _resolve_col(cols, ['On', 'on'])
            total_col = _resolve_col(cols, ['Total', 'total'])
            tunch_col = _resolve_col(cols, ['Tunch', 'tunch'])
            date_col = _resolve_col(cols, ['Date', 'date'])
            refno_col = _resolve_col(cols, ['Refno', 'refno', 'Ref No'])
            party_col = _resolve_col(cols, ['Party Name', 'party name', 'Party'])
            stamp_col = _resolve_col(cols, ['Stamp', 'stamp'])
            gr_col = _resolve_col(cols, ['Gr.Wt.', 'Gr Wt', 'Gross Wt'])
            net_col = _resolve_col(cols, ['Gold Std.', 'Net.Wt.', 'Net Wt'])
            fine_col = _resolve_col(cols, ['Fine', 'Sil.Fine', 'Sil Fine'])
            dia_col = _resolve_col(cols, ['Dia.Wt.', 'Dia Wt'])
            stn_col = _resolve_col(cols, ['Stn.Wt.', 'Stn Wt'])
            taxable_col = _resolve_col(cols, ['Taxable Val.', 'Taxable Value'])
            pc_col = _resolve_col(cols, ['Pc', 'pc'])

            records = []
            for r in raw_rows:
                item_name = _safe_str(r.get(item_col) if item_col else None)
                if len(item_name) < 2:
                    continue
                trans_type = _safe_str(r.get(type_col) if type_col else None, 'S').upper()
                if trans_type.isdigit():
                    continue

                tag_no = _safe_str(r.get(tag_col) if tag_col else None)
                labor_val, labor_on = parse_labor_value(tag_no)
                on_val = r.get(on_col) if on_col else None
                if on_val and str(on_val).replace('.', '').isdigit():
                    labor_val = float(on_val)

                total_amount = _safe_float(r.get(total_col) if total_col else None)
                sale_tunch = _safe_float(r.get(tunch_col) if tunch_col else None)

                records.append({
                    'type': 'sale' if trans_type in ('S', 'SALE') else 'sale_return',
                    'date': normalize_date(r.get(date_col) if date_col else ''),
                    'refno': _safe_str(r.get(refno_col) if refno_col else None),
                    'party_name': _safe_str(r.get(party_col) if party_col else None),
                    'item_name': item_name,
                    'stamp': normalize_stamp(r.get(stamp_col) if stamp_col else ''),
                    'tag_no': tag_no,
                    'gr_wt': _safe_float(r.get(gr_col) if gr_col else None) * KG_TO_GRAMS,
                    'net_wt': _safe_float(r.get(net_col) if net_col else None) * KG_TO_GRAMS,
                    'fine': _safe_float(r.get(fine_col) if fine_col else None) * KG_TO_GRAMS,
                    'labor': labor_val,
                    'labor_on': labor_on,
                    'dia_wt': _safe_float(r.get(dia_col) if dia_col else None) * KG_TO_GRAMS,
                    'stn_wt': _safe_float(r.get(stn_col) if stn_col else None) * KG_TO_GRAMS,
                    'tunch': str(sale_tunch),
                    'total_amount': total_amount,
                    'taxable_value': _safe_float(r.get(taxable_col) if taxable_col else None),
                    'total_pc': _safe_int(r.get(pc_col) if pc_col else None),
                })
            return records

        elif file_type == 'branch_transfer':
            item_col = _resolve_col(cols, ['Lnarr'])
            type_col = _resolve_col(cols, ['Type'])
            date_col = _resolve_col(cols, ['Date'])
            refno_col = _resolve_col(cols, ['Refno'])
            gr_col = _resolve_col(cols, ['Gr.Wt.'])
            net_col = _resolve_col(cols, ['Net.Wt.'])

            records = []
            for r in raw_rows:
                item_name = _safe_str(r.get(item_col) if item_col else None)
                if len(item_name) < 2:
                    continue
                if 'opening' in item_name.lower() and 'balance' in item_name.lower():
                    continue
                if 'total' in item_name.lower():
                    continue
                if item_name.isdigit():
                    continue

                trans_type = _safe_str(r.get(type_col) if type_col else None).upper()
                if not trans_type or trans_type in ('', 'NAN'):
                    continue

                records.append({
                    'type': 'receive' if trans_type == 'R' else 'issue',
                    'date': normalize_date(r.get(date_col) if date_col else ''),
                    'refno': _safe_str(r.get(refno_col) if refno_col else None),
                    'party_name': 'MMI Jewelly Branch',
                    'item_name': item_name,
                    'stamp': '',
                    'tag_no': '',
                    'gr_wt': _safe_float(r.get(gr_col) if gr_col else None) * KG_TO_GRAMS,
                    'net_wt': _safe_float(r.get(net_col) if net_col else None) * KG_TO_GRAMS,
                    'fine': 0.0,
                    'labor': 0.0,
                    'labor_on': None,
                    'dia_wt': 0.0,
                    'stn_wt': 0.0,
                    'tunch': '0',
                    'rate': 0.0,
                    'total_pc': 0,
                    'total_amount': 0.0,
                    'taxable_value': 0.0,
                })
            return records

        elif file_type == 'opening_stock':
            item_col = _resolve_col(cols, ['Item Name', 'Particular', 'item name', 'Stock'])
            stamp_col = _resolve_col(cols, ['Stamp', 'stamp'])
            unit_col = _resolve_col(cols, ['Unit', 'unit'])
            pc_col = _resolve_col(cols, ['Pc', 'pc', 'Pieces'])
            gr_col = _resolve_col(cols, ['Gr.Wt.', 'Gr Wt', 'Gross Wt'])
            net_col = _resolve_col(cols, ['Gold Std.', 'Net.Wt.', 'Net Wt'])
            fine_col = _resolve_col(cols, ['Sil.Fine', 'Fine', 'fine'])
            rate_col = _resolve_col(cols, ['Rate', 'rate'])
            total_col = _resolve_col(cols, ['Total', 'total'])
            tunch_col = _resolve_col(cols, ['Tunch', 'tunch'])
            wstg_col = _resolve_col(cols, ['Wstg', 'wstg'])

            records = []
            for r in raw_rows:
                item_name = _safe_str(r.get(item_col) if item_col else None)
                if len(item_name) < 2:
                    continue
                if 'total' in item_name.lower():
                    continue

                tunch_v = _safe_float(r.get(tunch_col) if tunch_col else None)
                wstg_v = _safe_float(r.get(wstg_col) if wstg_col else None)

                records.append({
                    'item_name': item_name,
                    'stamp': normalize_stamp(r.get(stamp_col) if stamp_col else ''),
                    'unit': _safe_str(r.get(unit_col) if unit_col else None),
                    'pc': _safe_int(r.get(pc_col) if pc_col else None),
                    'gr_wt': _safe_float(r.get(gr_col) if gr_col else None) * KG_TO_GRAMS,
                    'net_wt': _safe_float(r.get(net_col) if net_col else None) * KG_TO_GRAMS,
                    'fine': _safe_float(r.get(fine_col) if fine_col else None) * KG_TO_GRAMS,
                    'labor_wt': 0.0,
                    'labor_rs': 0.0,
                    'rate': _safe_float(r.get(rate_col) if rate_col else None),
                    'total': _safe_float(r.get(total_col) if total_col else None),
                })
            return records

        elif file_type == 'physical_stock':
            item_col = _resolve_col(cols, ['Item Name', 'Particular', 'item name', 'Stock'])
            stamp_col = _resolve_col(cols, ['Stamp', 'stamp'])
            gr_col = _resolve_col(cols, ['Gr.Wt.', 'Gr Wt', 'Gross Wt', 'Gross Weight'])
            net_col = _resolve_col(cols, ['Net.Wt.', 'Net Wt', 'Net Weight', 'Gold Std.'])
            pc_col = _resolve_col(cols, ['Pc', 'pc', 'Pieces', 'Pcs'])
            fine_col = _resolve_col(cols, ['Sil.Fine', 'Fine', 'fine'])

            has_net = net_col is not None
            records = []
            for r in raw_rows:
                item_name = _safe_str(r.get(item_col) if item_col else None)
                if len(item_name) < 2:
                    continue
                if 'total' in item_name.lower():
                    continue
                rec = {
                    'item_name': item_name,
                    'stamp': normalize_stamp(r.get(stamp_col) if stamp_col else ''),
                    'gr_wt': _safe_float(r.get(gr_col) if gr_col else None) * KG_TO_GRAMS,
                    'has_net': has_net,
                    'pc': _safe_int(r.get(pc_col) if pc_col else None),
                    'fine': _safe_float(r.get(fine_col) if fine_col else None) * KG_TO_GRAMS,
                }
                if has_net:
                    rec['net_wt'] = _safe_float(r.get(net_col) if net_col else None) * KG_TO_GRAMS
                else:
                    rec['net_wt'] = 0.0
                records.append(rec)
            return records

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing Excel file: {str(e)}")


def parse_excel_streaming(file_path: str, file_type: str) -> List[Dict]:
    """Memory-efficient Excel parser using openpyxl read-only/streaming mode.
    Avoids loading entire DataFrame into memory — critical for large files on constrained pods."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Step 1: Read first 20 rows to detect header
        header_names = None
        header_row_idx = None
        first_rows = []
        for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True)):
            str_vals = [str(v).strip() if v is not None else '' for v in row]
            first_rows.append(str_vals)
            if header_row_idx is None:
                row_str = ' '.join(s.lower() for s in str_vals if s)
                if 'item name' in row_str or 'particular' in row_str or 'party name' in row_str or 'lnarr' in row_str:
                    header_row_idx = i
                    header_names = str_vals

        if header_names is None:
            header_names = first_rows[0] if first_rows else []
            header_row_idx = 0

        wb.close()

        # Step 2: Re-open and stream all rows after header, converting each to a dict
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        records = []
        KG_TO_GRAMS = 1000

        # Build column name set for resolver
        cols_set = set(header_names)

        if file_type == 'purchase':
            item_col = _resolve_col(cols_set, ['Item Name', 'Particular', 'item name'])
            type_col = _resolve_col(cols_set, ['Type', 'type'])
            tag_col = _resolve_col(cols_set, ['Tag.No.', 'Tag No', 'tag no'])
            wt_rs_col = _resolve_col(cols_set, ['Wt/Rs', 'Wt Rs'])
            total_col = _resolve_col(cols_set, ['Total', 'total'])
            tunch_col = _resolve_col(cols_set, ['Tunch', 'tunch'])
            wstg_col = _resolve_col(cols_set, ['Wstg', 'wstg'])
            date_col = _resolve_col(cols_set, ['Date', 'date'])
            refno_col = _resolve_col(cols_set, ['Refno', 'refno', 'Ref No'])
            party_col = _resolve_col(cols_set, ['Party Name', 'party name', 'Party'])
            stamp_col = _resolve_col(cols_set, ['Stamp', 'stamp'])
            gr_col = _resolve_col(cols_set, ['Gr.Wt.', 'Gr Wt', 'Gross Wt'])
            net_col = _resolve_col(cols_set, ['Net.Wt.', 'Net Wt'])
            fine_col = _resolve_col(cols_set, ['Fine', 'Sil.Fine', 'Sil Fine', 'Silver Fine'])
            dia_col = _resolve_col(cols_set, ['Dia.Wt.', 'Dia Wt'])
            stn_col = _resolve_col(cols_set, ['Stn.Wt.', 'Stn Wt'])
            rate_col = _resolve_col(cols_set, ['Rate', 'rate'])
            pc_col = _resolve_col(cols_set, ['Pc', 'pc', 'Pieces'])

            col_map = {name: idx for idx, name in enumerate(header_names)}

            for row_num, row in enumerate(ws.iter_rows(values_only=True)):
                if row_num <= header_row_idx:
                    continue
                def _get(col_name):
                    if not col_name or col_name not in col_map:
                        return None
                    idx = col_map[col_name]
                    return str(row[idx]).strip() if idx < len(row) and row[idx] is not None else None

                item_name = _safe_str(_get(item_col))
                if len(item_name) < 2:
                    continue
                trans_type = _safe_str(_get(type_col), 'P').upper()
                if trans_type.isdigit():
                    continue
                tag_no = _safe_str(_get(tag_col))
                labor_val, labor_on = parse_labor_value(tag_no)
                wt_rs = _get(wt_rs_col)
                if wt_rs and str(wt_rs).replace('.', '').isdigit():
                    labor_val = float(wt_rs)
                total_amount = _safe_float(_get(total_col))
                tunch_v = _safe_float(_get(tunch_col))
                wstg_v = _safe_float(_get(wstg_col))
                purchase_tunch = tunch_v + wstg_v
                records.append({
                    'date': normalize_date(_get(date_col) or ''),
                    'type': 'purchase' if trans_type in ('P', 'PURCHASE') else 'purchase_return',
                    'refno': _safe_str(_get(refno_col)),
                    'party_name': _safe_str(_get(party_col)),
                    'item_name': item_name,
                    'stamp': normalize_stamp(_get(stamp_col) or ''),
                    'tag_no': tag_no,
                    'gr_wt': _safe_float(_get(gr_col)) * KG_TO_GRAMS,
                    'net_wt': _safe_float(_get(net_col)) * KG_TO_GRAMS,
                    'fine': _safe_float(_get(fine_col)) * KG_TO_GRAMS,
                    'labor': labor_val,
                    'labor_on': labor_on,
                    'dia_wt': _safe_float(_get(dia_col)) * KG_TO_GRAMS,
                    'stn_wt': _safe_float(_get(stn_col)) * KG_TO_GRAMS,
                    'tunch': str(purchase_tunch),
                    'rate': _safe_float(_get(rate_col)),
                    'total_pc': _safe_int(_get(pc_col)),
                    'total_amount': total_amount,
                })
        elif file_type == 'sale':
            item_col = _resolve_col(cols_set, ['Item Name', 'Particular', 'item name'])
            type_col = _resolve_col(cols_set, ['Type', 'type'])
            tag_col = _resolve_col(cols_set, ['Lbr. On Tag.No.', 'Tag.No.', 'Tag No'])
            on_col = _resolve_col(cols_set, ['On', 'on'])
            total_col = _resolve_col(cols_set, ['Total', 'total'])
            tunch_col = _resolve_col(cols_set, ['Tunch', 'tunch'])
            date_col = _resolve_col(cols_set, ['Date', 'date'])
            refno_col = _resolve_col(cols_set, ['Refno', 'refno', 'Ref No'])
            party_col = _resolve_col(cols_set, ['Party Name', 'party name', 'Party'])
            stamp_col = _resolve_col(cols_set, ['Stamp', 'stamp'])
            gr_col = _resolve_col(cols_set, ['Gr.Wt.', 'Gr Wt', 'Gross Wt'])
            net_col = _resolve_col(cols_set, ['Gold Std.', 'Net.Wt.', 'Net Wt'])
            fine_col = _resolve_col(cols_set, ['Fine', 'Sil.Fine', 'Sil Fine'])
            dia_col = _resolve_col(cols_set, ['Dia.Wt.', 'Dia Wt'])
            stn_col = _resolve_col(cols_set, ['Stn.Wt.', 'Stn Wt'])
            taxable_col = _resolve_col(cols_set, ['Taxable Val.', 'Taxable Value'])
            pc_col = _resolve_col(cols_set, ['Pc', 'pc'])

            col_map = {name: idx for idx, name in enumerate(header_names)}

            for row_num, row in enumerate(ws.iter_rows(values_only=True)):
                if row_num <= header_row_idx:
                    continue
                def _get(col_name):
                    if not col_name or col_name not in col_map:
                        return None
                    idx = col_map[col_name]
                    return str(row[idx]).strip() if idx < len(row) and row[idx] is not None else None

                item_name = _safe_str(_get(item_col))
                if len(item_name) < 2:
                    continue
                trans_type = _safe_str(_get(type_col), 'S').upper()
                if trans_type.isdigit():
                    continue
                tag_no = _safe_str(_get(tag_col))
                labor_val, labor_on = parse_labor_value(tag_no)
                on_val = _get(on_col)
                if on_val and str(on_val).replace('.', '').isdigit():
                    labor_val = float(on_val)
                total_amount = _safe_float(_get(total_col))
                sale_tunch = _safe_float(_get(tunch_col))
                records.append({
                    'type': 'sale' if trans_type in ('S', 'SALE') else 'sale_return',
                    'date': normalize_date(_get(date_col) or ''),
                    'refno': _safe_str(_get(refno_col)),
                    'party_name': _safe_str(_get(party_col)),
                    'item_name': item_name,
                    'stamp': normalize_stamp(_get(stamp_col) or ''),
                    'tag_no': tag_no,
                    'gr_wt': _safe_float(_get(gr_col)) * KG_TO_GRAMS,
                    'net_wt': _safe_float(_get(net_col)) * KG_TO_GRAMS,
                    'fine': _safe_float(_get(fine_col)) * KG_TO_GRAMS,
                    'labor': labor_val,
                    'labor_on': labor_on,
                    'dia_wt': _safe_float(_get(dia_col)) * KG_TO_GRAMS,
                    'stn_wt': _safe_float(_get(stn_col)) * KG_TO_GRAMS,
                    'tunch': str(sale_tunch),
                    'total_amount': total_amount,
                    'taxable_value': _safe_float(_get(taxable_col)),
                    'total_pc': _safe_int(_get(pc_col)),
                })
        elif file_type == 'opening_stock':
            item_col = _resolve_col(cols_set, ['Item Name', 'Particular', 'item name'])
            stamp_col = _resolve_col(cols_set, ['Stamp', 'stamp'])
            unit_col = _resolve_col(cols_set, ['Unit', 'unit'])
            gr_col = _resolve_col(cols_set, ['Gr.Wt.', 'Gr Wt', 'Gross Wt', 'Gross Weight'])
            net_col = _resolve_col(cols_set, ['Net.Wt.', 'Net Wt', 'Net Weight'])
            fine_col = _resolve_col(cols_set, ['Fine', 'Sil.Fine', 'Silver Fine'])
            pc_col = _resolve_col(cols_set, ['Pc', 'pc', 'Pieces', 'Pcs'])
            rate_col = _resolve_col(cols_set, ['Rate', 'rate'])
            total_col = _resolve_col(cols_set, ['Total', 'total', 'Amount'])

            col_map = {name: idx for idx, name in enumerate(header_names)}

            for row_num, row in enumerate(ws.iter_rows(values_only=True)):
                if row_num <= header_row_idx:
                    continue
                def _get(col_name):
                    if not col_name or col_name not in col_map:
                        return None
                    idx = col_map[col_name]
                    return str(row[idx]).strip() if idx < len(row) and row[idx] is not None else None

                item_name = _safe_str(_get(item_col))
                if len(item_name) < 2:
                    continue
                records.append({
                    'item_name': item_name,
                    'stamp': normalize_stamp(_get(stamp_col) or ''),
                    'unit': _safe_str(_get(unit_col)),
                    'pc': _safe_int(_get(pc_col)),
                    'gr_wt': _safe_float(_get(gr_col)) * KG_TO_GRAMS,
                    'net_wt': _safe_float(_get(net_col)) * KG_TO_GRAMS,
                    'fine': _safe_float(_get(fine_col)) * KG_TO_GRAMS,
                    'labor_wt': 0.0,
                    'labor_rs': 0.0,
                    'rate': _safe_float(_get(rate_col)),
                    'total': _safe_float(_get(total_col)),
                })

        elif file_type == 'physical_stock':
            item_col = _resolve_col(cols_set, ['Item Name', 'Particular', 'item name', 'Stock'])
            stamp_col = _resolve_col(cols_set, ['Stamp', 'stamp'])
            gr_col = _resolve_col(cols_set, ['Gr.Wt.', 'Gr Wt', 'Gross Wt', 'Gross Weight'])
            net_col = _resolve_col(cols_set, ['Net.Wt.', 'Net Wt', 'Net Weight', 'Gold Std.'])
            pc_col = _resolve_col(cols_set, ['Pc', 'pc', 'Pieces', 'Pcs'])
            fine_col = _resolve_col(cols_set, ['Sil.Fine', 'Fine', 'fine'])

            has_net = net_col is not None
            col_map = {name: idx for idx, name in enumerate(header_names)}

            for row_num, row in enumerate(ws.iter_rows(values_only=True)):
                if row_num <= header_row_idx:
                    continue
                def _get(col_name):
                    if not col_name or col_name not in col_map:
                        return None
                    idx = col_map[col_name]
                    return str(row[idx]).strip() if idx < len(row) and row[idx] is not None else None

                item_name = _safe_str(_get(item_col))
                if len(item_name) < 2:
                    continue
                if 'total' in item_name.lower():
                    continue
                rec = {
                    'item_name': item_name,
                    'stamp': normalize_stamp(_get(stamp_col) or ''),
                    'gr_wt': _safe_float(_get(gr_col)) * KG_TO_GRAMS,
                    'has_net': has_net,
                    'pc': _safe_int(_get(pc_col)),
                    'fine': _safe_float(_get(fine_col)) * KG_TO_GRAMS,
                }
                if has_net:
                    rec['net_wt'] = _safe_float(_get(net_col)) * KG_TO_GRAMS
                else:
                    rec['net_wt'] = 0.0
                records.append(rec)

        wb.close()
        logger.info(f"[Streaming parser] Parsed {len(records)} {file_type} records from {file_path}")
        return records
    except Exception as e:
        logger.error(f"[Streaming parser] Error: {e}", exc_info=True)
        return []

# ==================== UPLOAD ENDPOINTS ====================

@api_router.post("/opening-stock/upload")
async def upload_opening_stock(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Upload opening stock - Parse and MERGE items by name (sum weights regardless of stamp)"""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or Manager only")
    content = await file.read()
    
    try:
        # Parse in thread pool so we don't block the event loop
        loop = asyncio.get_event_loop()
        records = await loop.run_in_executor(_parse_executor, parse_excel_file, content, 'opening_stock')
        
        if not records:
            raise HTTPException(status_code=400, detail="No valid records found in file")
        
        # MERGE items by name - sum all weights for same item
        merged_items = {}
        for record in records:
            key = record['item_name'].strip().lower()
            if key not in merged_items:
                merged_items[key] = {
                    'item_name': record['item_name'],
                    'stamp': record.get('stamp', ''),
                    'unit': record.get('unit', ''),
                    'pc': 0,
                    'gr_wt': 0.0,
                    'net_wt': 0.0,
                    'fine': 0.0,
                    'labor_wt': record.get('labor_wt', 0.0),
                    'labor_rs': record.get('labor_rs', 0.0),
                    'rate': record.get('rate', 0.0),
                    'total': 0.0
                }
            
            # Sum weights (including negative values)
            merged_items[key]['gr_wt'] += record.get('gr_wt', 0)
            merged_items[key]['net_wt'] += record.get('net_wt', 0)
            merged_items[key]['fine'] += record.get('fine', 0)
            merged_items[key]['pc'] += record.get('pc', 0)
            merged_items[key]['total'] += record.get('total', 0)
            
            # Keep stamp if this entry has one
            if record.get('stamp') and not merged_items[key]['stamp']:
                merged_items[key]['stamp'] = record['stamp']
        
        # Clear existing opening stock
        await db.opening_stock.delete_many({})
        
        # Insert merged items
        stock_items = [OpeningStock(**item).model_dump() for item in merged_items.values()]
        await db.opening_stock.insert_many(stock_items)
        
        # Calculate totals for response
        total_net_wt = sum(item['net_wt'] for item in stock_items)
        total_gr_wt = sum(item['gr_wt'] for item in stock_items)
        
        await save_action('upload_opening_stock', f"Uploaded {len(stock_items)} merged opening stock items, total: {total_net_wt/1000:.3f} kg")
        
        # Auto-normalize stamps after upload
        await auto_normalize_stamps()
        
        return {
            "success": True,
            "count": len(stock_items),
            "original_rows": len(records),
            "merged_items": len(stock_items),
            "total_net_wt_kg": round(total_net_wt/1000, 3),
            "total_gr_wt_kg": round(total_gr_wt/1000, 3),
            "message": f"Merged {len(records)} rows into {len(stock_items)} items. Total: {total_net_wt/1000:.3f} kg"
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")


# ==================== AUTHENTICATION ENDPOINTS ====================

@api_router.post("/auth/login", response_model=Token)
async def login(request: LoginRequest):
    """Login endpoint - Returns JWT token (18 hours for regular users, 365 days for admin)"""
    user = await db.users.find_one({"username": request.username})
    
    if not user or not verify_password(request.password, user['password_hash']):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    if not user.get('is_active', True):
        raise HTTPException(status_code=401, detail="User account is inactive")
    
    # Create JWT token with different expiry for admin (perpetual - 365 days)
    if user.get('role') == 'admin':
        access_token = create_access_token({"sub": user['username'], "role": user['role']}, expire_hours=365*24)
    else:
        # Regular users: 18 hours
        access_token = create_access_token({"sub": user['username'], "role": user['role']})
    
    # Remove sensitive data
    user_data = {k: v for k, v in user.items() if k not in ['_id', 'password_hash']}
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": user_data
    }

@api_router.get("/auth/me")
async def get_current_user_info(current_user: dict = Depends(get_current_user)):
    """Get current logged-in user info"""
    return current_user

@api_router.post("/auth/logout")
async def logout():
    """Logout endpoint (client-side token removal)"""
    return {"message": "Logged out successfully"}

# ==================== USER MANAGEMENT (Admin Only) ====================

@api_router.post("/users/create")
async def create_user(
    request: CreateUserRequest,
    current_user: dict = Depends(get_current_user)
):
    """Create new user (Admin only)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can create users")
    
    # Check if username exists
    existing = await db.users.find_one({"username": request.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Validate role
    if request.role not in ['admin', 'manager', 'executive', 'polythene_executive']:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    # Create user
    user = User(
        username=request.username,
        password_hash=get_password_hash(request.password),
        full_name=request.full_name,
        role=request.role,
        created_by=current_user['username']
    )
    
    await db.users.insert_one(user.model_dump())
    
    return {
        "success": True,
        "message": f"User {request.username} created successfully",
        "user": {
            "username": user.username,
            "full_name": user.full_name,
            "role": user.role
        }
    }

@api_router.get("/users/list")
async def list_users(current_user: dict = Depends(get_current_user)):
    """List all users (Admin and Manager can view)"""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(100)
    return users

@api_router.delete("/users/{username}")
async def delete_user(
    username: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete user (Admin only)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can delete users")
    
    if username == current_user['username']:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    result = await db.users.delete_one({"username": username})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"success": True, "message": f"User {username} deleted"}

@api_router.put("/users/{username}")
async def update_user(
    username: str,
    request: Dict,
    current_user: dict = Depends(get_current_user)
):
    """Update user details (Admin only)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can update users")
    
    # Get existing user
    existing_user = await db.users.find_one({"username": username})
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prepare update data
    update_data = {}
    
    # Update full_name if provided
    if 'full_name' in request and request['full_name']:
        update_data['full_name'] = request['full_name']
    
    # Update role if provided
    if 'role' in request and request['role']:
        if request['role'] not in ['admin', 'manager', 'executive', 'polythene_executive']:
            raise HTTPException(status_code=400, detail="Invalid role")
        update_data['role'] = request['role']
    
    # Update active status if provided
    if 'is_active' in request:
        update_data['is_active'] = request['is_active']
    
    # Update password if provided
    if 'password' in request and request['password']:
        update_data['password_hash'] = get_password_hash(request['password'])
    
    # Update new username if provided and different
    if 'new_username' in request and request['new_username'] and request['new_username'] != username:
        # Check if new username already exists
        existing_new = await db.users.find_one({"username": request['new_username']})
        if existing_new:
            raise HTTPException(status_code=400, detail="New username already exists")
        update_data['username'] = request['new_username']
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    # Perform update
    await db.users.update_one(
        {"username": username},
        {"$set": update_data}
    )
    
    return {
        "success": True,
        "message": f"User {username} updated successfully",
        "updated_fields": list(update_data.keys())
    }

@api_router.post("/users/initialize-admin")
async def initialize_admin():
    """Create initial admin user if no users exist (one-time setup)"""
    count = await db.users.count_documents({})
    
    if count > 0:
        raise HTTPException(status_code=400, detail="Users already exist. Use login.")
    
    # Create default admin
    admin = User(
        username="admin",
        password_hash=get_password_hash("admin123"),  # Change on first login!
        full_name="System Administrator",
        role="admin",
        created_by="system"
    )
    
    await db.users.insert_one(admin.model_dump())
    
    return {
        "success": True,
        "message": "Admin user created. Use the credentials you set to login.",
        "username": "admin"
    }

async def batch_insert(collection, documents: list):
    """Insert documents in batches to handle large datasets"""
    total = 0
    for i in range(0, len(documents), BATCH_INSERT_SIZE):
        batch = documents[i:i + BATCH_INSERT_SIZE]
        await collection.insert_many(batch, ordered=False)
        total += len(batch)
    return total


def _prepare_transactions(records: list, batch_id: str) -> list:
    """Prepare transaction dicts with defaults (skip per-row Pydantic for speed)"""
    now = datetime.now(timezone.utc).isoformat()
    docs = []
    for r in records:
        r['batch_id'] = batch_id
        r.setdefault('id', str(uuid.uuid4()))
        r.setdefault('upload_date', now)
        r.setdefault('rate', 0.0)
        r.setdefault('total_amount', 0.0)
        r.setdefault('taxable_value', 0.0)
        docs.append(r)
    return docs


# ==================== CHUNKED UPLOAD (MongoDB-backed for multi-pod deployments) ====================

async def _save_upload_meta(upload_id: str, meta: dict):
    """Persist upload metadata to MongoDB (works across all pods)"""
    await db.upload_sessions.update_one(
        {"upload_id": upload_id},
        {"$set": {**meta, "upload_id": upload_id}},
        upsert=True
    )

async def _load_upload_meta(upload_id: str) -> dict:
    """Load upload metadata from MongoDB"""
    doc = await db.upload_sessions.find_one({"upload_id": upload_id}, {"_id": 0})
    return doc


@api_router.post("/upload/init")
async def init_chunked_upload(request: Dict, current_user: dict = Depends(get_current_user)):
    """Initialize a chunked file upload"""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or Manager only")
    file_type = request.get('file_type')
    if file_type not in ['purchase', 'sale', 'branch_transfer', 'opening_stock', 'physical_stock', 'master_stock', 'historical_sale', 'historical_purchase']:
        raise HTTPException(status_code=400, detail="Invalid file_type")

    # Physical stock must always use the direct upload flow, never chunked
    if file_type == 'physical_stock':
        raise HTTPException(status_code=400, detail="physical_stock uploads must use the direct upload flow")

    upload_id = str(uuid.uuid4())

    meta = {
        'file_type': file_type,
        'start_date': request.get('start_date'),
        'end_date': request.get('end_date'),
        'verification_date': request.get('verification_date'),
        'year': request.get('year'),
        'total_chunks': request.get('total_chunks', 0),
        'received': 0,
        'created_at': datetime.now(timezone.utc).isoformat(),
        'owner_username': current_user['username'],
    }
    await _save_upload_meta(upload_id, meta)

    return {"upload_id": upload_id}


@api_router.post("/upload/chunk/{upload_id}")
async def upload_chunk(upload_id: str, chunk_index: int, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Receive a single chunk of a large file — stored in MongoDB for cross-pod access"""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or Manager only")
    meta = await _load_upload_meta(upload_id)
    if not meta:
        raise HTTPException(status_code=404, detail="Upload session not found")
    if meta.get('owner_username') and meta['owner_username'] != current_user['username']:
        raise HTTPException(status_code=403, detail="Not the owner of this upload session")

    content = await file.read()

    # Store chunk in MongoDB (using Binary for efficient storage)
    from bson import Binary
    await db.upload_chunks.update_one(
        {"upload_id": upload_id, "chunk_index": chunk_index},
        {"$set": {"upload_id": upload_id, "chunk_index": chunk_index, "data": Binary(content)}},
        upsert=True
    )

    meta['received'] = meta.get('received', 0) + 1
    await _save_upload_meta(upload_id, meta)

    return {"received": meta['received'], "chunk_index": chunk_index}


async def _process_upload(upload_id: str, meta: dict):
    """Background task: reassemble chunks to temp file, parse, and insert into DB"""
    import tempfile as _tempfile
    tmp_path = None
    try:
        logger.info(f"[Upload {upload_id}] Starting processing, file_type={meta['file_type']}")
        meta['progress'] = 'Reassembling file from chunks...'
        await _save_upload_meta(upload_id, meta)
        
        # Write chunks directly to a temp file (avoids holding entire file in memory)
        tmp_fd, tmp_path = _tempfile.mkstemp(suffix='.xlsx')
        chunk_count = 0
        total_bytes = 0
        
        cursor = db.upload_chunks.find(
            {"upload_id": upload_id},
            {"chunk_index": 1, "data": 1, "_id": 0}
        ).sort("chunk_index", 1)
        
        import os as _os
        with _os.fdopen(tmp_fd, 'wb') as f:
            async for chunk_doc in cursor:
                data = chunk_doc['data']
                f.write(data)
                total_bytes += len(data)
                chunk_count += 1
        
        logger.info(f"[Upload {upload_id}] Wrote {chunk_count} chunks to disk ({total_bytes} bytes)")
        
        # Delete chunks from MongoDB immediately to free DB space
        await db.upload_chunks.delete_many({"upload_id": upload_id})

        meta['progress'] = f'Parsing Excel file ({total_bytes // 1024} KB)...'
        await _save_upload_meta(upload_id, meta)

        file_type = meta['file_type']

        parse_type = file_type
        if file_type in ('opening_stock', 'master_stock'):
            parse_type = 'opening_stock'
        elif file_type == 'historical_sale':
            parse_type = 'sale'
        elif file_type == 'historical_purchase':
            parse_type = 'purchase'

        # Use streaming parser for large file uploads (avoids loading entire file into memory)
        # This is critical for deployed environments with limited memory
        loop = asyncio.get_event_loop()
        records = await loop.run_in_executor(_parse_executor, parse_excel_streaming, tmp_path, parse_type)

        logger.info(f"[Upload {upload_id}] Parsed {len(records) if records else 0} records")

        if not records:
            meta['status'] = 'error'
            meta['error'] = 'No valid records found in file'
            await _save_upload_meta(upload_id, meta)
            return

        batch_id = str(uuid.uuid4())
        start_date = meta.get('start_date')
        end_date = meta.get('end_date')
        rec_count = len(records)
        meta['progress'] = f'Saving {rec_count} records to database...'
        await _save_upload_meta(upload_id, meta)

        if file_type in ('purchase', 'sale', 'branch_transfer'):
            deleted_count = 0
            # Map file_type to actual stored transaction types for deletion
            if file_type == 'branch_transfer':
                delete_types = ['issue', 'receive']
            else:
                delete_types = [file_type, f"{file_type}_return"]

            # Only delete dates that exist in the NEW file (prevents losing data for dates not in the file)
            new_dates = sorted(set(r.get('date', '') for r in records if r.get('date')))
            if new_dates:
                # Backup replaced records for undo
                old_records = await db.transactions.find(
                    {"type": {"$in": delete_types}, "date": {"$in": new_dates}}, {"_id": 0}
                ).to_list(None)
                if old_records:
                    await db.replaced_records.insert_one({
                        "batch_id": batch_id,
                        "records": old_records,
                        "replaced_at": datetime.now(timezone.utc).isoformat()
                    })
                delete_result = await db.transactions.delete_many({
                    "type": {"$in": delete_types},
                    "date": {"$in": new_dates}
                })
                deleted_count = delete_result.deleted_count

            transactions = _prepare_transactions(records, batch_id)
            await batch_insert(db.transactions, transactions)
            dates_str = f"{new_dates[0]} to {new_dates[-1]}" if new_dates else "unknown"
            message = f"Uploaded {len(transactions)} {file_type} records for {dates_str}"
            if deleted_count > 0:
                message += f" (replaced {deleted_count} old records)"
            await save_action(f'upload_{file_type}', message, {
                'batch_id': batch_id, 'file_type': file_type, 'count': len(transactions)
            })
            await auto_normalize_stamps()
            meta['status'] = 'complete'
            meta['result'] = {"success": True, "count": len(transactions), "replaced_count": deleted_count,
                              "batch_id": batch_id, "message": message}

        elif file_type == 'opening_stock':
            merged_items = {}
            for record in records:
                key = record['item_name'].strip().lower()
                if key not in merged_items:
                    merged_items[key] = {'item_name': record['item_name'], 'stamp': record.get('stamp', ''),
                        'unit': record.get('unit', ''), 'pc': 0, 'gr_wt': 0.0, 'net_wt': 0.0,
                        'fine': 0.0, 'labor_wt': 0.0, 'labor_rs': 0.0, 'rate': record.get('rate', 0.0), 'total': 0.0}
                merged_items[key]['gr_wt'] += record.get('gr_wt', 0)
                merged_items[key]['net_wt'] += record.get('net_wt', 0)
                merged_items[key]['fine'] += record.get('fine', 0)
                merged_items[key]['pc'] += record.get('pc', 0)
                merged_items[key]['total'] += record.get('total', 0)
                if record.get('stamp') and not merged_items[key]['stamp']:
                    merged_items[key]['stamp'] = record['stamp']
            await db.opening_stock.delete_many({})
            stock_items = [OpeningStock(**item).model_dump() for item in merged_items.values()]
            await db.opening_stock.insert_many(stock_items)
            total_net_wt = sum(i['net_wt'] for i in stock_items)
            await save_action('upload_opening_stock', f"Uploaded {len(stock_items)} merged opening stock items")
            await auto_normalize_stamps()
            meta['status'] = 'complete'
            meta['result'] = {"success": True, "count": len(stock_items),
                              "message": f"Opening stock uploaded: {len(stock_items)} items, {total_net_wt/1000:.3f} kg"}

        elif file_type == 'physical_stock':
            verification_date = meta.get('verification_date') or datetime.now(timezone.utc).isoformat()[:10]
            count, message = await _replace_physical_stock_for_date(records, verification_date)
            meta['status'] = 'complete'
            meta['result'] = {"success": True, "count": count,
                              "verification_date": verification_date, "message": message}

        elif file_type in ('historical_sale', 'historical_purchase'):
            year = meta.get('year', '2025')
            for record in records:
                record['batch_id'] = batch_id
                record['historical_year'] = year
                record['is_historical'] = True
            hist_docs = _prepare_transactions(records, batch_id)
            logger.info(f"[Upload {upload_id}] Inserting {len(hist_docs)} historical records into DB...")
            await batch_insert(db.historical_transactions, hist_docs)
            actual_type = 'sale' if file_type == 'historical_sale' else 'purchase'
            # Verify insertion
            verify_count = await db.historical_transactions.count_documents({"batch_id": batch_id})
            logger.info(f"[Upload {upload_id}] Verified {verify_count} records in DB for batch {batch_id}")
            meta['status'] = 'complete'
            meta['result'] = {"success": True, "count": verify_count, "year": year,
                              "message": f"Uploaded {verify_count} historical {actual_type} records for {year}"}
        else:
            meta['status'] = 'error'
            meta['error'] = f"Unsupported file_type: {file_type}"

        await _save_upload_meta(upload_id, meta)

    except Exception as e:
        logger.error(f"[Upload {upload_id}] FAILED: {e}", exc_info=True)
        meta['status'] = 'error'
        meta['error'] = str(e)
        try:
            await _save_upload_meta(upload_id, meta)
        except Exception:
            pass
    finally:
        # Always clean up temp file
        if tmp_path:
            try:
                import os as _os
                _os.unlink(tmp_path)
            except Exception:
                pass


@api_router.post("/upload/finalize/{upload_id}")
async def finalize_chunked_upload(upload_id: str, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    """Reassemble chunks and process the complete file in background"""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or Manager only")
    meta = await _load_upload_meta(upload_id)
    if not meta:
        raise HTTPException(status_code=404, detail="Upload session not found")
    if meta.get('owner_username') and meta['owner_username'] != current_user['username']:
        raise HTTPException(status_code=403, detail="Not the owner of this upload session")

    # Verify ALL chunks are present before processing
    chunk_count = await db.upload_chunks.count_documents({"upload_id": upload_id})
    expected = meta.get('total_chunks', 0)
    if chunk_count == 0:
        raise HTTPException(status_code=400, detail="No chunks received")
    if expected > 0 and chunk_count < expected:
        # Find which chunks are missing
        received = set()
        async for doc in db.upload_chunks.find({"upload_id": upload_id}, {"chunk_index": 1, "_id": 0}):
            received.add(doc["chunk_index"])
        missing = [i for i in range(expected) if i not in received]
        raise HTTPException(status_code=400, detail=f"Missing {len(missing)} of {expected} chunks: {missing[:10]}")

    # Mark as processing and kick off background task
    meta['status'] = 'processing'
    await _save_upload_meta(upload_id, meta)

    background_tasks.add_task(_process_upload, upload_id, meta)

    return {"status": "processing", "upload_id": upload_id,
            "message": "File is being processed. Poll /api/upload/status/{upload_id} for progress."}


@api_router.get("/upload/status/{upload_id}")
async def get_upload_status(upload_id: str, current_user: dict = Depends(get_current_user)):
    """Poll processing status of a chunked upload"""
    meta = await _load_upload_meta(upload_id)
    if not meta:
        raise HTTPException(status_code=404, detail="Upload session not found")
    if meta.get('owner_username') and meta['owner_username'] != current_user['username']:
        raise HTTPException(status_code=403, detail="Not the owner of this upload session")

    status = meta.get('status', 'unknown')

    if status == 'complete':
        result = meta.get('result', {})
        # Clean up session from MongoDB
        await db.upload_sessions.delete_one({"upload_id": upload_id})
        return {"status": "complete", **result}
    elif status == 'error':
        error_detail = meta.get('error', 'Unknown error')
        await db.upload_sessions.delete_one({"upload_id": upload_id})
        return {"status": "error", "detail": error_detail}
    else:
        progress = meta.get('progress', 'Processing...')
        return {"status": "processing", "message": progress}


# ==================== CLIENT-SIDE PARSED UPLOAD (no server-side Excel parsing — OOM-safe) ====================

def _parse_raw_rows(raw_rows: List[Dict], cols: set, file_type: str) -> List[Dict]:
    """Convert pre-parsed row dicts (from SheetJS in browser) to transaction records.
    Reuses the same column-mapping logic as parse_excel_file but without pandas/openpyxl."""
    KG_TO_GRAMS = 1000
    records = []

    if file_type == 'purchase':
        item_col = _resolve_col(cols, ['Item Name', 'Particular', 'item name'])
        type_col = _resolve_col(cols, ['Type', 'type'])
        tag_col = _resolve_col(cols, ['Tag.No.', 'Tag No', 'tag no'])
        wt_rs_col = _resolve_col(cols, ['Wt/Rs', 'Wt Rs'])
        total_col = _resolve_col(cols, ['Total', 'total'])
        tunch_col = _resolve_col(cols, ['Tunch', 'tunch'])
        wstg_col = _resolve_col(cols, ['Wstg', 'wstg'])
        date_col = _resolve_col(cols, ['Date', 'date'])
        refno_col = _resolve_col(cols, ['Refno', 'refno', 'Ref No'])
        party_col = _resolve_col(cols, ['Party Name', 'party name', 'Party'])
        stamp_col = _resolve_col(cols, ['Stamp', 'stamp'])
        gr_col = _resolve_col(cols, ['Gr.Wt.', 'Gr Wt', 'Gross Wt'])
        net_col = _resolve_col(cols, ['Net.Wt.', 'Net Wt'])
        fine_col = _resolve_col(cols, ['Fine', 'Sil.Fine', 'Sil Fine', 'Silver Fine'])
        dia_col = _resolve_col(cols, ['Dia.Wt.', 'Dia Wt'])
        stn_col = _resolve_col(cols, ['Stn.Wt.', 'Stn Wt'])
        rate_col = _resolve_col(cols, ['Rate', 'rate'])
        pc_col = _resolve_col(cols, ['Pc', 'pc', 'Pieces'])

        for r in raw_rows:
            item_name = _safe_str(r.get(item_col) if item_col else None)
            if len(item_name) < 2:
                continue
            trans_type = _safe_str(r.get(type_col) if type_col else None, 'P').upper()
            if trans_type.isdigit():
                continue
            tag_no = _safe_str(r.get(tag_col) if tag_col else None)
            labor_val, labor_on = parse_labor_value(tag_no)
            wt_rs = r.get(wt_rs_col) if wt_rs_col else None
            if wt_rs and str(wt_rs).replace('.', '').isdigit():
                labor_val = float(wt_rs)
            total_amount = _safe_float(r.get(total_col) if total_col else None)
            tunch_v = _safe_float(r.get(tunch_col) if tunch_col else None)
            wstg_v = _safe_float(r.get(wstg_col) if wstg_col else None)
            purchase_tunch = tunch_v + wstg_v
            records.append({
                'date': normalize_date(r.get(date_col) if date_col else ''),
                'type': 'purchase' if trans_type in ('P', 'PURCHASE') else 'purchase_return',
                'refno': _safe_str(r.get(refno_col) if refno_col else None),
                'party_name': _safe_str(r.get(party_col) if party_col else None),
                'item_name': item_name,
                'stamp': normalize_stamp(r.get(stamp_col) if stamp_col else ''),
                'tag_no': tag_no,
                'gr_wt': _safe_float(r.get(gr_col) if gr_col else None) * KG_TO_GRAMS,
                'net_wt': _safe_float(r.get(net_col) if net_col else None) * KG_TO_GRAMS,
                'fine': _safe_float(r.get(fine_col) if fine_col else None) * KG_TO_GRAMS,
                'labor': labor_val,
                'labor_on': labor_on,
                'dia_wt': _safe_float(r.get(dia_col) if dia_col else None) * KG_TO_GRAMS,
                'stn_wt': _safe_float(r.get(stn_col) if stn_col else None) * KG_TO_GRAMS,
                'tunch': str(purchase_tunch),
                'rate': _safe_float(r.get(rate_col) if rate_col else None),
                'total_pc': _safe_int(r.get(pc_col) if pc_col else None),
                'total_amount': total_amount,
            })

    elif file_type == 'sale':
        item_col = _resolve_col(cols, ['Item Name', 'Particular', 'item name'])
        type_col = _resolve_col(cols, ['Type', 'type'])
        tag_col = _resolve_col(cols, ['Lbr. On Tag.No.', 'Tag.No.', 'Tag No'])
        on_col = _resolve_col(cols, ['On', 'on'])
        total_col = _resolve_col(cols, ['Total', 'total'])
        tunch_col = _resolve_col(cols, ['Tunch', 'tunch'])
        date_col = _resolve_col(cols, ['Date', 'date'])
        refno_col = _resolve_col(cols, ['Refno', 'refno', 'Ref No'])
        party_col = _resolve_col(cols, ['Party Name', 'party name', 'Party'])
        stamp_col = _resolve_col(cols, ['Stamp', 'stamp'])
        gr_col = _resolve_col(cols, ['Gr.Wt.', 'Gr Wt', 'Gross Wt'])
        net_col = _resolve_col(cols, ['Gold Std.', 'Net.Wt.', 'Net Wt'])
        fine_col = _resolve_col(cols, ['Fine', 'Sil.Fine', 'Sil Fine'])
        dia_col = _resolve_col(cols, ['Dia.Wt.', 'Dia Wt'])
        stn_col = _resolve_col(cols, ['Stn.Wt.', 'Stn Wt'])
        taxable_col = _resolve_col(cols, ['Taxable Val.', 'Taxable Value'])
        pc_col = _resolve_col(cols, ['Pc', 'pc'])

        for r in raw_rows:
            item_name = _safe_str(r.get(item_col) if item_col else None)
            if len(item_name) < 2:
                continue
            trans_type = _safe_str(r.get(type_col) if type_col else None, 'S').upper()
            if trans_type.isdigit():
                continue
            tag_no = _safe_str(r.get(tag_col) if tag_col else None)
            labor_val, labor_on = parse_labor_value(tag_no)
            on_val = r.get(on_col) if on_col else None
            if on_val and str(on_val).replace('.', '').isdigit():
                labor_val = float(on_val)
            total_amount = _safe_float(r.get(total_col) if total_col else None)
            sale_tunch = _safe_float(r.get(tunch_col) if tunch_col else None)
            records.append({
                'type': 'sale' if trans_type in ('S', 'SALE') else 'sale_return',
                'date': normalize_date(r.get(date_col) if date_col else ''),
                'refno': _safe_str(r.get(refno_col) if refno_col else None),
                'party_name': _safe_str(r.get(party_col) if party_col else None),
                'item_name': item_name,
                'stamp': normalize_stamp(r.get(stamp_col) if stamp_col else ''),
                'tag_no': tag_no,
                'gr_wt': _safe_float(r.get(gr_col) if gr_col else None) * KG_TO_GRAMS,
                'net_wt': _safe_float(r.get(net_col) if net_col else None) * KG_TO_GRAMS,
                'fine': _safe_float(r.get(fine_col) if fine_col else None) * KG_TO_GRAMS,
                'labor': labor_val,
                'labor_on': labor_on,
                'dia_wt': _safe_float(r.get(dia_col) if dia_col else None) * KG_TO_GRAMS,
                'stn_wt': _safe_float(r.get(stn_col) if stn_col else None) * KG_TO_GRAMS,
                'tunch': str(sale_tunch),
                'total_amount': total_amount,
                'taxable_value': _safe_float(r.get(taxable_col) if taxable_col else None),
                'total_pc': _safe_int(r.get(pc_col) if pc_col else None),
            })
    return records


@api_router.post("/upload/client-batch")
async def client_batch_upload(request: Dict, current_user: dict = Depends(get_current_user)):
    """Accept a batch of pre-parsed rows from client-side Excel reading.
    No file upload, no Excel parsing on server — completely OOM-safe."""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    file_type = request.get('file_type')
    if not file_type:
        raise HTTPException(status_code=400, detail="file_type is required")

    batch_id = request.get('batch_id')
    if not batch_id:
        raise HTTPException(status_code=400, detail="batch_id is required")

    headers = request.get('headers', [])
    rows = request.get('rows', [])
    year = request.get('year', '2025')
    is_final = request.get('is_final', False)
    batch_index = request.get('batch_index', 0)

    if not rows:
        if is_final:
            # Final call with no rows — just return totals
            total = await db.historical_transactions.count_documents({"batch_id": batch_id})
            return {"success": True, "batch_records": 0, "total_so_far": total, "message": f"Upload complete. {total} records total."}
        raise HTTPException(status_code=400, detail="No rows in batch")

    # Convert rows (arrays) to dicts using headers
    raw_rows = []
    for row in rows:
        d = {}
        for i, h in enumerate(headers):
            d[h] = str(row[i]).strip() if i < len(row) and row[i] is not None else ''
        raw_rows.append(d)

    # Determine parse type
    parse_type = file_type
    if file_type in ('historical_sale',):
        parse_type = 'sale'
    elif file_type in ('historical_purchase',):
        parse_type = 'purchase'

    # Apply column mapping
    cols = set(headers)
    records = _parse_raw_rows(raw_rows, cols, parse_type)

    if not records:
        return {"success": True, "batch_records": 0, "total_so_far": 0, "message": "No valid records in this batch"}

    # Determine target collection
    is_historical = file_type.startswith('historical_')
    collection = db.historical_transactions if is_historical else db.transactions

    # Prepare and insert
    for rec in records:
        rec['batch_id'] = batch_id
        if is_historical:
            rec['historical_year'] = year
            rec['is_historical'] = True

    docs = _prepare_transactions(records, batch_id)
    await batch_insert(collection, docs)
    inserted_count = len(docs)
    del raw_rows, records, docs
    gc.collect()

    total = await collection.count_documents({"batch_id": batch_id})
    logger.info(f"[Client batch] batch_index={batch_index}, inserted={inserted_count}, total_so_far={total}")

    result = {"success": True, "batch_records": inserted_count, "total_so_far": total}
    if is_final:
        actual_type = parse_type
        result["message"] = f"Uploaded {total} historical {actual_type} records for {year}"
    return result


@api_router.post("/transactions/upload/{file_type}")
async def upload_transaction_file(
    file_type: str, 
    file: UploadFile = File(...),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Upload purchase, sale, or branch_transfer Excel file"""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or Manager only")
    if file_type not in ['purchase', 'sale', 'branch_transfer']:
        raise HTTPException(status_code=400, detail="file_type must be 'purchase', 'sale', or 'branch_transfer'")
    
    content = await file.read()
    
    # Parse in thread pool so we don't block the event loop
    loop = asyncio.get_event_loop()
    records = await loop.run_in_executor(_parse_executor, parse_excel_file, content, file_type)
    
    if not records:
        raise HTTPException(status_code=400, detail="No valid records found in file")
    
    batch_id = str(uuid.uuid4())
    
    # DELETE only dates that exist in the new file (prevents losing data for dates not in the file)
    deleted_count = 0
    if file_type == 'branch_transfer':
        delete_types = ['issue', 'receive']
    else:
        delete_types = [file_type, f"{file_type}_return"]

    new_dates = sorted(set(r.get('date', '') for r in records if r.get('date')))
    if new_dates:
        # Backup replaced records for undo
        old_records = await db.transactions.find(
            {"type": {"$in": delete_types}, "date": {"$in": new_dates}}, {"_id": 0}
        ).to_list(None)
        if old_records:
            await db.replaced_records.insert_one({
                "batch_id": batch_id,
                "records": old_records,
                "replaced_at": datetime.now(timezone.utc).isoformat()
            })
        delete_result = await db.transactions.delete_many({
            "type": {"$in": delete_types},
            "date": {"$in": new_dates}
        })
        deleted_count = delete_result.deleted_count
    
    # Prepare and batch-insert
    transactions = _prepare_transactions(records, batch_id)
    await batch_insert(db.transactions, transactions)
    
    dates_str = f"{new_dates[0]} to {new_dates[-1]}" if new_dates else "unknown"
    message = f"Uploaded {len(transactions)} {file_type} records for {dates_str}"
    if deleted_count > 0:
        message += f" (replaced {deleted_count} old records)"
    
    await save_action(
        f'upload_{file_type}',
        message,
        {
            'batch_id': batch_id,
            'file_name': file.filename,
            'file_type': file_type,
            'count': len(transactions),
            'start_date': start_date,
            'end_date': end_date
        }
    )
    
    # Auto-normalize stamps after upload
    await auto_normalize_stamps()
    
    return {
        "success": True,
        "count": len(transactions),
        "replaced_count": deleted_count,
        "batch_id": batch_id,
        "message": message
    }

@api_router.get("/executive/my-entries/{username}")
async def get_executive_entries(username: str, current_user: dict = Depends(get_current_user)):
    """Get stock entries by an executive — latest per stamp shown first"""
    # Only allow viewing own entries, or admin/manager can view any
    if current_user['username'] != username and current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Can only view your own entries")
    entries = await db.stock_entries.find(
        {'entered_by': username},
        {"_id": 0}
    ).sort('entry_date', -1).to_list(500)
    
    # Return latest entry per stamp (deduplicate by stamp, keep most recent)
    seen_stamps = set()
    latest_entries = []
    for e in entries:
        stamp = e.get('stamp', '')
        if stamp not in seen_stamps:
            seen_stamps.add(stamp)
            latest_entries.append(e)
    return latest_entries

@api_router.put("/executive/update-entry/{stamp}")
async def update_stock_entry(
    stamp: str,
    request: Dict,
    current_user: dict = Depends(get_current_user)
):
    """Update a rejected stock entry (same day only)"""
    entries = request.get('entries', [])
    verification_date = request.get('verification_date')
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    update_fields = {
        'entries': entries,
        'entry_date': datetime.now(timezone.utc).isoformat(),
        'status': 'pending'
    }
    if verification_date:
        update_fields['verification_date'] = verification_date
    
    # Try to update today's entry first, fallback to latest pending/rejected
    result = await db.stock_entries.update_one(
        {'stamp': stamp, 'entered_by': current_user['username'], 'entry_day': today, 'status': {'$in': ['pending', 'rejected']}},
        {'$set': update_fields}
    )
    
    if result.modified_count == 0:
        # Fallback: update the latest pending/rejected entry for this stamp
        fallback_fields = {
            'entries': entries,
            'entry_date': datetime.now(timezone.utc).isoformat(),
            'entry_day': today,
            'status': 'pending'
        }
        if verification_date:
            fallback_fields['verification_date'] = verification_date
        await db.stock_entries.update_one(
            {'stamp': stamp, 'entered_by': current_user['username'], 'status': {'$in': ['pending', 'rejected']}},
            {'$set': fallback_fields},
            upsert=False
        )
    
    return {'success': True, 'message': 'Entry updated'}

@api_router.get("/manager/all-entries")
async def get_all_entries(current_user: dict = Depends(get_current_user)):
    """Get all stock entries for manager — latest per stamp, sorted by entry_date desc"""
    if current_user['role'] not in ['manager', 'admin']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    entries = await db.stock_entries.find({}, {"_id": 0}).sort('entry_date', -1).to_list(1000)
    return entries

@api_router.delete("/executive/delete-entry/{stamp}/{username}")
async def delete_executive_entry(stamp: str, username: str, current_user: dict = Depends(get_current_user)):
    """Delete the latest stock entry for a stamp"""
    if current_user['username'] != username and current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Can only delete your own entries")
    
    # Delete the most recent entry for this stamp by this user
    latest = await db.stock_entries.find_one(
        {'stamp': stamp, 'entered_by': username},
        sort=[('entry_date', -1)]
    )
    if not latest:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    await db.stock_entries.delete_one({'_id': latest['_id']})
    
    return {'success': True, 'message': 'Entry deleted'}

# ==================== EXECUTIVE ENDPOINTS ====================

@api_router.post("/executive/stock-entry")
async def save_executive_stock_entry(
    request: Dict,
    current_user: dict = Depends(get_current_user)
):
    """Save stock entry from executive (for manager approval).
    Rules:
    - Same stamp + same day = update existing (overwrite)
    - Different day = new entry (previous day's entry is locked/historical)
    - Approved entries from previous days remain untouched
    - Each stamp shown by its last submission timestamp
    """
    
    if current_user['role'] not in ['executive', 'manager', 'admin']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    stamp = request.get('stamp')
    entries = request.get('entries', [])
    entered_by = current_user['username']  # Always use authenticated user, never trust client
    verification_date = request.get('verification_date')  # Date for which stock is being entered
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    if not verification_date:
        # Check if there's a rejected entry for this stamp — inherit its verification_date
        rejected = await db.stock_entries.find_one(
            {'stamp': stamp, 'entered_by': entered_by, 'status': 'rejected'},
            {"_id": 0},
            sort=[('entry_date', -1)]
        )
        if rejected and rejected.get('verification_date'):
            verification_date = rejected['verification_date']
        else:
            verification_date = today
    
    # Save stock entry keyed by stamp + user + today's date
    entry_record = {
        'stamp': stamp,
        'entries': entries,
        'entered_by': entered_by,
        'entry_date': datetime.now(timezone.utc).isoformat(),
        'entry_day': today,
        'verification_date': verification_date,
        'status': 'pending',
        'approved_by': None,
        'approved_at': None,
    }
    
    # Check if an entry exists for this stamp + user + today
    existing_today = await db.stock_entries.find_one({
        'stamp': stamp, 'entered_by': entered_by, 'entry_day': today
    })
    
    if existing_today:
        # Same day — overwrite (update values, reset status to pending)
        entry_record['iteration'] = existing_today.get('iteration', 0) + 1
        await db.stock_entries.update_one(
            {'stamp': stamp, 'entered_by': entered_by, 'entry_day': today},
            {'$set': entry_record}
        )
    else:
        # New day — insert new entry (old entries remain as history)
        entry_record['iteration'] = 1
        await db.stock_entries.insert_one(entry_record)
    
    # If there was a same-day approval in stamp_approvals, clear it so manager can re-approve
    await db.stamp_approvals.delete_one({'stamp': stamp, 'approval_day': today})
    
    # Create notification for manager
    await db.notifications.insert_one({
        'id': str(uuid.uuid4()),
        'category': 'stamp',
        'type': 'stock_entry',
        'message': f'{entered_by} submitted stock for {stamp}',
        'severity': 'info',
        'target_user': 'manager',
        'stamp': stamp,
        'timestamp': datetime.now(timezone.utc).isoformat(),
        'read': False
    })
    
    # Log to activity log for accountability
    await db.activity_log.insert_one({
        'user': entered_by,
        'user_role': current_user['role'],
        'action_type': 'stock_entry',
        'description': f'Submitted stock for {stamp} ({len(entries)} items)',
        'details': {
            'stamp': stamp,
            'items_count': len(entries),
            'iteration': entry_record['iteration']
        },
        'timestamp': datetime.now(timezone.utc).isoformat()
    })
    
    return {'success': True, 'message': 'Stock entry saved successfully'}

@api_router.get("/manager/approval-details/{stamp}")
async def get_approval_details(stamp: str, verification_date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get detailed approval data — compares entered stock vs expected closing stock
    for the verification_date. Auto-refreshes when transactions change.
    If verification_date is provided, looks up that specific entry instead of the latest."""
    
    if current_user['role'] not in ['manager', 'admin']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # If verification_date provided, find the SPECIFIC entry for that date
    if verification_date:
        entry = await db.stock_entries.find_one(
            {'stamp': stamp, 'verification_date': verification_date, 'status': {'$in': ['pending', 'approved']}},
            {"_id": 0},
            sort=[('entry_date', -1)]
        )
    else:
        entry = None
    
    # Fallback: get latest pending or approved entry for this stamp
    if not entry:
        entry = await db.stock_entries.find_one(
            {'stamp': stamp, 'status': {'$in': ['pending', 'approved']}},
            {"_id": 0},
            sort=[('entry_date', -1)]
        )
    
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    verification_date = entry.get('verification_date', entry.get('entry_day', datetime.now(timezone.utc).strftime('%Y-%m-%d')))
    
    # Get ALL items in this stamp from master
    master_items = await db.master_items.find({'stamp': stamp}, {"_id": 0}).to_list(1000)
    master_item_names = {m['item_name'] for m in master_items}
    
    # Create map of entered weights
    entered_map = {}
    for entered in entry.get('entries', []):
        entered_map[entered['item_name']] = entered['gross_wt']
    
    # Calculate expected closing stock for the verification_date
    # Use get_current_inventory (baseline-aware) and extract stamp items
    current_inv = await get_current_inventory(as_of_date=verification_date)
    stamp_items_list = current_inv.get('by_stamp', {}).get(stamp, [])
    closing_stock = {}
    for si in stamp_items_list:
        closing_stock[si['item_name']] = round(si['gr_wt'] / 1000, 3)
    
    # Build comparison for ALL items in stamp
    comparison = []
    total_entered = 0.0
    total_book = 0.0
    
    all_item_names = sorted(master_item_names | set(closing_stock.keys()))
    
    for item_name in all_item_names:
        entered_gross = entered_map.get(item_name, 0.0)
        book_gross = closing_stock.get(item_name, 0.0)
        difference = round(entered_gross - book_gross, 3)
        
        comparison.append({
            'item_name': item_name,
            'entered_gross': entered_gross,
            'book_gross': book_gross,
            'difference': difference,
            'was_entered': item_name in entered_map,
            'is_mapped': item_name not in master_item_names
        })
        
        total_entered += entered_gross
        total_book += book_gross
    
    return {
        'entry': entry,
        'verification_date': verification_date,
        'comparison': comparison,
        'total_items': len(comparison),
        'items_entered': len(entered_map),
        'total_entered': round(total_entered, 3),
        'total_book': round(total_book, 3),
        'total_difference': round(total_entered - total_book, 3)
    }


# ==================== MANAGER ENDPOINTS ====================

@api_router.get("/manager/pending-approvals")
async def get_pending_approvals(current_user: dict = Depends(get_current_user)):
    """Get all pending stock entries for manager approval"""
    
    if current_user['role'] not in ['manager', 'admin']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    entries = await db.stock_entries.find({'status': 'pending'}, {"_id": 0}).to_list(100)
    return entries

@api_router.get("/polythene/all")
async def get_all_polythene_adjustments(current_user: dict = Depends(get_current_user)):
    """Get ALL polythene adjustments from all time (admin and executive)"""
    if current_user['role'] not in ['admin', 'executive']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    entries = await db.polythene_adjustments.find({}, {"_id": 0}).sort('created_at', -1).to_list(None)
    return entries

@api_router.get("/polythene/item/{item_name}")
async def get_item_polythene_history(item_name: str, current_user: dict = Depends(get_current_user)):
    """Get all polythene adjustments for a specific item"""
    
    entries = await db.polythene_adjustments.find(
        {'item_name': item_name},
        {"_id": 0}
    ).sort('created_at', -1).to_list(1000)
    
    return entries


@api_router.put("/manager/update-verification-date/{stamp}")
async def update_verification_date(stamp: str, request: Dict, current_user: dict = Depends(get_current_user)):
    """Admin/Manager can update verification_date on any entry to recalculate book values"""
    if current_user['role'] not in ['manager', 'admin']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    new_date = request.get('verification_date')
    if not new_date:
        raise HTTPException(status_code=400, detail="verification_date required")
    
    # Find latest entry for this stamp
    entry = await db.stock_entries.find_one(
        {'stamp': stamp, 'status': {'$in': ['pending', 'approved']}},
        sort=[('entry_date', -1)]
    )
    
    if entry:
        await db.stock_entries.update_one(
            {'_id': entry['_id']},
            {'$set': {'verification_date': new_date}}
        )
    else:
        # Try any entry for this stamp
        await db.stock_entries.update_many(
            {'stamp': stamp},
            {'$set': {'verification_date': new_date}}
        )
    
    return {'success': True, 'message': f'Verification date updated to {new_date}'}

@api_router.post("/manager/approve-stamp")
async def approve_stamp(
    request: Dict,
    current_user: dict = Depends(get_current_user)
):
    """Approve or reject a stamp's stock entry.
    Only affects the latest pending/approved entry for this stamp.
    Old approved entries from previous days remain untouched.
    """
    
    if current_user['role'] not in ['manager', 'admin']:
        raise HTTPException(status_code=403, detail="Only managers can approve")
    
    stamp = request.get('stamp')
    approve = request.get('approve')
    total_difference = request.get('total_difference', 0)
    
    # Get the LATEST pending or approved entry for this stamp
    entry = await db.stock_entries.find_one(
        {'stamp': stamp, 'status': {'$in': ['pending', 'approved']}},
        sort=[('entry_date', -1)]
    )
    
    if not entry:
        raise HTTPException(status_code=404, detail="No pending entry found for this stamp")
    
    iteration = entry.get('iteration', 1)
    entry_day = entry.get('entry_day', datetime.now(timezone.utc).strftime('%Y-%m-%d'))
    now_iso = datetime.now(timezone.utc).isoformat()
    
    # Update ONLY this specific entry (by _id to be precise)
    await db.stock_entries.update_one(
        {'_id': entry['_id']},
        {'$set': {
            'status': 'approved' if approve else 'rejected',
            'approved_by': current_user['username'],
            'approved_at': now_iso,
            'rejection_message': request.get('rejection_message', '') if not approve else None
        }}
    )
    
    # If approving, record approval (keyed by stamp + day so previous approvals remain)
    if approve:
        await db.stamp_approvals.update_one(
            {'stamp': stamp, 'approval_day': entry_day},
            {'$set': {
                'stamp': stamp,
                'is_approved': True,
                'approved_by': current_user['username'],
                'approved_at': now_iso,
                'approval_day': entry_day,
                'iterations': iteration,
                'total_difference': total_difference
            }},
            upsert=True
        )
        
        # Also write to stamp_verifications so dashboard sees this as verified
        # Use the entry's verification_date (the date stock was counted FOR), not today
        entry_verification_date = entry.get('verification_date', entry_day)
        diff_kg = total_difference / 1000 if total_difference else 0
        is_match = abs(total_difference) <= 50
        await db.stamp_verifications.update_one(
            {'stamp': stamp, 'verification_date': entry_verification_date},
            {'$set': {
                'stamp': stamp,
                'physical_gross_wt': 0,
                'book_gross_wt': 0,
                'difference': diff_kg,
                'is_match': is_match,
                'verification_date': entry_verification_date,
                'verified_at': now_iso,
                'approved_by': current_user['username']
            }},
            upsert=True
        )
    else:
        # Rejecting — remove today's approval lock only (not historical)
        await db.stamp_approvals.delete_one({'stamp': stamp, 'approval_day': entry_day})
    
    # Notify admin
    is_matching = abs(total_difference) <= 50
    
    if approve:
        if is_matching:
            notification_message = f'{current_user["username"]} approved {stamp} - ✓ MATCHING (Diff: {total_difference/1000:.3f}kg)'
        else:
            notification_message = f'{current_user["username"]} approved {stamp} - ⚠️ NOT MATCHING (Diff: {total_difference/1000:.3f}kg) - APPROVED DESPITE MISMATCH'
    else:
        rejection_msg = request.get('rejection_message', '')
        notification_message = f'{current_user["username"]} rejected {stamp} (Diff: {total_difference/1000:.3f}kg)'
        if rejection_msg:
            notification_message += f' - Message: "{rejection_msg}"'
    
    await db.notifications.insert_one({
        'id': str(uuid.uuid4()),
        'category': 'stamp',
        'type': 'stamp_approval',
        'message': notification_message,
        'severity': 'success' if (approve and is_matching) else 'warning',
        'target_user': entry.get('entered_by', 'admin'),
        'stamp': stamp,
        'details': {
            'approved_by': current_user['username'],
            'iterations': iteration,
            'total_difference_kg': total_difference / 1000,
            'is_matching': is_matching,
            'entered_by': entry.get('entered_by') if entry else 'unknown',
            'action': 'approved' if approve else 'rejected',
            'approved_despite_mismatch': approve and not is_matching,
            'rejection_message': request.get('rejection_message') if not approve else None
        },
        'timestamp': datetime.now(timezone.utc).isoformat(),
        'read': False
    })

    # Also notify admin if the executive's entry was approved/rejected
    if entry.get('entered_by') != 'admin':
        await db.notifications.insert_one({
            'id': str(uuid.uuid4()),
            'category': 'stamp',
            'type': 'stamp_approval',
            'message': notification_message,
            'severity': 'success' if (approve and is_matching) else 'warning',
            'target_user': 'admin',
            'stamp': stamp,
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'read': False
        })
    
    await db.activity_log.insert_one({
        'user': current_user['username'],
        'user_role': current_user['role'],
        'action_type': 'stamp_approval' if approve else 'stamp_rejection',
        'description': f'{"Approved" if approve else "Rejected"} {stamp} by {entry.get("entered_by") if entry else "unknown"} - Diff: {total_difference/1000:.3f}kg',
        'details': {
            'stamp': stamp,
            'action': 'approved' if approve else 'rejected',
            'total_difference_kg': total_difference / 1000,
            'iterations': iteration,
            'entered_by': entry.get('entered_by') if entry else 'unknown'
        },
        'timestamp': datetime.now(timezone.utc).isoformat()
    })
    
    _inv_cache.invalidate()
    return {'success': True, 'message': f'{stamp} {"approved" if approve else "rejected"}', 'iterations': iteration}

@api_router.get("/stamp-verification/history")
async def get_stamp_verification_history(current_user: dict = Depends(get_current_user)):
    """Get verification history for all stamps"""
    
    # Get stamps from ALL relevant sources, not just master_items
    stamps_set = set()
    for coll_name in ['master_items', 'opening_stock', 'physical_stock', 'transactions']:
        coll = db[coll_name]
        stamps = await coll.distinct('stamp')
        stamps_set.update(s for s in stamps if s and s != 'Unassigned')
    # Also include stamp_assignments stamps
    sa_stamps = await db.stamp_assignments.distinct('stamp')
    stamps_set.update(s for s in sa_stamps if s and s != 'Unassigned')
    
    all_stamps = sorted(stamps_set, key=stamp_sort_key)
    
    # Get latest verification for each
    history = []
    for stamp in all_stamps:
        
        # Check stamp_verifications first (physical verification)
        latest = await db.stamp_verifications.find_one(
            {'stamp': stamp},
            {"_id": 0},
            sort=[('verified_at', -1)]
        )
        
        # Also check stamp_approvals (manager approval = verification) — get latest
        approval = await db.stamp_approvals.find_one(
            {'stamp': stamp, 'is_approved': True},
            {"_id": 0},
            sort=[('approved_at', -1)]
        )
        
        # Use whichever is more recent
        verified_date = None
        is_match = None
        difference = None
        
        if latest:
            verified_date = latest.get('verification_date')
            is_match = latest.get('is_match')
            raw_diff = latest.get('difference', 0)
            # Normalize difference to kg (could be stored as grams or kg)
            difference = round(raw_diff / 1000, 3) if abs(raw_diff) > 100 else round(raw_diff, 3)
        
        if approval:
            approval_date = approval.get('approved_at', '')
            latest_date = latest.get('verified_at', '') if latest else ''
            if approval_date > latest_date:
                verified_date = approval_date[:10] if approval_date else None
                diff_kg = approval.get('total_difference', 0) / 1000 if approval.get('total_difference') else 0
                is_match = abs(diff_kg) < 0.05
                difference = round(diff_kg, 3)
        
        history.append({
            'stamp': stamp,
            'last_verified_date': verified_date,
            'verified_by': approval.get('approved_by') if approval else (latest.get('verified_at') if latest else None),
            'is_match': is_match,
            'difference': difference
        })
    
    return history

@api_router.get("/notifications/my")
async def get_my_notifications(current_user: dict = Depends(get_current_user)):
    """Get notifications for current user — matches by target_user or role"""
    
    notifications = await db.notifications.find(
        {'$or': [
            {'target_user': current_user['username']},
            {'target_user': current_user['role']},
            {'target_user': 'all'},
            {'for_role': {'$in': [current_user['role'], 'all']}}
        ]},
        {"_id": 0}
    ).sort('timestamp', -1).limit(50).to_list(50)
    
    return notifications

@api_router.post("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(get_current_user)):
    """Mark notification as read (matches same scope as fetch query)"""
    await db.notifications.update_one(
        {'id': notification_id, '$or': [
            {'target_user': current_user['username']},
            {'target_user': current_user['role']},
            {'target_user': 'all'},
            {'for_role': {'$in': [current_user['role'], 'all']}}
        ]},
        {'$set': {'read': True}}
    )


# ==================== POLYTHENE EXECUTIVE ENDPOINTS ====================

@api_router.post("/polythene/adjust")
async def adjust_polythene(
    item_name: str,
    poly_weight: float,
    operation: str,
    current_user: dict = Depends(get_current_user)
):
    """Adjust polythene weight for an item (gross weight changes, net stays same)"""
    
    if current_user['role'] not in ['polythene_executive', 'admin']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    actual_user = current_user['username']  # Always use authenticated user
    
    # Dedup: reject if identical entry by same user within 20 seconds
    cutoff = (datetime.now(timezone.utc) - timedelta(seconds=20)).isoformat()
    existing = await db.polythene_adjustments.find_one({
        'item_name': item_name,
        'poly_weight': poly_weight,
        'operation': operation,
        'adjusted_by': actual_user,
        'created_at': {'$gte': cutoff}
    })
    if existing:
        return {'success': True, 'message': 'Duplicate skipped — identical entry within last 20s', 'duplicate': True}
    
    # Save polythene adjustment
    adjustment = {
        'id': str(uuid.uuid4()),
        'item_name': item_name,
        'poly_weight': poly_weight,
        'operation': operation,
        'adjusted_by': actual_user,
        'created_at': datetime.now(timezone.utc).isoformat(),
        'date': datetime.now(timezone.utc).date().isoformat()
    }
    
    await db.polythene_adjustments.insert_one(adjustment)
    
    await db.notifications.insert_one({
        'id': str(uuid.uuid4()),
        'category': 'polythene',
        'type': 'polythene_adjustment',
        'message': f'{actual_user} {operation}ed {poly_weight} kg polythene for {item_name}',
        'severity': 'info',
        'target_user': 'admin',
        'item_name': item_name,
        'timestamp': datetime.now(timezone.utc).isoformat(),
        'read': False
    })

    await db.activity_log.insert_one({
        'user': actual_user,
        'user_role': current_user['role'],
        'action_type': 'polythene_adjustment',
        'description': f'{operation.upper()} {poly_weight} kg polythene for {item_name}',
        'details': {'item': item_name, 'weight': poly_weight, 'operation': operation},
        'timestamp': datetime.now(timezone.utc).isoformat()
    })
    
    return {'success': True, 'message': 'Polythene adjustment saved'}

@api_router.post("/polythene/adjust-batch")
async def adjust_polythene_batch(
    request: Dict,
    current_user: dict = Depends(get_current_user)
):
    """Save multiple polythene adjustments at once"""
    
    if current_user['role'] not in ['polythene_executive', 'admin']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    entries = request.get('entries', [])
    actual_user = current_user['username']  # Always use authenticated user
    
    # Dedup: check for identical entries by same user within 20 seconds
    cutoff = (datetime.now(timezone.utc) - timedelta(seconds=20)).isoformat()
    
    saved_entries = []
    skipped = 0
    
    for entry in entries:
        existing = await db.polythene_adjustments.find_one({
            'item_name': entry['item_name'],
            'poly_weight': entry['poly_weight'],
            'operation': entry['operation'],
            'adjusted_by': actual_user,
            'created_at': {'$gte': cutoff}
        })
        if existing:
            skipped += 1
            continue

        adjustment = {
            'id': str(uuid.uuid4()),
            'item_name': entry['item_name'],
            'stamp': entry.get('stamp', ''),
            'poly_weight': entry['poly_weight'],
            'operation': entry['operation'],
            'adjusted_by': actual_user,
            'created_at': datetime.now(timezone.utc).isoformat(),
            'date': datetime.now(timezone.utc).date().isoformat()
        }
        
        saved_entries.append(adjustment)
        
        # Log activity
        await db.activity_log.insert_one({
            'user': actual_user,
            'user_role': current_user['role'],
            'action_type': 'polythene_adjustment',
            'description': f'{entry["operation"].upper()} {entry["poly_weight"]} kg polythene for {entry["item_name"]}',
            'details': entry,
            'timestamp': datetime.now(timezone.utc).isoformat()
        })
    
    # Insert all at once
    if saved_entries:
        await db.polythene_adjustments.insert_many(saved_entries)
        await db.notifications.insert_one({
            'id': str(uuid.uuid4()),
            'category': 'polythene',
            'type': 'polythene_batch',
            'message': f'{actual_user} adjusted polythene for {len(saved_entries)} items',
            'severity': 'info',
            'target_user': 'admin',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'read': False
        })
    
    _inv_cache.invalidate()
    return {
        'success': True,
        'message': f'{len(saved_entries)} polythene adjustments saved',
        'saved': len(saved_entries),
        'skipped': skipped,
        'count': len(saved_entries)
    }

@api_router.get("/polythene/today/{username}")
async def get_today_polythene_entries(username: str, current_user: dict = Depends(get_current_user)):
    """Get ALL polythene entries by user for today (no limit)"""
    if current_user['username'] != username and current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Can only view your own entries")
    today = datetime.now(timezone.utc).date().isoformat()
    
    entries = await db.polythene_adjustments.find(
        {'adjusted_by': username, 'date': today},
        {"_id": 0}
    ).to_list(None)  # Increased limit to 10000
    
    return entries

@api_router.delete("/polythene/{entry_id}")
async def delete_polythene_entry(entry_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a polythene entry (admin or own entry for polythene_executive)"""
    entry = await db.polythene_adjustments.find_one({'id': entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    if current_user['role'] == 'admin':
        pass  # admin can delete any entry
    elif current_user['role'] == 'polythene_executive' and entry.get('adjusted_by') == current_user['username']:
        pass  # polythene_executive can delete own entries
    else:
        raise HTTPException(status_code=403, detail="You can only delete your own polythene entries")
    result = await db.polythene_adjustments.delete_one({'id': entry_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    _inv_cache.invalidate()
    return {'success': True}

# ==================== ADMIN ACCOUNTABILITY ====================

@api_router.get("/activity-log")
async def get_activity_log(current_user: dict = Depends(get_current_user)):
    """Get complete activity log for admin accountability"""
    
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    activities = await db.activity_log.find({}, {"_id": 0}).sort('timestamp', -1).limit(200).to_list(200)
    return activities




@api_router.post("/mappings/create-new-item")
async def create_new_item_from_unmapped(
    transaction_name: str,
    stamp: str = "Unassigned",
    current_user: dict = Depends(get_current_user)
):
    """Create a completely new item from unmapped transaction name"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Check for duplicates
    existing = await db.master_items.find_one({"item_name": transaction_name})
    if existing:
        return {'success': False, 'message': f'Item "{transaction_name}" already exists in master items'}
    
    # Add to master_items as new item
    new_item = {
        'item_name': transaction_name,
        'stamp': stamp,
        'gr_wt': 0.0,
        'net_wt': 0.0,
        'is_master': True
    }
    
    await db.master_items.insert_one(new_item)
    
    # Also add to opening_stock with 0 quantity
    new_stock = {
        'item_name': transaction_name,
        'stamp': stamp,
        'unit': 'kg',
        'pc': 0,
        'gr_wt': 0.0,
        'net_wt': 0.0,
        'fine': 0.0,
        'labor_wt': 0.0,
        'labor_rs': 0.0,
        'rate': 0.0,
        'total': 0.0
    }
    
    await db.opening_stock.insert_one(new_stock)
    
    return {'success': True, 'message': f'New item "{transaction_name}" created with stamp: {stamp}'}


@api_router.get("/transactions")
async def get_transactions(type: Optional[str] = None, limit: int = 5000, current_user: dict = Depends(get_current_user)):
    """Get all transactions"""
    query = {} if not type else {"type": type}
    transactions = await db.transactions.find(query, {"_id": 0}).sort("date", -1).to_list(limit)
    return transactions

@api_router.get("/inventory/current")
async def get_current_inventory_endpoint(current_user: dict = Depends(get_current_user)):
    """Calculate current inventory: Opening Stock + Purchases - Sales (cached 30s)"""
    cached = _inv_cache.get('current_inventory')
    if cached:
        return cached
    result = await get_current_inventory()
    _inv_cache.set('current_inventory', result)
    return result


async def _replace_physical_stock_for_date(records: list, verification_date: str):
    """Shared helper: merge parsed records by (item name + stamp), delete only the selected date's rows,
    insert the new merged rows. Returns (count, message)."""
    merged_items = {}
    for record in records:
        name_key = record['item_name'].strip().lower()
        stamp_val = record.get('stamp', '') or ''
        # Merge by (name + stamp) to preserve separate stamp entries
        if stamp_val and stamp_val != 'Unassigned':
            merge_key = f"{name_key}||{stamp_val.strip().lower()}"
        else:
            merge_key = name_key
        if merge_key not in merged_items:
            merged_items[merge_key] = {
                'item_name': record['item_name'],
                'stamp': stamp_val,
                'pc': 0, 'gr_wt': 0.0, 'net_wt': 0.0, 'fine': 0.0,
                'verification_date': verification_date,
            }
        merged_items[merge_key]['gr_wt'] += record.get('gr_wt', 0)
        merged_items[merge_key]['net_wt'] += record.get('net_wt', 0)
        merged_items[merge_key]['fine'] += record.get('fine', 0)
        merged_items[merge_key]['pc'] += record.get('pc', 0)
        if stamp_val and not merged_items[merge_key]['stamp']:
            merged_items[merge_key]['stamp'] = stamp_val

    # Delete ONLY rows for the selected date — other dates untouched
    await db.physical_stock.delete_many({'verification_date': verification_date})

    stock_items = [PhysicalStock(**item).model_dump() for item in merged_items.values()]
    if stock_items:
        await db.physical_stock.insert_many(stock_items)
    await auto_normalize_stamps()

    total_net_wt = sum(i['net_wt'] for i in stock_items)
    message = f"Physical stock snapshot for {verification_date}: {len(stock_items)} items, {total_net_wt/1000:.3f} kg"
    return len(stock_items), message


@api_router.post("/physical-stock/upload")
async def upload_physical_stock(
    file: UploadFile = File(...),
    verification_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Upload physical stock file — replaces snapshot for the selected verification_date ONLY.
    Rows for other dates are preserved."""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or manager only")
    if not verification_date:
        raise HTTPException(status_code=400, detail="verification_date is required")
    content = await file.read()
    
    try:
        loop = asyncio.get_event_loop()
        records = await loop.run_in_executor(_parse_executor, parse_excel_file, content, 'physical_stock')
        
        if not records:
            raise HTTPException(status_code=400, detail="No valid records found in file")
        
        count, message = await _replace_physical_stock_for_date(records, verification_date)
        
        return {
            "success": True,
            "count": count,
            "verification_date": verification_date,
            "message": message,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")


@api_router.post("/physical-stock/upload-preview")
async def upload_physical_stock_preview(
    file: UploadFile = File(...),
    verification_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Parse physical stock file and return a preview diff. Creates a draft session."""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or manager only")

    if not verification_date:
        raise HTTPException(status_code=400, detail="verification_date is required")

    content = await file.read()
    loop = asyncio.get_event_loop()
    records = await loop.run_in_executor(_parse_executor, parse_excel_file, content, 'physical_stock')

    if not records:
        raise HTTPException(status_code=400, detail="No valid records found in file")

    has_net = any(r.get('has_net', False) for r in records)
    update_mode = 'gross_and_net' if has_net else 'gross_only'

    uploaded = {}
    for rec in records:
        item_key = rec['item_name'].strip().lower()
        if item_key not in uploaded:
            uploaded[item_key] = {'item_name': rec['item_name'], 'stamp': rec.get('stamp', ''), 'gr_wt': 0.0, 'net_wt': 0.0}
        uploaded[item_key]['gr_wt'] += rec.get('gr_wt', 0)
        uploaded[item_key]['net_wt'] += rec.get('net_wt', 0)

    base = await get_effective_physical_base_for_date(verification_date)
    has_existing_snapshot = await db.physical_stock.count_documents({'verification_date': verification_date}) > 0

    # Build comprehensive name→base_key reverse lookup
    all_mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    all_groups = await db.item_groups.find({}, {"_id": 0}).to_list(1000)
    ps_mapping_dict, _, _ = build_group_maps(all_groups, all_mappings)

    # Step 1: direct base key lookup
    name_to_base_key = {k: k for k in base}

    # Step 2: group members → base key (prefer member's own key; NO leader fallback)
    # Groups are for display only. Each member must resolve to its own base entry.
    for g in all_groups:
        for member in g.get('members', []):
            member_key = member.strip().lower()
            if member_key in base and member_key not in name_to_base_key:
                name_to_base_key[member_key] = member_key

    # Step 3: mapping transaction_name → master_name → base key (individual level only)
    for m in all_mappings:
        txn_key = m['transaction_name'].strip().lower()
        master_name = m['master_name'].strip()
        master_key = master_name.strip().lower()
        # Resolve to the master_name's own base entry (NOT group leader)
        if master_key in base:
            if txn_key not in name_to_base_key:
                name_to_base_key[txn_key] = master_key
            if master_key not in name_to_base_key:
                name_to_base_key[master_key] = master_key

    # Resolve uploaded items to base keys using the comprehensive lookup
    # No longer merges group members — each uploaded item stays as its own row
    resolved_uploads = {}  # base_key -> {base_item, gr_wt, net_wt}
    unmatched_uploads = []
    for item_key, upl in uploaded.items():
        resolved_base_key = name_to_base_key.get(item_key)

        if not resolved_base_key:
            # Fallback: try mapping resolution to master_name (individual level, no leader merge)
            raw_name = upl['item_name'].strip()
            master = ps_mapping_dict.get(raw_name, raw_name)
            master_key = master.strip().lower()
            if master_key in base:
                resolved_base_key = master_key

        if not resolved_base_key:
            unmatched_uploads.append(upl)
            continue

        base_item = base[resolved_base_key]

        # Merge into resolved_uploads keyed by the base entry key
        if resolved_base_key not in resolved_uploads:
            resolved_uploads[resolved_base_key] = {
                'base_item': base_item, 'gr_wt': 0.0, 'net_wt': 0.0,
            }
        resolved_uploads[resolved_base_key]['gr_wt'] += upl['gr_wt']
        resolved_uploads[resolved_base_key]['net_wt'] += upl['net_wt']

    preview_rows = []

    # Unmatched items
    for upl in unmatched_uploads:
        preview_rows.append({
            'item_name': upl['item_name'], 'stamp': upl.get('stamp', ''),
            'update_mode': update_mode, 'is_negative_grouped': False,
            'old_gr_wt': 0, 'new_gr_wt': round(upl['gr_wt'], 3), 'gr_delta': round(upl['gr_wt'], 3),
            'old_net_wt': 0, 'new_net_wt': round(upl['net_wt'], 3) if has_net else 0,
            'net_delta': round(upl['net_wt'], 3) if has_net else 0,
            'status': 'unmatched',
        })

    # Matched items (merged)
    for base_key, merged in resolved_uploads.items():
        base_item = merged['base_item']
        old_gr = base_item.get('gr_wt', 0)
        old_net = base_item.get('net_wt', 0)
        new_gr = round(merged['gr_wt'], 3)
        is_neg_grouped = base_item.get('is_negative_grouped', False)
        new_net = round(old_net, 3) if (is_neg_grouped or not has_net) else round(merged['net_wt'], 3)

        preview_rows.append({
            'item_name': base_item['item_name'], 'stamp': base_item.get('stamp', ''),
            'update_mode': update_mode, 'is_negative_grouped': is_neg_grouped,
            'old_gr_wt': round(old_gr, 3), 'new_gr_wt': new_gr, 'gr_delta': round(new_gr - old_gr, 3),
            'old_net_wt': round(old_net, 3), 'new_net_wt': new_net, 'net_delta': round(new_net - old_net, 3),
            'status': 'pending',
        })

    # Create draft session
    session_id = str(uuid.uuid4())
    draft_items = []
    for r in preview_rows:
        draft_items.append({
            'item_name': r['item_name'], 'stamp': r.get('stamp', ''),
            'status': r['status'], 'update_mode': r.get('update_mode', update_mode),
            'is_negative_grouped': r.get('is_negative_grouped', False),
            'old_gr_wt': r.get('old_gr_wt', 0), 'proposed_gr_wt': r.get('new_gr_wt', 0),
            'final_gr_wt': r.get('old_gr_wt', 0),
            'gr_delta': 0,
            'old_net_wt': r.get('old_net_wt', 0), 'proposed_net_wt': r.get('new_net_wt', 0),
            'final_net_wt': r.get('old_net_wt', 0),
            'net_delta': 0,
        })
    draft_items.sort(key=lambda x: ((x.get('stamp', '') or '').lower(), x['item_name'].lower()))

    await db.physical_stock_update_sessions.insert_one({
        'session_id': session_id,
        'verification_date': verification_date,
        'session_state': 'draft',
        'created_at': datetime.now(timezone.utc).isoformat(),
        'applied_by': current_user['username'],
        'update_mode': update_mode,
        'uploaded_count': len(preview_rows),
        'applied_count': 0, 'rejected_count': 0,
        'unmatched_count': sum(1 for r in preview_rows if r['status'] == 'unmatched'),
        'skipped_count': 0,
        'items': draft_items,
    })

    matched_count = sum(1 for r in preview_rows if r['status'] == 'pending')
    return {
        "success": True,
        "preview_session_id": session_id,
        "update_mode": update_mode,
        "verification_date": verification_date,
        "has_existing_snapshot": has_existing_snapshot,
        "preview_rows": preview_rows,
        "summary": {
            "total_uploaded": len(preview_rows),
            "matched": matched_count,
            "unmatched": sum(1 for r in preview_rows if r['status'] == 'unmatched'),
            "base_item_count": len(base),
        }
    }


@api_router.post("/physical-stock/apply-updates")
async def apply_physical_stock_updates(
    request: Dict,
    current_user: dict = Depends(get_current_user)
):
    """Apply approved items within an existing preview session."""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or manager only")

    items = request.get('items', [])
    verification_date = request.get('verification_date')
    preview_session_id = request.get('preview_session_id')

    if not verification_date:
        raise HTTPException(status_code=400, detail="verification_date is required")
    if not items:
        raise HTTPException(status_code=400, detail="No items to apply")

    # Materialize snapshot if needed
    has_snapshot = await db.physical_stock.count_documents({'verification_date': verification_date}) > 0
    if not has_snapshot:
        book_base = await get_effective_physical_base_for_date(verification_date)
        if book_base:
            bulk_docs = [{
                'item_name': bi['item_name'], 'stamp': bi.get('stamp', ''),
                'gr_wt': bi['gr_wt'], 'net_wt': bi['net_wt'],
                'is_negative_grouped': bi.get('is_negative_grouped', False),
                'verification_date': verification_date,
            } for bi in book_base.values()]
            if bulk_docs:
                await db.physical_stock.insert_many(bulk_docs)

    updated_count = 0
    results = []
    applied_items_detail = []
    for item in items:
        item_name = item.get('item_name', '')
        key = item_name.strip().lower()
        update_mode = item.get('update_mode', 'gross_only')
        new_gr = item.get('new_gr_wt', 0)
        new_net = item.get('new_net_wt', 0)
        is_neg_grouped = item.get('is_negative_grouped', False)

        existing = await db.physical_stock.find_one(
            {'item_name': {'$regex': f'^{re.escape(key)}$', '$options': 'i'}, 'verification_date': verification_date},
            {"_id": 0}
        )
        if not existing:
            results.append({'item_name': item_name, 'status': 'skipped', 'reason': 'not_found_for_date'})
            continue

        old_gr = existing.get('gr_wt', 0)
        old_net = existing.get('net_wt', 0)
        update_fields = {'gr_wt': new_gr}
        if is_neg_grouped or update_mode == 'gross_only':
            update_fields['net_wt'] = old_net
        else:
            update_fields['net_wt'] = new_net

        await db.physical_stock.update_one(
            {'item_name': existing['item_name'], 'verification_date': verification_date},
            {'$set': update_fields}
        )
        updated_count += 1
        applied_items_detail.append({
            'item_name': existing['item_name'],
            'old_gr_wt': round(old_gr, 3), 'final_gr_wt': round(new_gr, 3), 'gr_delta': round(new_gr - old_gr, 3),
            'old_net_wt': round(old_net, 3), 'final_net_wt': round(update_fields['net_wt'], 3),
            'net_delta': round(update_fields['net_wt'] - old_net, 3),
        })
        results.append({'item_name': existing['item_name'], 'status': 'applied'})

        # Upsert inventory baseline — physical stock becomes the new starting point
        baseline_key = existing['item_name'].strip().lower()
        await db.inventory_baselines.update_one(
            {'item_key': baseline_key},
            {'$set': {
                'item_key': baseline_key,
                'item_name': existing['item_name'],
                'baseline_date': verification_date,
                'gr_wt': round(new_gr, 3),
                'net_wt': round(update_fields['net_wt'], 3),
                'stamp': existing.get('stamp', ''),
                'updated_at': datetime.now(timezone.utc).isoformat(),
                'session_id': preview_session_id or '',
            }},
            upsert=True
        )

    # Update draft session if provided
    if preview_session_id and updated_count > 0:
        session = await db.physical_stock_update_sessions.find_one({'session_id': preview_session_id})
        if session:
            applied_names = {d['item_name'].lower() for d in applied_items_detail}
            updated_items = []
            for si in session.get('items', []):
                name_lower = si['item_name'].strip().lower()
                if name_lower in applied_names and si.get('status') == 'pending':
                    detail = next((d for d in applied_items_detail if d['item_name'].lower() == name_lower), None)
                    if detail:
                        si['status'] = 'applied'
                        si['final_gr_wt'] = detail['final_gr_wt']
                        si['final_net_wt'] = detail['final_net_wt']
                        si['gr_delta'] = detail['gr_delta']
                        si['net_delta'] = detail['net_delta']
                updated_items.append(si)

            applied_rows = [i for i in updated_items if i.get('status') == 'applied']
            await db.physical_stock_update_sessions.update_one(
                {'session_id': preview_session_id},
                {'$set': {
                    'items': updated_items,
                    'applied_count': len(applied_rows),
                    'applied_at': datetime.now(timezone.utc).isoformat(),
                    'applied_by': current_user['username'],
                    'old_total_gr_wt': round(sum(i.get('old_gr_wt', 0) for i in applied_rows), 3),
                    'new_total_gr_wt': round(sum(i.get('final_gr_wt', 0) for i in applied_rows), 3),
                    'gr_delta_total': round(sum(i.get('gr_delta', 0) for i in applied_rows), 3),
                    'old_total_net_wt': round(sum(i.get('old_net_wt', 0) for i in applied_rows), 3),
                    'new_total_net_wt': round(sum(i.get('final_net_wt', 0) for i in applied_rows), 3),
                    'net_delta_total': round(sum(i.get('net_delta', 0) for i in applied_rows), 3),
                }}
            )

    await save_action(
        'physical_stock_partial_update',
        f"Partial physical stock update for {verification_date}: {updated_count} items",
        {'updated_count': updated_count, 'verification_date': verification_date, 'by': current_user['username']}
    )

    return {
        "success": True,
        "updated_count": updated_count,
        "skipped_count": len(items) - updated_count,
        "verification_date": verification_date,
        "results": results,
        "message": f"Physical stock updated for {verification_date}: {updated_count} item{'s' if updated_count != 1 else ''} applied"
    }

@api_router.post("/physical-stock/finalize-session")
async def finalize_physical_stock_session(
    request: Dict,
    current_user: dict = Depends(get_current_user)
):
    """Finalize a draft session: remaining pending rows become rejected. Empty sessions get abandoned."""
    session_id = request.get('session_id')
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id is required")

    session = await db.physical_stock_update_sessions.find_one({'session_id': session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if session.get('session_state') != 'draft':
        return {"success": True, "message": "Session already finalized"}

    items = session.get('items', [])
    applied_count = sum(1 for i in items if i.get('status') == 'applied')

    if applied_count == 0:
        # Abandon empty session
        await db.physical_stock_update_sessions.update_one(
            {'session_id': session_id},
            {'$set': {'session_state': 'abandoned'}}
        )
        return {"success": True, "message": "Session abandoned (no items applied)"}

    # Mark remaining pending as rejected — reset final weights to old values
    for i in items:
        if i.get('status') == 'pending':
            i['status'] = 'rejected'
            i['final_gr_wt'] = i.get('old_gr_wt', 0)
            i['final_net_wt'] = i.get('old_net_wt', 0)
            i['gr_delta'] = 0
            i['net_delta'] = 0

    rejected_count = sum(1 for i in items if i.get('status') == 'rejected')
    await db.physical_stock_update_sessions.update_one(
        {'session_id': session_id},
        {'$set': {
            'session_state': 'finalized',
            'finalized_at': datetime.now(timezone.utc).isoformat(),
            'items': items,
            'rejected_count': rejected_count,
        }}
    )
    _inv_cache.invalidate()
    return {"success": True, "message": f"Session finalized: {applied_count} applied, {rejected_count} rejected"}

@api_router.post("/physical-stock/update-history/{session_id}/reverse")
async def reverse_physical_stock_session(
    session_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Reverse the latest unreversed session for its date. Restores old weights."""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or manager only")

    session = await db.physical_stock_update_sessions.find_one({'session_id': session_id}, {'_id': 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if session.get('is_reversed'):
        raise HTTPException(status_code=400, detail="Session already reversed")
    if session.get('session_state') not in ('finalized', 'draft'):
        raise HTTPException(status_code=400, detail="Only finalized/draft sessions can be reversed")

    v_date = session['verification_date']

    # Check it's the latest unreversed session for this date
    later = await db.physical_stock_update_sessions.find_one({
        'verification_date': v_date,
        'session_state': {'$in': ['finalized', 'draft']},
        'is_reversed': {'$ne': True},
        'applied_at': {'$gt': session.get('applied_at', session.get('created_at', ''))},
    })
    if later:
        raise HTTPException(status_code=400, detail="Cannot reverse: a later unreversed session exists for this date. Reverse the latest session first.")

    # Restore old weights for applied rows
    restored = 0
    for item in session.get('items', []):
        if item.get('status') != 'applied':
            continue
        key = item['item_name'].strip().lower()
        await db.physical_stock.update_one(
            {'item_name': {'$regex': f'^{re.escape(key)}$', '$options': 'i'}, 'verification_date': v_date},
            {'$set': {'gr_wt': item['old_gr_wt'], 'net_wt': item['old_net_wt']}}
        )
        # Remove inventory baseline only if it belongs to THIS session
        existing_bl = await db.inventory_baselines.find_one({'item_key': key}, {'_id': 0})
        if existing_bl and existing_bl.get('session_id') == session_id:
            await db.inventory_baselines.delete_one({'item_key': key})
        elif existing_bl and existing_bl.get('session_id') != session_id:
            pass  # Baseline belongs to a different session — keep it
        else:
            # No session_id recorded or no baseline — safe to delete
            await db.inventory_baselines.delete_one({'item_key': key})
        restored += 1

    await db.physical_stock_update_sessions.update_one(
        {'session_id': session_id},
        {'$set': {
            'is_reversed': True,
            'reversed_at': datetime.now(timezone.utc).isoformat(),
            'reversed_by': current_user['username'],
            'session_state': 'reversed',
        }}
    )

    await save_action(
        'physical_stock_reverse',
        f"Reversed physical stock session for {v_date}: {restored} items restored",
        {'session_id': session_id, 'verification_date': v_date, 'restored': restored, 'by': current_user['username']}
    )

    _inv_cache.invalidate()
    return {"success": True, "restored_count": restored, "message": f"Reversed: {restored} items restored to previous values"}

@api_router.post("/physical-stock/fix-group-baselines")
async def fix_group_baselines(current_user: dict = Depends(get_current_user)):
    """Fix existing group-level baselines by splitting them into member-level baselines.
    Uses get_current_inventory (without baselines) to compute each member's actual
    stock as of the baseline date, then assigns each member its share of the combined
    baseline based on those proportions."""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")

    all_groups = await db.item_groups.find({}, {"_id": 0}).to_list(1000)
    all_baselines = await db.inventory_baselines.find({}, {"_id": 0}).to_list(None)

    group_leaders = {g['group_name'] for g in all_groups}
    group_members_map = {g['group_name']: g.get('members', []) for g in all_groups}

    # Index all baselines by item_key for quick lookup
    baseline_keys = {bl['item_key'] for bl in all_baselines}

    # Find group-level baselines that need splitting.
    # A baseline is "group-level" ONLY if the leader has a baseline but NOT all members
    # have their own separate baselines. If every member already has its own baseline,
    # the leader's baseline is just its individual entry — skip it.
    to_fix = []
    for bl in all_baselines:
        bl_name = bl['item_name']
        if bl_name in group_leaders and len(group_members_map.get(bl_name, [])) >= 2:
            members = group_members_map[bl_name]
            # Check if all OTHER members (non-leader) already have their own baselines
            non_leader_members = [m for m in members if m != bl_name]
            all_members_have_baselines = all(
                m.strip().lower() in baseline_keys for m in non_leader_members
            )
            if not all_members_have_baselines:
                to_fix.append(bl)

    if not to_fix:
        return {"success": True, "fixed_count": 0, "details": [], "message": "No group-level baselines found"}

    fixed = []
    master_items = await db.master_items.find({}, {"_id": 0}).to_list(None)
    master_stamp_dict = {m['item_name']: m['stamp'] for m in master_items}

    for bl in to_fix:
        bl_name = bl['item_name']
        bl_key = bl['item_key']
        members = group_members_map[bl_name]
        baseline_date = bl['baseline_date']
        session_id = bl.get('session_id', '')

        # Temporarily remove this baseline so we can compute raw book stock
        await db.inventory_baselines.delete_one({'item_key': bl_key})

        # Compute inventory as of baseline_date WITHOUT the baseline
        inv = await get_current_inventory(as_of_date=baseline_date)
        all_inv_items = inv.get('inventory', []) + inv.get('negative_items', [])

        # Find the group item and get member-level breakdown
        group_item = next((it for it in all_inv_items if it['item_name'] == bl_name), None)
        member_book = {}
        if group_item and group_item.get('members'):
            for m in group_item['members']:
                member_book[m['item_name']] = {'gr_wt': m['gr_wt'], 'net_wt': m['net_wt']}
        else:
            # Group item exists but no member breakdown — assign all to leader
            if group_item:
                member_book[bl_name] = {'gr_wt': group_item['gr_wt'], 'net_wt': group_item['net_wt']}

        # Compute what the baseline should be for each member:
        # baseline_member = book_member + (baseline_total - book_total) * (book_member / book_total)
        # Simplified: each member's baseline = baseline_total * (book_member / book_total)
        # But better: use book value + proportional delta
        total_book_gr = sum(m['gr_wt'] for m in member_book.values()) or 1
        total_book_net = sum(m['net_wt'] for m in member_book.values()) or 1
        baseline_total_gr = bl['gr_wt']
        baseline_total_net = bl['net_wt']
        delta_gr = baseline_total_gr - total_book_gr
        delta_net = baseline_total_net - total_book_net

        member_baselines = {}
        for m_name in members:
            m_data = member_book.get(m_name, {'gr_wt': 0, 'net_wt': 0})
            # Distribute delta proportionally based on absolute book values
            abs_total = sum(abs(m['gr_wt']) for m in member_book.values()) or 1
            ratio = abs(m_data['gr_wt']) / abs_total if abs_total > 0 else (1.0 / len(members))
            m_baseline_gr = round(m_data['gr_wt'] + delta_gr * ratio, 3)
            m_baseline_net = round(m_data['net_wt'] + delta_net * ratio, 3)

            m_key = m_name.strip().lower()
            await db.inventory_baselines.update_one(
                {'item_key': m_key},
                {'$set': {
                    'item_key': m_key,
                    'item_name': m_name,
                    'baseline_date': baseline_date,
                    'gr_wt': m_baseline_gr,
                    'net_wt': m_baseline_net,
                    'stamp': master_stamp_dict.get(m_name, bl.get('stamp', '')),
                    'updated_at': datetime.now(timezone.utc).isoformat(),
                    'session_id': session_id,
                }},
                upsert=True
            )
            member_baselines[m_name] = m_baseline_gr

        fixed.append({
            'leader': bl_name,
            'members_created': list(member_baselines.keys()),
            'original_gr': bl['gr_wt'],
            'book_at_date': {m: d['gr_wt'] for m, d in member_book.items()},
            'new_baselines': member_baselines,
        })

    return {
        "success": True,
        "fixed_count": len(fixed),
        "details": fixed,
        "message": f"Split {len(fixed)} group-level baseline(s) into member-level baselines"
    }


@api_router.post("/physical-stock/restore-group-baselines")
async def restore_group_baselines(current_user: dict = Depends(get_current_user)):
    """One-time fix: Delete all member baselines for groups, restore the original
    group-level baseline from the physical_stock session, then run the split once.
    This fixes double-split corruption."""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")

    all_groups = await db.item_groups.find({}, {"_id": 0}).to_list(1000)
    group_members_map = {g['group_name']: g.get('members', []) for g in all_groups}
    master_items = await db.master_items.find({}, {"_id": 0}).to_list(None)
    master_stamp_dict = {m['item_name']: m['stamp'] for m in master_items}

    results = []
    for gname, members in group_members_map.items():
        if len(members) < 2:
            continue

        # Delete ALL baselines for this group's members
        deleted = 0
        for m in members:
            m_key = m.strip().lower()
            r = await db.inventory_baselines.delete_many({'item_key': m_key})
            deleted += r.deleted_count

        if deleted == 0:
            continue

        # Find the original physical stock session that created these baselines
        # Look for sessions with verification_date, find any that applied this group's leader
        sessions = await db.physical_stock_update_sessions.find(
            {'session_state': 'finalized', 'is_reversed': {'$ne': True}},
            {'_id': 0}
        ).sort('applied_at', -1).to_list(50)

        # Find the session that has data for this group
        orig_session = None
        orig_gr = 0
        orig_net = 0
        for sess in sessions:
            for item in sess.get('items', []):
                item_name = item.get('item_name', '').strip()
                if item_name == gname and item.get('status') == 'applied':
                    orig_session = sess
                    orig_gr = item.get('final_gr_wt', 0)
                    orig_net = item.get('final_net_wt', 0)
                    break
            if orig_session:
                break

        if not orig_session or (orig_gr == 0 and orig_net == 0):
            results.append({'group': gname, 'status': 'no_session_found', 'deleted': deleted})
            continue

        baseline_date = orig_session['verification_date']

        # Now compute book stock as of baseline_date WITHOUT any baselines
        _inv_cache.invalidate()
        inv = await get_current_inventory(as_of_date=baseline_date)

        # Get member-level book stock from by_stamp (individual level)
        member_book = {}
        for stamp_items in inv.get('by_stamp', {}).values():
            for si in stamp_items:
                if si['item_name'] in members:
                    member_book[si['item_name']] = si.get('gr_wt', 0)

        # Split the original physical stock value proportionally
        total_book = sum(abs(v) for v in member_book.values()) or 1
        member_baselines = {}
        for m_name in members:
            m_book = member_book.get(m_name, 0)
            ratio = abs(m_book) / total_book if total_book > 0 else 1.0 / len(members)
            m_baseline_gr = round(orig_gr * ratio, 3)

            # Net weight: proportional too
            m_baseline_net = round(orig_net * ratio, 3)

            m_key = m_name.strip().lower()
            await db.inventory_baselines.update_one(
                {'item_key': m_key},
                {'$set': {
                    'item_key': m_key,
                    'item_name': m_name,
                    'baseline_date': baseline_date,
                    'gr_wt': m_baseline_gr,
                    'net_wt': m_baseline_net,
                    'stamp': master_stamp_dict.get(m_name, ''),
                    'updated_at': datetime.now(timezone.utc).isoformat(),
                    'session_id': orig_session.get('session_id', ''),
                    'is_member_baseline': True,
                }},
                upsert=True
            )
            member_baselines[m_name] = m_baseline_gr

        results.append({
            'group': gname,
            'original_physical_gr': orig_gr,
            'baseline_date': baseline_date,
            'member_baselines': member_baselines,
            'book_at_date': member_book,
        })

    _inv_cache.invalidate()
    return {"success": True, "results": results}

@api_router.get("/physical-stock/dates")
async def get_physical_stock_dates(current_user: dict = Depends(get_current_user)):
    """Return distinct physical stock verification_date values sorted descending."""
    dates = await db.physical_stock.distinct("verification_date")
    dates = sorted([d for d in dates if d], reverse=True)
    return {"dates": dates}

@api_router.get("/physical-stock/update-history")
async def get_physical_stock_update_history(
    verification_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Return session summaries for physical stock updates (excluding abandoned and empty drafts)."""
    query = {'session_state': {'$in': ['finalized', 'reversed']}}
    if verification_date:
        query['verification_date'] = verification_date
    # Also include drafts that have applied items (still open)
    draft_query = {'session_state': 'draft', 'applied_count': {'$gt': 0}}
    if verification_date:
        draft_query['verification_date'] = verification_date

    finalized = await db.physical_stock_update_sessions.find(
        query, {"_id": 0, "items": 0}
    ).sort("created_at", -1).to_list(100)
    drafts = await db.physical_stock_update_sessions.find(
        draft_query, {"_id": 0, "items": 0}
    ).sort("created_at", -1).to_list(20)
    sessions = drafts + finalized

    # Determine which session is reversible (latest unreversed per date)
    latest_unreversed = {}
    for s in sessions:
        vd = s['verification_date']
        if not s.get('is_reversed') and s.get('session_state') in ('finalized', 'draft') and s.get('applied_count', 0) > 0:
            if vd not in latest_unreversed:
                latest_unreversed[vd] = s['session_id']
    for s in sessions:
        s['reversible'] = s['session_id'] == latest_unreversed.get(s['verification_date'])

    return {"sessions": sessions}

@api_router.get("/physical-stock/update-history/{session_id}")
async def get_physical_stock_update_session(
    session_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Return full session details including sorted item rows."""
    session = await db.physical_stock_update_sessions.find_one(
        {"session_id": session_id}, {"_id": 0}
    )
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    return session

@api_router.get("/physical-stock/compare")
async def compare_physical_with_book(verification_date: str, current_user: dict = Depends(get_current_user)):
    """Compare physical stock with book stock for a specific date.
    Book stock = computed closing stock as of verification_date.
    Physical stock = saved snapshot for that date."""
    
    if not verification_date:
        raise HTTPException(status_code=400, detail="verification_date is required")
    
    # Get date-scoped book stock using same identity model as current stock page
    book_base = await _flat_base_from_inventory(verification_date)
    book_items = book_base  # already keyed by normalized name
    
    # Get physical stock snapshot for this date
    physical = await db.physical_stock.find({'verification_date': verification_date}, {"_id": 0}).to_list(None)
    physical_items = {item['item_name'].strip().lower(): item for item in physical}
    
    # Compare
    matches = []
    discrepancies = []
    only_in_book = []
    only_in_physical = []
    
    for key, book_item in book_items.items():
        if key in physical_items:
            phys_item = physical_items[key]
            book_net = book_item['net_wt']
            phys_net = phys_item.get('net_wt', 0)
            book_gross = book_item.get('gr_wt', 0)
            phys_gross = phys_item.get('gr_wt', 0)
            diff_net = phys_net - book_net
            diff_gross = phys_gross - book_gross
            
            comparison = {
                'item_name': book_item['item_name'],
                'stamp': book_item.get('stamp', ''),
                'book_net_wt': book_net,
                'physical_net_wt': phys_net,
                'book_gross_wt': book_gross,
                'physical_gross_wt': phys_gross,
                'difference': diff_net,
                'difference_kg': round(diff_net/1000, 3),
                'gross_difference': diff_gross,
                'gross_difference_kg': round(diff_gross/1000, 3),
                'match_percentage': round((min(abs(book_net), abs(phys_net)) / max(abs(book_net), abs(phys_net)) * 100) if max(abs(book_net), abs(phys_net)) > 0 else 100, 2)
            }
            
            # Use gross difference for classification when net values are identical (gross-only files)
            classify_diff = diff_gross if (abs(diff_net) < 1 and abs(diff_gross) >= 1) else diff_net
            if abs(classify_diff) < 10:
                matches.append(comparison)
            else:
                discrepancies.append(comparison)
        else:
            only_in_book.append({
                'item_name': book_item['item_name'],
                'stamp': book_item.get('stamp', ''),
                'book_net_wt': book_item['net_wt'],
                'book_net_wt_kg': round(book_item['net_wt']/1000, 3)
            })
    
    for key in physical_items:
        if key not in book_items:
            only_in_physical.append({
                'item_name': physical_items[key]['item_name'],
                'stamp': physical_items[key].get('stamp', ''),
                'physical_net_wt': physical_items[key].get('net_wt', 0),
                'physical_net_wt_kg': round(physical_items[key].get('net_wt', 0)/1000, 3)
            })
    
    discrepancies.sort(key=lambda x: abs(x['difference']), reverse=True)
    
    total_book_net = sum(item['net_wt'] for item in book_items.values())
    total_physical_net = sum(item.get('net_wt', 0) for item in physical_items.values())
    total_book_gross = sum(item.get('gr_wt', 0) for item in book_items.values())
    total_physical_gross = sum(item.get('gr_wt', 0) for item in physical_items.values())
    
    return {
        "summary": {
            "total_book_kg": round(total_book_net/1000, 3),
            "total_physical_kg": round(total_physical_net/1000, 3),
            "total_difference_kg": round((total_physical_net - total_book_net)/1000, 3),
            "total_book_gross_kg": round(total_book_gross/1000, 3),
            "total_physical_gross_kg": round(total_physical_gross/1000, 3),
            "total_difference_gross_kg": round((total_physical_gross - total_book_gross)/1000, 3),
            "match_count": len(matches),
            "discrepancy_count": len(discrepancies),
            "only_in_book_count": len(only_in_book),
            "only_in_physical_count": len(only_in_physical)
        },
        "matches": matches[:50],
        "discrepancies": discrepancies[:50],
        "only_in_book": only_in_book[:50],
        "only_in_physical": only_in_physical[:50]
    }

@api_router.get("/history/recent-uploads")
async def get_recent_uploads(current_user: dict = Depends(get_current_user)):
    """Get ALL file uploads for undo selection (admin only)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    actions = await db.action_history.find(
        {"action_type": {"$in": [
            "upload_purchase", "upload_sale", "upload_branch_transfer",
            "upload_opening_stock", "upload_master_stock", "upload_physical_stock",
            "client_batch_upload"
        ]}},
        {"_id": 0}
    ).sort("timestamp", -1).to_list(1000)
    
    return actions


@api_router.get("/inventory/stamp-breakdown/{stamp}")
async def get_stamp_breakdown(stamp: str, current_user: dict = Depends(get_current_user)):
    """Get detailed breakdown for a specific stamp using the authoritative inventory calculation"""
    
    # Use the single source of truth: get_current_inventory
    inventory_response = await get_current_inventory()
    all_items = inventory_response.get('inventory', []) + inventory_response.get('negative_items', [])
    
    # Filter items in this stamp
    stamp_items = [item for item in all_items if item.get('stamp') == stamp]
    
    # Calculate totals from inventory
    current_gross = sum(item.get('gr_wt', 0) for item in stamp_items)
    current_net = sum(item.get('net_wt', 0) for item in stamp_items)
    
    # Get opening stock for breakdown display
    opening = await db.opening_stock.find({"stamp": stamp}, {"_id": 0}).to_list(1000)
    opening_gross = sum(item.get('gr_wt', 0) for item in opening)
    opening_net = sum(item.get('net_wt', 0) for item in opening)
    
    # Collect all item names (master + mapped) for transaction queries
    item_names = [item['item_name'] for item in stamp_items]
    all_mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    mapped_names = [m['transaction_name'] for m in all_mappings if m['master_name'] in item_names]
    all_names = list(set(item_names + mapped_names))
    
    # Get purchase/sale breakdown for display
    purchases = await db.transactions.find({
        "item_name": {"$in": all_names},
        "type": {"$in": ["purchase", "purchase_return"]}
    }, {"_id": 0}).to_list(None)
    purchase_gross = sum(t.get('gr_wt', 0) for t in purchases)
    purchase_net = sum(t.get('net_wt', 0) for t in purchases)
    
    sales = await db.transactions.find({
        "item_name": {"$in": all_names},
        "type": {"$in": ["sale", "sale_return"]}
    }, {"_id": 0}).to_list(None)
    sale_gross = sum(t.get('gr_wt', 0) for t in sales)
    sale_net = sum(t.get('net_wt', 0) for t in sales)
    
    # Count master vs mapped
    master_count = len([i for i in stamp_items if not any(m['transaction_name'] == i['item_name'] for m in all_mappings)])
    mapped_count = len(stamp_items) - master_count
    
    return {
        "stamp": stamp,
        "opening_gross": round(opening_gross, 3),
        "opening_net": round(opening_net, 3),
        "purchase_gross": round(purchase_gross, 3),
        "purchase_net": round(purchase_net, 3),
        "sale_gross": round(sale_gross, 3),
        "sale_net": round(sale_net, 3),
        "current_gross": round(current_gross, 3),
        "current_net": round(current_net, 3),
        "item_count": len(stamp_items),
        "mapped_count": mapped_count
    }


@api_router.post("/admin/normalize-stamps")
async def normalize_all_stamps(current_user: dict = Depends(get_current_user)):
    """Normalize all stamps to CAPS format (Admin only)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    total_updated = await auto_normalize_stamps()
    
    if total_updated == 0:
        return {"success": True, "message": "All stamps already normalized!", "stamps_updated": 0}
    
    await save_action('normalize_stamps', f'Normalized {total_updated} documents to CAPS format', {'count': total_updated})
    
    return {
        "success": True,
        "message": f"Normalized {total_updated} documents to CAPS format",
        "total_documents": total_updated
    }

@api_router.post("/history/undo-upload")
async def undo_upload(batch_id: str, current_user: dict = Depends(get_current_user)):
    """Undo a specific file upload by batch_id — restores previously replaced data"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Find the action
    action = await db.action_history.find_one({"data_snapshot.batch_id": batch_id})
    if not action:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    # Delete all transactions with this batch_id
    delete_result = await db.transactions.delete_many({"batch_id": batch_id})
    
    # Restore backed-up records if they exist
    restored_count = 0
    backup = await db.replaced_records.find_one({"batch_id": batch_id}, {"_id": 0})
    if backup and backup.get("records"):
        await batch_insert(db.transactions, backup["records"])
        restored_count = len(backup["records"])
        await db.replaced_records.delete_one({"batch_id": batch_id})
    
    # Mark action as undone
    await db.action_history.update_one(
        {"data_snapshot.batch_id": batch_id},
        {"$set": {"can_undo": False}}
    )
    
    msg = f"Undone: {action.get('description', 'Upload')}. Removed {delete_result.deleted_count} records"
    if restored_count > 0:
        msg += f", restored {restored_count} previous records"
    
    return {
        "success": True,
        "message": msg,
        "deleted_count": delete_result.deleted_count,
        "restored_count": restored_count
    }


@api_router.post("/master-stock/upload")
async def upload_master_stock(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Upload STOCK 2026 as master reference - FINAL item names and stamps"""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or Manager only")
    content = await file.read()
    
    try:
        import pandas as pd
        df = pd.read_excel(BytesIO(content), header=0)
        df = df.fillna('')
        df.columns = df.columns.str.strip()
        
        # Remove totals row
        df = df[~df['Item Name'].astype(str).str.lower().str.contains('total', na=False)]
        
        # Clear existing opening stock and master items
        await db.opening_stock.delete_many({})
        await db.master_items.delete_many({})
        
        opening_items = []
        master_items = []
        
        for _, row in df.iterrows():
            item_name = str(row.get('Item Name', '')).strip()
            if not item_name or len(item_name) < 2:
                continue
            
            stamp = str(row.get('Stamp', '')).strip()
            gr_wt = float(row.get('Gross weigth', 0) or 0)
            net_wt = float(row.get('Net Weight', 0) or 0)
            
            # Opening stock
            opening_items.append({
                'item_name': item_name,
                'stamp': stamp,
                'unit': 'kg',
                'pc': 0,
                'gr_wt': gr_wt,
                'net_wt': net_wt,
                'fine': 0.0,
                'labor_wt': 0.0,
                'labor_rs': 0.0,
                'rate': 0.0,
                'total': 0.0
            })
            
            # Master reference
            master_items.append({
                'item_name': item_name,
                'stamp': stamp,
                'gr_wt': gr_wt,
                'net_wt': net_wt,
                'is_master': True
            })
        
        await db.opening_stock.insert_many(opening_items)
        await db.master_items.insert_many(master_items)
        
        # Auto-normalize stamps after upload
        await auto_normalize_stamps()
        
        total_net = sum(i['net_wt'] for i in opening_items)
        
        return {
            "success": True,
            "count": len(opening_items),
            "total_net_kg": round(total_net / 1000, 3),
            "message": f"Master stock uploaded: {len(opening_items)} items, {total_net/1000:.3f} kg"
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")

@api_router.post("/purchase-ledger/upload")
async def upload_purchase_ledger(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Upload PURCHASE_CUMUL file to create/update purchase rate ledger"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    content = await file.read()
    
    try:
        import pandas as pd
        df = pd.read_excel(BytesIO(content), header=2)
        df = df.fillna(0)
        df.columns = df.columns.str.strip()
        
        await db.purchase_ledger.delete_many({})
        
        ledger_items = []
        
        for _, row in df.iterrows():
            item_name = str(row.get('Particular', '')).strip()
            if not item_name or len(item_name) < 2:
                continue
            
            if 'total' in item_name.lower():
                continue
            
            less = float(row.get('Less', 0) or 0)
            sil_fine = float(row.get('Sil.Fine', 0) or 0)
            total = float(row.get('Total', 0) or 0)
            
            if less > 0:
                purchase_tunch = (sil_fine / less * 100)
                labour_per_kg = (total / less)
                
                ledger_items.append({
                    'item_name': item_name,
                    'purchase_tunch': purchase_tunch,
                    'labour_per_kg': labour_per_kg,
                    'total_purchased_kg': less,
                    'total_fine_kg': sil_fine,
                    'total_labour': total
                })
        
        if ledger_items:
            await db.purchase_ledger.insert_many(ledger_items)
        
        return {
            "success": True,
            "count": len(ledger_items),
            "message": f"Purchase ledger created with {len(ledger_items)} items"
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")

@api_router.get("/analytics/customer-profit")
async def get_customer_profit(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Calculate profit per customer (silver & labour).
    Sale returns are treated as purchases — we 'buy back' at the return rate,
    profiting if the return rate is below our purchase cost."""
    
    query = {}
    if start_date and end_date:
        end_date_with_time = end_date + ' 23:59:59'
        query['date'] = {'$gte': start_date, '$lte': end_date_with_time}
    
    # Get all sales + returns
    transactions = await db.transactions.find(
        {**query, "type": {"$in": ["sale", "sale_return"]}},
        {"_id": 0}
    ).to_list(None)
    
    # Get purchase ledger — GROUP AWARE
    ledger = await db.purchase_ledger.find({}, {"_id": 0}).to_list(None)
    all_mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    all_groups = await db.item_groups.find({}, {"_id": 0}).to_list(1000)
    grp_ledger = build_group_ledger(ledger, all_groups, all_mappings)
    mapping_dict, member_to_leader, _ = build_group_maps(all_groups, all_mappings)
    
    # Group by customer
    customer_profit = defaultdict(lambda: {
        'customer_name': '',
        'silver_profit_kg': 0.0,
        'labour_profit_inr': 0.0,
        'total_sold_kg': 0.0,
        'transaction_count': 0
    })
    
    for txn in transactions:
        customer = txn.get('party_name', 'Unknown')
        if not customer:
            continue
        
        raw_item_name = txn.get('item_name', '')
        leader_name = resolve_to_leader(raw_item_name, mapping_dict, member_to_leader)
        
        txn_tunch = float(txn.get('tunch', 0) or 0)
        txn_net_wt = txn.get('net_wt', 0)
        txn_total = txn.get('total_amount', 0) or txn.get('labor', 0)  # Total Rs (labour charges)
        is_return = txn['type'] == 'sale_return'
        
        # Get purchase cost from GROUP-AWARE ledger
        ledger_item = grp_ledger.get(leader_name) or grp_ledger.get(raw_item_name)
        if ledger_item:
            purchase_tunch = ledger_item.get('purchase_tunch', 0)
            purchase_cost_per_gram = ledger_item.get('labour_per_kg', 0) / 1000
        else:
            purchase_tunch = 0
            purchase_cost_per_gram = 0
        
        if is_return:
            # SALE RETURN → treated as PURCHASE from customer
            # We "buy back" goods at return_tunch/return_rate
            # Profit = (our_cost_basis - return_cost) for both silver & labour
            abs_wt = abs(txn_net_wt)
            abs_total = abs(txn_total)
            
            # Silver profit: difference between our purchase cost (what goods are worth to us)
            # and the return tunch (what we're accepting back at)
            # Positive if return_tunch < purchase_tunch (we accepted cheap)
            silver_profit_grams = (purchase_tunch - txn_tunch) * abs_wt / 100
            silver_profit_kg = silver_profit_grams / 1000
            
            # Labour/Total profit: we refund abs_total, goods cost us purchase_cost * weight
            # Positive if we refunded less than goods cost us
            labour_profit = (purchase_cost_per_gram * abs_wt) - abs_total
            
            # Returns reduce the net sold weight
            customer_profit[customer]['total_sold_kg'] -= abs_wt / 1000
        else:
            # REGULAR SALE → profit = (sale_rate - purchase_cost) * weight
            silver_profit_grams = (txn_tunch - purchase_tunch) * txn_net_wt / 100
            silver_profit_kg = silver_profit_grams / 1000
            
            # Labour/Total profit: we charged txn_total, our cost is purchase_cost * weight
            labour_profit = txn_total - (purchase_cost_per_gram * txn_net_wt)
            
            customer_profit[customer]['total_sold_kg'] += txn_net_wt / 1000
        
        customer_profit[customer]['customer_name'] = customer
        customer_profit[customer]['silver_profit_kg'] += silver_profit_kg
        customer_profit[customer]['labour_profit_inr'] += labour_profit
        customer_profit[customer]['transaction_count'] += 1
    
    # Convert to list, round values, and sort
    for v in customer_profit.values():
        v['silver_profit_kg'] = round(v['silver_profit_kg'], 3)
        v['labour_profit_inr'] = round(v['labour_profit_inr'], 2)
        v['total_sold_kg'] = round(v['total_sold_kg'], 3)
    
    customers = sorted(
        [v for v in customer_profit.values()],
        key=lambda x: x['silver_profit_kg'],
        reverse=True
    )
    
    return {
        "customers": customers,
        "total_customers": len(customers)
    }

@api_router.get("/analytics/supplier-profit")
async def get_supplier_profit(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Calculate profit per supplier based on items they supply"""
    
    query = {}
    if start_date and end_date:
        end_date_with_time = end_date + ' 23:59:59'
        query['date'] = {'$gte': start_date, '$lte': end_date_with_time}
    
    # Get all transactions
    all_transactions = await db.transactions.find(query, {"_id": 0}).to_list(None)
    
    # Group-aware mappings
    all_mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    all_groups = await db.item_groups.find({}, {"_id": 0}).to_list(1000)
    s_mapping_dict, s_member_to_leader, _ = build_group_maps(all_groups, all_mappings)
    
    def _resolve_supplier(name):
        return resolve_to_leader(name, s_mapping_dict, s_member_to_leader)
    
    # Group by supplier: track what each supplier sells us (resolved to leaders)
    supplier_data = defaultdict(lambda: {
        'supplier_name': '',
        'items': defaultdict(lambda: {'purchases': [], 'sales': []}),
        'total_purchased_kg': 0.0,
        'silver_profit_kg': 0.0,
        'labor_profit_inr': 0.0
    })
    
    # Organize transactions — resolve item names to group leaders
    for trans in all_transactions:
        item_name = _resolve_supplier(trans.get('item_name', ''))
        
        if trans['type'] in ['purchase', 'purchase_return']:
            supplier = trans.get('party_name', 'Unknown')
            if supplier:
                supplier_data[supplier]['supplier_name'] = supplier
                supplier_data[supplier]['items'][item_name]['purchases'].append(trans)
                supplier_data[supplier]['total_purchased_kg'] += trans.get('net_wt', 0) / 1000
        
        elif trans['type'] in ['sale', 'sale_return']:
            # Track sales for all items (to calculate profit later)
            for supplier in supplier_data.keys():
                if item_name in supplier_data[supplier]['items']:
                    supplier_data[supplier]['items'][item_name]['sales'].append(trans)
    
    # Calculate profit per supplier
    supplier_profits = []
    
    for supplier, data in supplier_data.items():
        supplier_silver_profit = 0.0
        supplier_labor_profit = 0.0
        total_purchased_from_supplier = 0.0
        
        for item_name, item_data in data['items'].items():
            purchases = item_data['purchases']
            sales = item_data['sales']
            
            # Skip if no purchases or sales
            if not purchases or not sales:
                continue
            
            # Calculate purchase weight from this supplier for this item
            purchase_wt = sum(p.get('net_wt', 0) for p in purchases)
            sale_wt = sum(s.get('net_wt', 0) for s in sales)
            
            if abs(purchase_wt) < 0.001 or abs(sale_wt) < 0.001:
                continue
            
            # Calculate average tunch
            avg_purchase_tunch = sum(float(p.get('tunch', 0) or 0) * abs(p.get('net_wt', 0)) for p in purchases) / sum(abs(p.get('net_wt', 0)) for p in purchases)
            avg_sale_tunch = sum(float(s.get('tunch', 0) or 0) * abs(s.get('net_wt', 0)) for s in sales) / sum(abs(s.get('net_wt', 0)) for s in sales)
            
            # Calculate labour rates
            # Use total_amount as the labour value (in silver trading, Total column = labour Rs)
            purch_total_wt = sum(abs(p.get('net_wt', 0)) for p in purchases)
            if purch_total_wt > 0:
                purchase_labour_per_gram = sum(abs(p.get('total_amount', 0) or p.get('labor', 0)) for p in purchases) / purch_total_wt
            else:
                purchase_labour_per_gram = 0
            
            sale_total_wt = sum(abs(s.get('net_wt', 0)) for s in sales)
            if sale_total_wt > 0:
                sale_labour_per_gram = sum(abs(s.get('total_amount', 0) or s.get('labor', 0)) for s in sales) / sale_total_wt
            else:
                sale_labour_per_gram = 0
            
            # Calculate profit for THIS ITEM based on weight purchased from THIS SUPPLIER
            # Silver profit = (sale_tunch - purchase_tunch) * weight_purchased_from_supplier / 100
            item_silver_profit = (avg_sale_tunch - avg_purchase_tunch) * purchase_wt / 100 / 1000  # Convert to kg
            
            # Labor profit = (sale_labour - purchase_labour) * weight_purchased_from_supplier
            item_labor_profit = (sale_labour_per_gram - purchase_labour_per_gram) * purchase_wt
            
            supplier_silver_profit += item_silver_profit
            supplier_labor_profit += item_labor_profit
            total_purchased_from_supplier += purchase_wt / 1000
        
        if total_purchased_from_supplier > 0:
            supplier_profits.append({
                'supplier_name': supplier,
                'total_purchased_kg': round(total_purchased_from_supplier, 3),
                'silver_profit_kg': round(supplier_silver_profit, 3),
                'labor_profit_inr': round(supplier_labor_profit, 2),
                'items_count': len([k for k, v in data['items'].items() if v['purchases'] and v['sales']])
            })
    
    # Sort by total profit (silver + labor converted to kg equivalent)
    supplier_profits.sort(key=lambda x: x['silver_profit_kg'], reverse=True)
    
    return {
        "suppliers": supplier_profits,
        "total_suppliers": len(supplier_profits)
    }


@api_router.get("/purchase-ledger/all")
async def get_purchase_ledger(current_user: dict = Depends(get_current_user)):
    """Get all purchase rate ledger items"""
    ledger = await db.purchase_ledger.find({}, {"_id": 0}).sort("item_name", 1).to_list(None)
    return ledger

@api_router.get("/mappings/unmapped")
async def get_unmapped_items(current_user: dict = Depends(get_current_user)):
    """Get all unmapped items from transactions AND historical_transactions"""
    # Get item names from both collections
    transactions = await db.transactions.find({}, {"_id": 0, "item_name": 1}).to_list(None)
    historical_names = set()
    async for doc in db.historical_transactions.find({}, {"_id": 0, "item_name": 1}):
        historical_names.add(doc['item_name'])
    trans_names = set(t['item_name'] for t in transactions) | historical_names
    
    # Get all master item names
    master = await db.master_items.find({}, {"_id": 0, "item_name": 1}).to_list(None)
    master_names = set(m['item_name'] for m in master)
    
    # Get existing mappings
    mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    mapped_names = set(m['transaction_name'] for m in mappings)
    
    # Find unmapped: in transactions but not in master and not already mapped
    unmapped = []
    for name in trans_names:
        if name not in master_names and name not in mapped_names:
            # Filter out purely numeric names (e.g., "136" from branch transfer summary lines)
            if name.isdigit():
                continue
            # Filter out test data items
            if name.startswith('TEST_SILVER_ITEM_') or name.startswith('Item ') or name.startswith('Batch'):
                continue
            unmapped.append(name)
    
    return {
        "unmapped_items": sorted(unmapped),
        "count": len(unmapped)
    }

@api_router.post("/mappings/create")
async def create_mapping(transaction_name: str, master_name: str, current_user: dict = Depends(get_current_user)):
    """Create a mapping from transaction name to master name"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    # Verify master name exists
    master = await db.master_items.find_one({"item_name": master_name}, {"_id": 0})
    if not master:
        raise HTTPException(status_code=404, detail="Master item not found")
    
    # Check if mapping already exists
    existing = await db.item_mappings.find_one({"transaction_name": transaction_name})
    if existing:
        # Update existing
        await db.item_mappings.update_one(
            {"transaction_name": transaction_name},
            {"$set": {"master_name": master_name}}
        )
    else:
        # Create new
        mapping = ItemMapping(
            transaction_name=transaction_name,
            master_name=master_name
        )
        await db.item_mappings.insert_one(mapping.model_dump())
    
    return {"success": True, "message": f"Mapped '{transaction_name}' → '{master_name}'"}

@api_router.post("/stamp-verification/save")
async def save_stamp_verification(request: Dict, current_user: dict = Depends(get_current_user)):
    """Save stamp verification record"""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or Manager only")
    stamp = request.get('stamp', '')
    # Normalize stamp format to match master items
    import re as _re
    _match = _re.search(r'(\d+)', stamp)
    if _match:
        stamp = f'STAMP {_match.group(1)}'
    
    physical_gross_wt = request.get('physical_gross_wt', 0)
    book_gross_wt = request.get('book_gross_wt', 0)
    difference = request.get('difference', 0)
    is_match = request.get('is_match', False)
    verification_date = request.get('verification_date', datetime.now(timezone.utc).isoformat()[:10])
    
    # Save to history
    await save_action(
        'stamp_verification',
        f"{stamp} verified: {'MATCH' if is_match else 'MISMATCH'} (Diff: {difference/1000:.3f} kg)",
        {
            'stamp': stamp,
            'physical_gross_wt': physical_gross_wt,
            'book_gross_wt': book_gross_wt,
            'difference': difference,
            'is_match': is_match,
            'verification_date': verification_date
        }
    )
    
    # Save stamp verification record
    verification = {
        'stamp': stamp,
        'physical_gross_wt': physical_gross_wt,
        'book_gross_wt': book_gross_wt,
        'difference': difference,
        'is_match': is_match,
        'verification_date': verification_date,
        'verified_at': datetime.now(timezone.utc).isoformat(),
        'verified_by': current_user['username']
    }
    
    # Update or insert
    await db.stamp_verifications.update_one(
        {'stamp': stamp, 'verification_date': verification_date},
        {'$set': verification},
        upsert=True
    )
    
    # Check if all stamps verified
    total_stamps = await db.master_items.distinct('stamp')
    verified_stamps = await db.stamp_verifications.distinct('stamp', {'verification_date': verification_date})
    
    notification_msg = f"{stamp} {'matched' if is_match else 'mismatched'}"
    
    if len(verified_stamps) >= len(total_stamps):
        notification_msg += " - ALL STAMPS VERIFIED!"
        
        # Create notification for admin
        await db.notifications.insert_one({
            'id': str(uuid.uuid4()),
            'category': 'stamp',
            'type': 'full_stock_match',
            'message': f'Full stock verification complete for {verification_date}',
            'severity': 'success',
            'target_user': 'admin',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'read': False
        })
    else:
        # Individual stamp notification
        await db.notifications.insert_one({
            'id': str(uuid.uuid4()),
            'category': 'stamp',
            'type': 'stamp_verification',
            'message': notification_msg,
            'severity': 'warning' if not is_match else 'info',
            'target_user': 'admin',
            'stamp': stamp,
            'is_match': is_match,
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'read': False
        })
    
    return {
        'success': True,
        'message': notification_msg,
        'verified_stamps': len(verified_stamps),
        'total_stamps': len(total_stamps)
    }

@api_router.get("/stamp-verification/all")
async def get_all_verifications(current_user: dict = Depends(get_current_user)):
    """Get all saved stamp verifications with details"""
    verifications = await db.stamp_verifications.find({}, {"_id": 0}).sort("verified_at", -1).to_list(500)
    for v in verifications:
        v['difference_kg'] = round(v.get('difference', 0) / 1000, 3) if abs(v.get('difference', 0)) > 1 else round(v.get('difference', 0), 3)
    return {"verifications": verifications}

@api_router.delete("/stamp-verification/{stamp}/{verification_date}")
async def delete_stamp_verification(stamp: str, verification_date: str, current_user: dict = Depends(get_current_user)):
    """Delete a stamp verification record (admin/manager only)"""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or Manager only")
    
    # Normalize stamp
    _match = re.search(r'(\d+)', stamp)
    if _match:
        stamp = f'STAMP {_match.group(1)}'
    
    result = await db.stamp_verifications.delete_one({'stamp': stamp, 'verification_date': verification_date})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Verification not found")
    
    await save_action('delete_verification', f"Deleted verification for {stamp} on {verification_date}", user=current_user)
    return {"success": True, "message": f"Verification for {stamp} on {verification_date} deleted"}

@api_router.get("/mappings/all")
async def get_all_mappings(current_user: dict = Depends(get_current_user)):
    """Get all item mappings"""
    mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    return mappings

@api_router.delete("/mappings/{transaction_name}")
async def delete_mapping(transaction_name: str, current_user: dict = Depends(get_current_user)):
    """Delete a mapping"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    result = await db.item_mappings.delete_one({"transaction_name": transaction_name})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Mapping not found")
    return {"success": True, "message": "Mapping deleted"}

@api_router.get("/master-items")
async def get_master_items(search: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get all master items with optional search"""
    query = {}
    if search:
        query = {"item_name": {"$regex": search, "$options": "i"}}
    
    items = await db.master_items.find(query, {"_id": 0}).sort("item_name", 1).to_list(1000)
    
    # Return with cache-control headers to prevent browser caching of stamp data
    return JSONResponse(
        content=items,
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0"
        }
    )

@api_router.get("/analytics/party-analysis")
async def get_party_analysis(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Analyze parties (customers and suppliers) with silver weight comparisons"""
    
    query = {}
    if start_date and end_date:
        end_date_with_time = end_date + ' 23:59:59'
        query['date'] = {'$gte': start_date, '$lte': end_date_with_time}
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(None)
    
    customers = defaultdict(lambda: {
        'party_name': '',
        'total_sales_value': 0.0,
        'total_net_wt': 0.0,
        'total_fine_wt': 0.0,
        'total_gr_wt': 0.0,
        'transaction_count': 0
    })
    
    suppliers = defaultdict(lambda: {
        'party_name': '',
        'total_purchases_value': 0.0,
        'total_net_wt': 0.0,
        'total_fine_wt': 0.0,
        'total_gr_wt': 0.0,
        'transaction_count': 0
    })
    
    for trans in transactions:
        party = trans.get('party_name', 'Unknown')
        if not party:
            continue
        
        amount = trans.get('total_amount', 0)
        net_wt = trans.get('net_wt', 0)
        fine_wt = trans.get('fine', 0)
        gr_wt = trans.get('gr_wt', 0)
        
        if trans['type'] in ['sale', 'sale_return']:
            multiplier = 1 if trans['type'] == 'sale' else -1
            customers[party]['party_name'] = party
            customers[party]['total_sales_value'] += amount * multiplier
            customers[party]['total_net_wt'] += net_wt * multiplier
            customers[party]['total_fine_wt'] += fine_wt * multiplier
            customers[party]['total_gr_wt'] += gr_wt * multiplier
            customers[party]['transaction_count'] += 1
        
        elif trans['type'] in ['purchase', 'purchase_return']:
            multiplier = 1 if trans['type'] == 'purchase' else -1
            suppliers[party]['party_name'] = party
            suppliers[party]['total_purchases_value'] += amount * multiplier
            suppliers[party]['total_net_wt'] += net_wt * multiplier
            suppliers[party]['total_fine_wt'] += fine_wt * multiplier
            suppliers[party]['total_gr_wt'] += gr_wt * multiplier
            suppliers[party]['transaction_count'] += 1
    
    # Convert to lists, round values, and sort by net weight
    for party_data in list(customers.values()) + list(suppliers.values()):
        party_data['total_net_wt'] = round(party_data['total_net_wt'], 3)
        party_data['total_fine_wt'] = round(party_data['total_fine_wt'], 3)
        party_data['total_gr_wt'] = round(party_data['total_gr_wt'], 3)
        party_data['total_sales_value'] = round(party_data.get('total_sales_value', 0), 2)
        party_data['total_purchases_value'] = round(party_data.get('total_purchases_value', 0), 2)
    
    customers_list = sorted(
        [v for v in customers.values()],
        key=lambda x: x['total_net_wt'],
        reverse=True
    )
    
    suppliers_list = sorted(
        [v for v in suppliers.values()],
        key=lambda x: x['total_net_wt'],
        reverse=True
    )
    
    return {
        "customers": customers_list[:50],
        "suppliers": suppliers_list[:50],
        "top_customer": customers_list[0] if customers_list else None,
        "top_supplier": suppliers_list[0] if suppliers_list else None
    }

@api_router.get("/analytics/sales-summary")
async def get_sales_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get total sales summary with net weight, fine weight, and labour (including returns)"""
    
    # Items to exclude
    EXCLUDED_ITEMS = ["SILVER ORNAMENTS"]
    
    query = {"type": {"$in": ["sale", "sale_return"]}}
    if start_date and end_date:
        # Add time component to end_date to include full day
        end_date_with_time = end_date + ' 23:59:59'
        query['date'] = {'$gte': start_date, '$lte': end_date_with_time}
    
    # Get ALL sale transactions (S and SR - SR have negative values already)
    sales_transactions = await db.transactions.find(query, {"_id": 0}).to_list(None)
    
    # Filter out excluded items
    sales_transactions = [t for t in sales_transactions if t['item_name'] not in EXCLUDED_ITEMS]
    
    # Sum including negative values (SR rows are already negative)
    total_net_wt = sum(t.get('net_wt', 0) for t in sales_transactions)
    total_fine_wt = sum(t.get('fine', 0) for t in sales_transactions)
    total_labor = sum((t.get('total_amount', 0) or t.get('labor', 0)) for t in sales_transactions)
    total_sales_value = sum(t.get('total_amount', 0) for t in sales_transactions)
    
    return {
        "total_net_wt_kg": round(total_net_wt / 1000, 3),
        "total_fine_wt_kg": round(total_fine_wt / 1000, 3),
        "total_labor": round(total_labor, 2),
        "total_sales_value": round(total_sales_value, 2),
        "transaction_count": len(sales_transactions)
    }

@api_router.get("/analytics/profit")
async def calculate_profit(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Calculate profit: Silver profit (in kg) and Labour profit (in INR)"""
    
    # Items to exclude from profit calculation
    EXCLUDED_ITEMS = ["SILVER ORNAMENTS", "COURIER", "EMERALD MURTI", "FRAME NEW", "NAJARIA"]
    
    # Get all master items to check for stamps
    master_items = await db.master_items.find({}, {"_id": 0, "item_name": 1, "stamp": 1}).to_list(None)
    master_stamps = {m['item_name']: m.get('stamp', 'Unassigned') for m in master_items}
    
    # Get item mappings + groups for group-aware resolution
    mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    all_groups = await db.item_groups.find({}, {"_id": 0}).to_list(1000)
    p_mapping_dict, p_member_to_leader, _ = build_group_maps(all_groups, mappings)
    
    def _resolve_profit(name):
        return resolve_to_leader(name, p_mapping_dict, p_member_to_leader)
    
    query = {}
    if start_date and end_date:
        # Add time component to end_date to include full day
        end_date_with_time = end_date + ' 23:59:59'
        query['date'] = {'$gte': start_date, '$lte': end_date_with_time}
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(None)
    
    # Filter out excluded items AND items without stamps (unmapped)
    filtered_transactions = []
    for t in transactions:
        trans_name = t['item_name']
        leader_name = _resolve_profit(trans_name)
        
        # Skip if in excluded list
        if leader_name in EXCLUDED_ITEMS:
            continue
        
        # Skip if no stamp or Unassigned
        item_stamp = master_stamps.get(leader_name, master_stamps.get(p_mapping_dict.get(trans_name, trans_name), 'Unassigned'))
        if not item_stamp or item_stamp == 'Unassigned':
            continue
        
        filtered_transactions.append(t)
    
    transactions = filtered_transactions
    
    # Group-aware purchase ledger
    all_ledger = await db.purchase_ledger.find({}, {"_id": 0}).to_list(None)
    grp_ledger = build_group_ledger(all_ledger, all_groups, mappings)
    
    # Group transactions by LEADER item for profit calculation
    item_transactions = defaultdict(lambda: {'purchases': [], 'sales': []})
    
    for trans in transactions:
        raw_name = trans.get('item_name', '')
        if not raw_name:
            continue
        # Resolve to GROUP LEADER for consistent grouping
        item_name = _resolve_profit(raw_name)
        
        trans_data = {
            'date': trans.get('date'),
            'net_wt': trans.get('net_wt', 0),
            'tunch': float(trans.get('tunch', 0) or 0),
            'labor': trans.get('labor', 0),
            'total_amount': trans.get('total_amount', 0)
        }
        
        if trans['type'] in ['purchase', 'purchase_return']:
            item_transactions[item_name]['purchases'].append(trans_data)
        elif trans['type'] == 'sale':
            item_transactions[item_name]['sales'].append(trans_data)
        elif trans['type'] == 'sale_return':
            # Treat sale_return as purchase (buying back from customer)
            item_transactions[item_name]['purchases'].append(trans_data)
    
    # Calculate profits per user's formula
    total_silver_profit_kg = 0.0  # Silver profit in KG
    total_labor_profit_inr = 0.0  # Labour profit in INR
    item_profits = []
    
    for item_name, data in item_transactions.items():
        purchases = data['purchases']
        sales = data['sales']
        
        # Skip if no sales
        if not sales:
            continue
        
        # If no purchases in this period, try to get cost basis from GROUP-AWARE ledger
        if not purchases:
            ledger_item = grp_ledger.get(item_name)
            if ledger_item:
                # Use cumulative purchase data as cost basis
                purchases = [{
                    'net_wt': ledger_item.get('total_purchased_kg', 0) * 1000,  # Convert to grams
                    'tunch': ledger_item.get('purchase_tunch', 0),
                    'labor': ledger_item.get('total_labour', 0),  # Total labour Rs (NOT per-kg rate)
                    'total_amount': ledger_item.get('total_labour', 0)
                }]
            else:
                # No purchase history - skip this item
                continue
        
        # Calculate total and average values
        total_purchase_wt = sum(p['net_wt'] for p in purchases)
        total_sale_wt = sum(s['net_wt'] for s in sales)
        
        if abs(total_purchase_wt) < 0.001 or abs(total_sale_wt) < 0.001:
            continue
        
        # Average tunch weighted by ABSOLUTE net weight (to handle negative returns correctly)
        avg_purchase_tunch = sum(p['tunch'] * abs(p['net_wt']) for p in purchases) / sum(abs(p['net_wt']) for p in purchases) if purchases else 0
        avg_sale_tunch = sum(s['tunch'] * abs(s['net_wt']) for s in sales) / sum(abs(s['net_wt']) for s in sales) if sales else 0
        
        # 1. Silver Profit (kg) = (sale tunch - purchase tunch) * sale net weight / 100
        silver_profit_grams = (avg_sale_tunch - avg_purchase_tunch) * total_sale_wt / 100
        silver_profit_kg = silver_profit_grams / 1000  # Convert to kg
        
        # 2. Labour Profit (INR)
        # In silver trading, 'total_amount' (the "Total" column) IS the labour Rs per line
        # Sale labour = sum of total_amount across all sale transactions for this item
        total_sale_labour = sum(abs(s.get('total_amount', 0) or s.get('labor', 0)) for s in sales)
        
        # Purchase labour cost: ALWAYS prefer purchase ledger (individual txns often have labor=0)
        ledger_item = grp_ledger.get(item_name)
        if ledger_item and ledger_item.get('labour_per_kg', 0) > 0:
            purchase_labour_per_gram = ledger_item['labour_per_kg'] / 1000
        elif purchases and sum(abs(p['net_wt']) for p in purchases) > 0:
            purchase_labour_per_gram = sum(abs(p.get('total_amount', 0) or p.get('labor', 0)) for p in purchases) / sum(abs(p['net_wt']) for p in purchases)
        else:
            purchase_labour_per_gram = 0
        
        # Labour profit = total sale labour income - purchase labour cost for sold quantity
        labor_profit_inr = total_sale_labour - (purchase_labour_per_gram * abs(total_sale_wt))
        
        total_silver_profit_kg += silver_profit_kg
        total_labor_profit_inr += labor_profit_inr
        
        item_profits.append({
            'item_name': item_name,
            'silver_profit_kg': round(silver_profit_kg, 3),
            'labor_profit_inr': round(labor_profit_inr, 2),
            'avg_purchase_tunch': round(avg_purchase_tunch, 2),
            'avg_sale_tunch': round(avg_sale_tunch, 2),
            'net_wt_sold_kg': round(total_sale_wt / 1000, 3)
        })
    
    # Sort by silver profit
    item_profits.sort(key=lambda x: x['silver_profit_kg'], reverse=True)
    
    # Calculate total sales/purchases value
    total_sales_value = sum(t['total_amount'] for t in transactions if t['type'] == 'sale')
    total_purchase_value = sum(t['total_amount'] for t in transactions if t['type'] == 'purchase')
    
    return {
        "silver_profit_kg": round(total_silver_profit_kg, 3),
        "labor_profit_inr": round(total_labor_profit_inr, 2),
        "total_sales_value": round(total_sales_value, 2),
        "total_purchase_value": round(total_purchase_value, 2),
        "all_items": item_profits,
        "total_items_analyzed": len(item_profits)
    }

@api_router.get("/history/actions")
async def get_action_history(limit: int = 20, current_user: dict = Depends(get_current_user)):
    """Get recent actions for undo/redo"""
    actions = await db.action_history.find({}, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)
    return actions

@api_router.post("/history/undo")
async def undo_last_action(current_user: dict = Depends(get_current_user)):
    """Undo last action"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    last_action = await db.action_history.find_one({"can_undo": True}, {"_id": 0}, sort=[("timestamp", -1)])
    
    if not last_action:
        raise HTTPException(status_code=404, detail="No action to undo")
    
    # Mark as undone
    await db.action_history.update_one(
        {"id": last_action['id']},
        {"$set": {"can_undo": False}}
    )
    
    return {
        "success": True,
        "message": f"Undone: {last_action['description']}",
        "action": last_action
    }

@api_router.post("/system/reset")
async def reset_system(request: ResetRequest, current_user: dict = Depends(get_current_user)):
    """Selective system reset with password protection (admin only)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    if request.password != "CLOSE":
        raise HTTPException(status_code=403, detail="Invalid password")
    
    if not request.categories:
        raise HTTPException(status_code=400, detail="No categories selected for reset")
    
    # Map category names to DB collections and transaction type filters
    results = {}
    
    if 'sales' in request.categories:
        r = await db.transactions.delete_many({"type": {"$in": ["sale", "sale_return"]}})
        results['sales'] = r.deleted_count
    
    if 'purchases' in request.categories:
        r = await db.transactions.delete_many({"type": {"$in": ["purchase", "purchase_return"]}})
        results['purchases'] = r.deleted_count
    
    if 'issues' in request.categories:
        r = await db.transactions.delete_many({"type": {"$in": ["issue", "receive"]}})
        results['issues'] = r.deleted_count
    
    if 'polythene' in request.categories:
        r = await db.polythene_adjustments.delete_many({})
        results['polythene'] = r.deleted_count
    
    if 'mappings' in request.categories:
        r = await db.item_mappings.delete_many({})
        results['mappings'] = r.deleted_count
    
    if 'physical_stock' in request.categories:
        r1 = await db.physical_stock.delete_many({})
        r2 = await db.physical_inventory.delete_many({})
        r3 = await db.stock_entries.delete_many({})
        results['physical_stock'] = r1.deleted_count + r2.deleted_count + r3.deleted_count
    
    if 'purchase_ledger' in request.categories:
        r = await db.purchase_ledger.delete_many({})
        results['purchase_ledger'] = r.deleted_count
    
    if 'notifications' in request.categories:
        r1 = await db.notifications.delete_many({})
        r2 = await db.activity_log.delete_many({})
        results['notifications'] = r1.deleted_count + r2.deleted_count
    
    if 'history' in request.categories:
        r = await db.action_history.delete_many({})
        results['history'] = r.deleted_count
    
    if 'master_stock' in request.categories:
        # Zero out quantities but keep items & stamps intact so mappings/groups don't break
        r1 = await db.master_items.update_many({}, {"$set": {"gr_wt": 0, "net_wt": 0}})
        r2 = await db.opening_stock.update_many({}, {"$set": {"gr_wt": 0, "net_wt": 0, "fine": 0, "labor_wt": 0, "labor_rs": 0, "rate": 0, "total": 0, "pc": 0}})
        results['master_stock'] = f"{r1.modified_count} items zeroed, {r2.modified_count} opening stock zeroed"
    
    if 'all_data' in request.categories:
        # Nuclear option: clear everything except users, keep master items structure intact
        for coll in ['transactions', 'polythene_adjustments', 'item_mappings',
                      'physical_inventory', 'physical_stock', 'stock_entries', 'purchase_ledger',
                      'notifications', 'activity_log', 'action_history',
                      'stamp_approvals', 'inventory_snapshots']:
            await db[coll].delete_many({})
        # Zero out master stock quantities but keep items & stamps
        await db.master_items.update_many({}, {"$set": {"gr_wt": 0, "net_wt": 0}})
        await db.opening_stock.update_many({}, {"$set": {"gr_wt": 0, "net_wt": 0, "fine": 0, "labor_wt": 0, "labor_rs": 0, "rate": 0, "total": 0, "pc": 0}})
        results['all_data'] = 'All cleared, master stock zeroed (items/stamps/mappings preserved)'
    
    desc = ', '.join(f"{k}: {v}" for k, v in results.items())
    await save_action('system_reset', f"Selective reset: {desc}")
    
    return {
        "success": True,
        "results": results,
        "message": f"Reset complete: {desc}"
    }


@api_router.post("/system/fix-dates")
async def fix_swapped_dates(current_user: dict = Depends(get_current_user)):
    """Fix dates that were incorrectly stored due to month/day swap bug in normalize_date.
    The bug: pd.to_datetime('YYYY-MM-DD', dayfirst=True) swaps month/day for ISO strings.
    Strategy: detect dates where month/day are both ≤12, and swapping would produce a date
    that fills a gap in the otherwise sequential date stream."""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")

    from datetime import datetime as dt_cls, timedelta
    fixed_count = 0
    fix_log = []

    all_dates = sorted(await db.transactions.distinct('date'))
    if not all_dates:
        return {"success": True, "fixed_count": 0, "fixes": [], "message": "No dates to fix"}

    # Parse all dates
    date_objs = []
    for d in all_dates:
        try:
            date_objs.append((d, dt_cls.strptime(d, '%Y-%m-%d')))
        except ValueError:
            continue

    # Detect suspicious gaps: if sorted dates have a gap > 28 days, something is wrong
    # For each date where month ≤ 12 and day ≤ 12 (swappable), check if swapping
    # produces a date that falls within the expected range
    if len(date_objs) < 2:
        return {"success": True, "fixed_count": 0, "fixes": [], "message": "Not enough dates"}

    # Find dates where day > 12 (unambiguous) to establish the valid date range
    unambiguous = [(d, dt) for d, dt in date_objs if dt.day > 12]
    if unambiguous:
        min_date = min(dt for _, dt in unambiguous) - timedelta(days=15)
        max_date = max(dt for _, dt in unambiguous) + timedelta(days=15)
    else:
        min_date = date_objs[0][1] - timedelta(days=15)
        max_date = date_objs[-1][1]

    # For each swappable date, check if swapping puts it in the valid range
    for date_str, dt in date_objs:
        if dt.month == dt.day:
            continue  # Swap would give same result
        if dt.day > 12 or dt.month > 12:
            continue  # Not swappable
        # Try swapping
        try:
            swapped_dt = dt_cls(dt.year, dt.day, dt.month)
        except ValueError:
            continue
        swapped_str = swapped_dt.strftime('%Y-%m-%d')

        # If current date is outside the valid range but swapped is inside, fix it
        current_in_range = min_date <= dt <= max_date
        swapped_in_range = min_date <= swapped_dt <= max_date

        if not current_in_range and swapped_in_range:
            result = await db.transactions.update_many(
                {'date': date_str},
                {'$set': {'date': swapped_str}}
            )
            if result.modified_count > 0:
                fix_log.append(f"{date_str} → {swapped_str} ({result.modified_count} txns)")
                fixed_count += result.modified_count

    await save_action('fix_dates', f"Fixed {fixed_count} transaction dates (month/day swap correction)")

    return {
        "success": True,
        "fixed_count": fixed_count,
        "fixes": fix_log,
        "message": f"Fixed {fixed_count} transactions across {len(fix_log)} dates"
    }


@api_router.get("/debug/item-closing/{item_name}")
async def debug_item_closing(item_name: str, as_of_date: str = None, current_user: dict = Depends(get_current_user)):
    """Debug endpoint: shows full breakdown of closing stock calculation for an item."""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin/Manager only")

    # Find item in master
    master_item = await db.master_items.find_one({'item_name': item_name}, {'_id': 0})
    if not master_item:
        raise HTTPException(status_code=404, detail=f"Item '{item_name}' not found in master items")

    stamp = master_item.get('stamp', 'Unassigned')

    # Opening stock
    opening_docs = await db.opening_stock.find({'item_name': item_name}, {'_id': 0}).to_list(10)
    opening_gr = sum(d.get('gr_wt', 0) for d in opening_docs)

    # Find all names that map to this item
    mappings = await db.item_mappings.find({'master_name': item_name}, {'_id': 0}).to_list(100)
    all_names = [item_name] + [m['transaction_name'] for m in mappings]

    # Transactions
    query = {'item_name': {'$in': all_names}}
    if as_of_date:
        query['date'] = {'$lte': as_of_date + ' 23:59:59'}

    txns = await db.transactions.find(query, {'_id': 0}).to_list(None)

    from collections import defaultdict
    by_date_type = defaultdict(lambda: defaultdict(float))
    total_in = 0
    total_out = 0
    for t in txns:
        gr = t.get('gr_wt', 0)
        by_date_type[t['date']][t['type']] += gr
        if t['type'] in ['purchase', 'purchase_return', 'receive']:
            total_in += gr
        else:
            total_out += gr

    # Polythene
    poly_query = {'item_name': {'$in': all_names}}
    if as_of_date:
        poly_query['date'] = {'$lte': as_of_date + ' 23:59:59'}
    polythene = await db.polythene_adjustments.find(poly_query, {'_id': 0}).to_list(100)
    poly_total = 0
    for p in polythene:
        pw = p['poly_weight'] * 1000  # kg to grams
        if p['operation'] == 'add':
            poly_total += pw
        else:
            poly_total -= pw

    closing = opening_gr + total_in - total_out + poly_total

    daily_breakdown = []
    for d in sorted(by_date_type.keys()):
        day_data = {'date': d}
        for tp, val in by_date_type[d].items():
            day_data[tp] = round(val / 1000, 3)
        daily_breakdown.append(day_data)

    return {
        'item_name': item_name,
        'stamp': stamp,
        'as_of_date': as_of_date or 'all dates',
        'all_transaction_names': all_names,
        'opening_gr_wt_kg': round(opening_gr / 1000, 3),
        'total_inflow_kg': round(total_in / 1000, 3),
        'total_outflow_kg': round(total_out / 1000, 3),
        'polythene_adjustment_kg': round(poly_total / 1000, 3),
        'closing_gr_wt_kg': round(closing / 1000, 3),
        'transaction_count': len(txns),
        'daily_breakdown': daily_breakdown,
        'polythene_entries': [{
            'date': p.get('date', '?'),
            'operation': p['operation'],
            'weight_kg': p['poly_weight']
        } for p in polythene]
    }



@api_router.get("/stats")
async def get_stats(current_user: dict = Depends(get_current_user)):
    """Get dashboard statistics"""
    total_transactions = await db.transactions.count_documents({})
    total_purchases = await db.transactions.count_documents({"type": "purchase"})
    total_sales = await db.transactions.count_documents({"type": "sale"})
    total_opening_stock = await db.opening_stock.count_documents({})
    
    # Get unique parties
    all_parties = await db.transactions.distinct("party_name")
    total_parties = len([p for p in all_parties if p])
    
    # Get transaction date range
    date_range = {"from_date": None, "to_date": None}
    if total_transactions > 0:
        pipeline = [{"$group": {"_id": None, "min_date": {"$min": "$date"}, "max_date": {"$max": "$date"}}}]
        async for doc in db.transactions.aggregate(pipeline):
            date_range["from_date"] = doc.get("min_date")
            date_range["to_date"] = doc.get("max_date")
    
    # Get unique items count
    all_items = await db.transactions.distinct("item_name")
    total_items = len([i for i in all_items if i])
    
    return {
        "total_transactions": total_transactions,
        "total_purchases": total_purchases,
        "total_sales": total_sales,
        "total_opening_stock": total_opening_stock,
        "total_parties": total_parties,
        "total_items": total_items,
        "date_range": date_range
    }


@api_router.get("/stats/transaction-summary")
async def get_transaction_summary(current_user: dict = Depends(get_current_user)):
    """Get detailed transaction summary with weight totals per type — useful for reconciliation."""
    pipeline = [
        {'$group': {
            '_id': '$type',
            'count': {'$sum': 1},
            'total_gr_wt': {'$sum': '$gr_wt'},
            'total_net_wt': {'$sum': '$net_wt'},
        }},
        {'$sort': {'_id': 1}}
    ]
    results = {}
    async for doc in db.transactions.aggregate(pipeline):
        results[doc['_id']] = {
            'count': doc['count'],
            'gr_wt_kg': round(doc['total_gr_wt'] / 1000, 3),
            'net_wt_kg': round(doc['total_net_wt'] / 1000, 3),
        }

    # Calculate summary totals
    sale_gr = results.get('sale', {}).get('gr_wt_kg', 0)
    sale_ret_gr = results.get('sale_return', {}).get('gr_wt_kg', 0)
    purchase_gr = results.get('purchase', {}).get('gr_wt_kg', 0)
    purchase_ret_gr = results.get('purchase_return', {}).get('gr_wt_kg', 0)
    issue_gr = results.get('issue', {}).get('gr_wt_kg', 0)
    receive_gr = results.get('receive', {}).get('gr_wt_kg', 0)

    return {
        "by_type": results,
        "summary": {
            "net_sale_gr_wt_kg": round(sale_gr + sale_ret_gr, 3),
            "net_purchase_gr_wt_kg": round(purchase_gr + purchase_ret_gr, 3),
            "net_branch_transfer_gr_wt_kg": round(issue_gr - receive_gr, 3),
            "total_outflow_gr_wt_kg": round(sale_gr + sale_ret_gr + issue_gr, 3),
        },
        "dates": sorted(await db.transactions.distinct('date'))
    }

@api_router.delete("/transactions/all")
async def clear_all_transactions(current_user: dict = Depends(get_current_user)):
    """Clear all transactions only (preserves opening stock and master data)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    result = await db.transactions.delete_many({})
    return {"success": True, "deleted_count": result.deleted_count}

@api_router.get("/item/{item_name}")
async def get_item_detail(item_name: str, current_user: dict = Depends(get_current_user)):
    """Get detailed information about a specific item"""
    
    # Get all transactions for this item (and mapped names)
    all_mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    mapping_dict = {m['transaction_name']: m['master_name'] for m in all_mappings}
    reverse_map = defaultdict(list)
    for txn, master in mapping_dict.items():
        reverse_map[master].append(txn)
    
    # Collect all names that map to this item
    search_names = [item_name] + reverse_map.get(item_name, [])
    
    transactions = await db.transactions.find(
        {"item_name": {"$in": search_names}}, 
        {"_id": 0}
    ).sort("date", -1).to_list(1000)
    
    # Calculate statistics
    purchases = [t for t in transactions if t['type'] in ['purchase', 'purchase_return']]
    sales = [t for t in transactions if t['type'] in ['sale', 'sale_return']]
    
    # Weighted average tunch
    abs_purchase_wt = sum(abs(t.get('net_wt', 0)) for t in purchases)
    abs_sale_wt = sum(abs(t.get('net_wt', 0)) for t in sales)
    avg_purchase_tunch = (sum(float(t.get('tunch', 0) or 0) * abs(t.get('net_wt', 0)) for t in purchases) / abs_purchase_wt) if abs_purchase_wt > 0 else 0
    avg_sale_tunch = (sum(float(t.get('tunch', 0) or 0) * abs(t.get('net_wt', 0)) for t in sales) / abs_sale_wt) if abs_sale_wt > 0 else 0
    
    # Weighted average labour per kg — labour charge is in total_amount, not labor field
    def _get_labour(t):
        """Get labour charge from transaction. Uses total_amount (actual labour Rs) not labor field."""
        return float(t.get('total_amount', 0) or t.get('labor', 0) or 0)
    
    avg_purchase_labour = (sum(_get_labour(t) for t in purchases) / (abs_purchase_wt / 1000)) if abs_purchase_wt > 0 else 0
    avg_sale_labour = (sum(_get_labour(t) for t in sales) / (abs_sale_wt / 1000)) if abs_sale_wt > 0 else 0
    
    # Get ACCURATE current stock from get_current_inventory() (matches Current Stock page)
    inv_response = await get_current_inventory()
    current_stock_kg = 0
    current_gr_wt_kg = 0
    item_fine = 0
    item_labor = 0
    # Search in individual items (by_stamp)
    for stamp_items in inv_response.get('by_stamp', {}).values():
        for si in stamp_items:
            if si['item_name'] == item_name:
                current_stock_kg = round(si.get('net_wt', 0) / 1000, 3)
                current_gr_wt_kg = round(si.get('gr_wt', 0) / 1000, 3)
                item_fine = round(si.get('fine', 0) / 1000, 3)
                item_labor = si.get('labor', 0)
                break
    
    # Get purchase ledger info
    ledger = await db.purchase_ledger.find_one({"item_name": item_name}, {"_id": 0})
    has_purchase_rate = ledger is not None
    purchase_tunch_ledger = ledger.get('purchase_tunch', 0) if ledger else 0
    labour_per_kg_ledger = ledger.get('labour_per_kg', 0) if ledger else 0
    
    # Get current stamp from master_items
    master = await db.master_items.find_one({"item_name": item_name}, {"_id": 0})
    current_stamp = master.get('stamp', 'Unassigned') if master else 'Unassigned'
    
    return {
        "item_name": item_name,
        "current_stamp": current_stamp,
        "current_stock_kg": current_stock_kg,
        "current_gr_wt_kg": current_gr_wt_kg,
        "fine_kg": item_fine,
        "labor_value": item_labor,
        "total_purchases": len(purchases),
        "total_sales": len(sales),
        "avg_purchase_tunch": round(avg_purchase_tunch, 2),
        "avg_sale_tunch": round(avg_sale_tunch, 2),
        "tunch_margin": round(avg_sale_tunch - avg_purchase_tunch, 2),
        "avg_purchase_labour": round(avg_purchase_labour, 2),
        "avg_sale_labour": round(avg_sale_labour, 2),
        "labour_margin": round(avg_sale_labour - avg_purchase_labour, 2),
        "has_purchase_rate": has_purchase_rate,
        "purchase_tunch_ledger": purchase_tunch_ledger,
        "labour_per_kg_ledger": labour_per_kg_ledger,
        "recent_transactions": transactions[:20],
    }


@api_router.post("/item/{item_name}/set-purchase-rate")
async def set_purchase_rate(item_name: str, request: Dict, current_user: dict = Depends(get_current_user)):
    """Set or update purchase tunch and labour rate for an item in the purchase ledger.
    Used for items that have no purchase transactions but need fine/labour calculations."""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    purchase_tunch = request.get('purchase_tunch')
    labour_per_kg = request.get('labour_per_kg')
    
    if purchase_tunch is None and labour_per_kg is None:
        raise HTTPException(status_code=400, detail="Provide purchase_tunch and/or labour_per_kg")
    
    update_fields = {'item_name': item_name, 'updated_at': datetime.now(timezone.utc).isoformat()}
    if purchase_tunch is not None:
        update_fields['purchase_tunch'] = float(purchase_tunch)
    if labour_per_kg is not None:
        update_fields['labour_per_kg'] = float(labour_per_kg)
    
    await db.purchase_ledger.update_one(
        {'item_name': item_name},
        {'$set': update_fields},
        upsert=True
    )
    
    _inv_cache.invalidate()
    return {'success': True, 'message': f'Purchase rate updated for {item_name}'}


@api_router.post("/item/{item_name}/assign-stamp")
async def assign_stamp_to_item(item_name: str, stamp: str = Query(...), current_user: dict = Depends(get_current_user)):
    """Assign stamp to all instances of an item"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Normalize stamp to consistent format "STAMP X" (ALL CAPS)
    import re
    if stamp and stamp.lower() != 'unassigned':
        match = re.search(r'(\d+)', stamp)
        if match:
            stamp = f'STAMP {match.group(1)}'
    
    # Update master_items (single source of truth)
    result_master = await db.master_items.update_many(
        {"item_name": item_name},
        {"$set": {"stamp": stamp}}
    )
    
    # Update all transactions
    result1 = await db.transactions.update_many(
        {"item_name": item_name},
        {"$set": {"stamp": stamp}}
    )
    
    # Update opening stock
    result2 = await db.opening_stock.update_many(
        {"item_name": item_name},
        {"$set": {"stamp": stamp}}
    )
    
    await save_action('assign_stamp', f"Assigned stamp '{stamp}' to '{item_name}'")
    
    return {
        "success": True,
        "message": f"Stamp '{stamp}' assigned to '{item_name}'",
        "master_items_updated": result_master.modified_count,
        "transactions_updated": result1.modified_count,
        "opening_stock_updated": result2.modified_count
    }

# ==================== ITEM CATEGORIZATION & BUFFER MANAGEMENT ====================

@api_router.post("/item-buffers/categorize")
async def categorize_items(current_user: dict = Depends(get_current_user)):
    """Rotation-based buffer calculation: 2.73-month stock cycle with seasonal lead times."""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")

    now = datetime.now(timezone.utc)
    season_key = get_current_season(now.month)
    season = SEASON_PROFILES[season_key]
    lead_time_days = season['lead_time_days']
    target_total_stock = season['target_total_stock_kg']

    # 1. Load mappings + groups for item resolution
    all_mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    mapping_dict = {m['transaction_name']: m['master_name'] for m in all_mappings}
    groups = await db.item_groups.find({}, {"_id": 0}).to_list(1000)
    member_to_group = {}
    for g in groups:
        for member in g.get('members', []):
            member_to_group[member] = g['group_name']

    def resolve(name):
        master = mapping_dict.get(name, name)
        return member_to_group.get(master, master)

    # 2. Master items + current inventory
    master_items = await db.master_items.find({}, {"_id": 0}).to_list(None)
    master_dict = {m['item_name']: m for m in master_items}
    inv_response = await get_current_inventory()
    inv_dict = {item['item_name']: item for item in inv_response['inventory']}
    inv_dict.update({item['item_name']: item for item in inv_response.get('negative_items', [])})

    # 3. Aggregate monthly sales (current + historical)
    monthly_pipeline = [
        {"$match": {"type": {"$in": ["sale", "sale_return"]}}},
        {"$project": {"item_name": 1, "net_wt": 1, "month": {"$substr": ["$date", 5, 2]}}},
        {"$group": {"_id": {"item": "$item_name", "month": "$month"}, "total_wt": {"$sum": "$net_wt"}}},
    ]
    item_monthly = defaultdict(lambda: defaultdict(float))
    async for doc in db.transactions.aggregate(monthly_pipeline):
        item = resolve(doc['_id']['item'])
        try:
            m = int(doc['_id']['month'])
        except (ValueError, TypeError):
            continue
        item_monthly[item][m] += abs(doc['total_wt']) / 1000

    async for doc in db.historical_transactions.aggregate(monthly_pipeline):
        item = resolve(doc['_id']['item'])
        try:
            m = int(doc['_id']['month'])
        except (ValueError, TypeError):
            continue
        item_monthly[item][m] += abs(doc['total_wt']) / 1000

    years_with_data = await db.historical_transactions.distinct("historical_year")
    num_years = max(len(years_with_data), 1)

    # 4. Resolve stock per group/item
    group_stock = defaultdict(float)
    group_stamps = {}
    for inv_item in list(inv_response['inventory']) + list(inv_response.get('negative_items', [])):
        name = inv_item['item_name']
        gname = resolve(name)
        group_stock[gname] += inv_item.get('net_wt', 0) / 1000
        if gname not in group_stamps:
            group_stamps[gname] = inv_item.get('stamp', 'Unassigned')
    for name in master_dict:
        gname = resolve(name)
        if gname not in group_stamps:
            group_stamps[gname] = master_dict.get(name, {}).get('stamp', 'Unassigned')

    # 5. Calculate per-item velocities
    all_item_names = set(list(group_stock.keys()) + list(item_monthly.keys()))
    velocities = []
    for gname in all_item_names:
        month_data = item_monthly.get(gname, {})
        # Seasonal velocity: avg monthly sales during current season's months
        season_months = season['months']
        season_total = sum(month_data.get(m, 0) for m in season_months)
        season_velocity = (season_total / num_years) / max(len(season_months), 1)
        # Overall average velocity
        total_all = sum(month_data.values())
        overall_velocity = (total_all / num_years) / 12.0
        # Use higher of seasonal vs overall
        effective_velocity = max(season_velocity, overall_velocity)

        velocities.append({
            'item_name': gname,
            'stamp': group_stamps.get(gname, 'Unassigned'),
            'monthly_velocity_kg': effective_velocity,
            'season_velocity_kg': season_velocity,
            'overall_velocity_kg': overall_velocity,
            'total_sold_kg': total_all / num_years,
            'current_stock_kg': round(group_stock.get(gname, 0), 3),
        })

    # 6. Tier assignment (quartile-based)
    vel_values = [v['monthly_velocity_kg'] for v in velocities if v['monthly_velocity_kg'] > 0]
    if vel_values:
        sorted_vals = sorted(vel_values)
        n = len(sorted_vals)
        q25 = float(sorted_vals[max(0, int(n * 0.25) - 1)])
        q50 = float(sorted_vals[max(0, int(n * 0.50) - 1)])
        q75 = float(sorted_vals[max(0, int(n * 0.75) - 1)])
    else:
        q25 = q50 = q75 = 0

    # Total monthly velocity across all items (for share calculation)
    total_monthly_velocity = sum(v['monthly_velocity_kg'] for v in velocities)

    # 7. Build buffer docs using rotation model
    buffer_docs = []
    group_names_set = {g['group_name'] for g in groups}
    for v in velocities:
        vel = v['monthly_velocity_kg']
        if vel <= 0:
            tier, tier_num = 'dead', 4
        elif vel <= q25:
            tier, tier_num = 'slow', 3
        elif vel <= q50:
            tier, tier_num = 'medium', 2
        elif vel <= q75:
            tier, tier_num = 'fast', 1
        else:
            tier, tier_num = 'fastest', 0

        # --- Core rotation-based calculation ---
        # Minimum stock = 2.73 months of sales (full rotation cycle worth)
        minimum_stock = round(vel * ROTATION_CYCLE_MONTHS, 3)

        # Reorder buffer = stock consumed during order lead time
        daily_velocity = vel / 30.0
        reorder_buffer = round(daily_velocity * lead_time_days, 3)

        # Upper buffer (target) = item's proportional share of target total stock
        # This is the aspirational level to push sales
        if total_monthly_velocity > 0:
            item_share = vel / total_monthly_velocity
            upper_target = round(item_share * target_total_stock, 3)
        else:
            upper_target = minimum_stock

        # Upper target should be at least minimum_stock
        upper_target = max(upper_target, minimum_stock)

        current_stock_kg = v['current_stock_kg']

        # Status logic:
        # red = below reorder buffer (CRITICAL - will run out before order arrives)
        # yellow = below minimum stock (needs restocking soon)
        # green = at or above minimum stock
        if vel <= 0:
            status = 'green'  # dead items — no concern
        elif current_stock_kg < reorder_buffer:
            status = 'red'
        elif current_stock_kg < minimum_stock:
            status = 'yellow'
        else:
            status = 'green'

        # season_boost = ratio of seasonal velocity to overall velocity (>1 means seasonal demand is higher)
        season_boost = round(v['season_velocity_kg'] / v['overall_velocity_kg'], 2) if v['overall_velocity_kg'] > 0 else 1.0

        buffer_docs.append({
            'item_name': v['item_name'],
            'stamp': v['stamp'],
            'tier': tier, 'tier_num': tier_num,
            'monthly_velocity_kg': round(vel, 3),
            'season_velocity_kg': round(v['season_velocity_kg'], 3),
            'overall_velocity_kg': round(v['overall_velocity_kg'], 3),
            'season_boost': season_boost,
            'total_sold_kg': round(v['total_sold_kg'], 3),
            'minimum_stock_kg': minimum_stock,
            'reorder_buffer_kg': reorder_buffer,
            'upper_target_kg': upper_target,
            'lead_time_days': lead_time_days,
            'current_stock_kg': current_stock_kg,
            'status': status,
            'current_season': season_key,
            'season_label': season['label'],
            'is_group': v['item_name'] in group_names_set,
            'updated_at': datetime.now(timezone.utc).isoformat()
        })

    # Save to DB
    new_names = {doc['item_name'] for doc in buffer_docs}
    await db.item_buffers.delete_many({'item_name': {'$nin': list(new_names)}})
    for doc in buffer_docs:
        await db.item_buffers.update_one(
            {'item_name': doc['item_name']},
            {'$set': doc},
            upsert=True
        )

    tier_counts = defaultdict(int)
    for d in buffer_docs:
        tier_counts[d['tier']] += 1

    total_current_stock = round(sum(v['current_stock_kg'] for v in velocities), 2)
    await save_action('categorize_items', f"Categorized {len(buffer_docs)} items | Season: {season_key} | Lead: {lead_time_days}d | Total stock: {total_current_stock} kg", user=current_user)

    return {
        "success": True,
        "total_items": len(buffer_docs),
        "tiers": dict(tier_counts),
        "thresholds": {"q25": round(q25, 3), "q50": round(q50, 3), "q75": round(q75, 3)},
        "current_season": season_key,
        "season_label": season['label'],
        "lead_time_days": lead_time_days,
        "target_total_stock_kg": target_total_stock,
        "total_current_stock_kg": total_current_stock,
        "rotation_months": ROTATION_CYCLE_MONTHS,
        "years_analyzed": num_years
    }

@api_router.get("/item-buffers")
async def get_item_buffers(
    stamp: Optional[str] = None,
    tier: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all item buffer configurations with optional filters"""
    query = {}
    if stamp:
        query['stamp'] = stamp
    if tier:
        query['tier'] = tier
    if status:
        query['status'] = status
    
    items = await db.item_buffers.find(query, {"_id": 0}).sort("tier_num", 1).to_list(None)
    
    # Refresh current stock and status using INDIVIDUAL item data (not group totals)
    inv_response = await get_current_inventory()
    # Build dict from by_stamp (individual level) for accurate per-item stock
    inv_dict = {}
    for stamp_items in inv_response.get('by_stamp', {}).values():
        for si in stamp_items:
            inv_dict[si['item_name']] = si
    # Also include group-level entries for items that are group leaders
    for item in inv_response.get('inventory', []) + inv_response.get('negative_items', []):
        if item['item_name'] not in inv_dict:
            inv_dict[item['item_name']] = item
    
    for item in items:
        inv_item = inv_dict.get(item['item_name'])
        current = round(inv_item['net_wt'] / 1000, 3) if inv_item else 0
        item['current_stock_kg'] = current
        min_stock = item.get('minimum_stock_kg', 0)
        reorder = item.get('reorder_buffer_kg', 0)
        vel = item.get('monthly_velocity_kg', 0)
        
        if vel <= 0:
            item['status'] = 'green'
        elif current < reorder:
            item['status'] = 'red'
        elif current < min_stock:
            item['status'] = 'yellow'
        else:
            item['status'] = 'green'
    
    return {"items": items, "total": len(items)}

@api_router.put("/item-buffers/{item_name}")
async def update_item_buffer(item_name: str, minimum_stock_kg: float = Query(...), current_user: dict = Depends(get_current_user)):
    """Update minimum stock for an item"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    result = await db.item_buffers.update_one(
        {'item_name': item_name},
        {'$set': {'minimum_stock_kg': round(minimum_stock_kg, 3), 'updated_at': datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found in buffers")
    return {"success": True, "message": f"Minimum stock for '{item_name}' set to {minimum_stock_kg} kg"}

# ==================== ITEM GROUPS (merge similar items) ====================

@api_router.get("/item-groups")
async def get_item_groups(current_user: dict = Depends(get_current_user)):
    """Get all item groups with their members and mapped items"""
    groups = await db.item_groups.find({}, {"_id": 0}).to_list(1000)
    # Also get item mappings to show which items map to each member
    mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    mapping_by_master = defaultdict(list)
    for m in mappings:
        mapping_by_master[m['master_name']].append(m['transaction_name'])
    for g in groups:
        g['mapped_items'] = {}
        for member in g.get('members', []):
            g['mapped_items'][member] = mapping_by_master.get(member, [])
    return {"groups": groups}


@api_router.post("/item-groups")
async def save_item_group(group: ItemGroup, current_user: dict = Depends(get_current_user)):
    """Create or update an item group"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    if len(group.members) < 2:
        raise HTTPException(status_code=400, detail="Group needs at least 2 members")
    await db.item_groups.update_one(
        {'group_name': group.group_name},
        {'$set': {'group_name': group.group_name, 'members': group.members,
                  'updated_at': datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"success": True, "message": f"Group '{group.group_name}' saved with {len(group.members)} members"}


@api_router.delete("/item-groups/{group_name}")
async def delete_item_group(group_name: str, current_user: dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    await db.item_groups.delete_one({'group_name': group_name})
    return {"success": True}


@api_router.get("/item-groups/suggestions")
async def suggest_item_groups(current_user: dict = Depends(get_current_user)):
    """List all master items + auto-detected groups from mappings"""
    items = await db.master_items.find({}, {"_id": 0, "item_name": 1, "stamp": 1}).to_list(None)
    existing = await db.item_groups.find({}, {"_id": 0}).to_list(1000)
    grouped_items = set()
    for g in existing:
        grouped_items.update(g.get('members', []))

    # Auto-detect: master items that have mappings pointing to them
    mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    master_names = {i['item_name'] for i in items}
    mapping_by_master = defaultdict(list)
    for m in mappings:
        mapping_by_master[m['master_name']].append(m['transaction_name'])
    # Items that have transaction names mapping to them AND those names are also master items
    auto_suggestions = []
    for master, txn_names in mapping_by_master.items():
        related_masters = [t for t in txn_names if t in master_names and t != master]
        if related_masters:
            auto_suggestions.append({
                'leader': master,
                'members': [master] + related_masters
            })

    return {"items": items, "already_grouped": list(grouped_items), "auto_suggestions": auto_suggestions}


# ==================== STAMP DETAIL (click a stamp to see items + assign) ====================

@api_router.get("/stamps/{stamp_name}/detail")
async def get_stamp_detail(stamp_name: str, current_user: dict = Depends(get_current_user)):
    """Get all items in a stamp with stock info (per-stamp, not grouped)"""
    master_items = await db.master_items.find(
        {"stamp": stamp_name}, {"_id": 0}
    ).to_list(None)
    inv_response = await get_current_inventory()
    
    # Use stamp_items (ungrouped, per-stamp) for correct per-stamp lookup
    inv_dict = {}
    for item in inv_response.get('stamp_items', inv_response.get('inventory', [])):
        inv_dict[item['item_name']] = item

    items_with_stock = []
    total_net_wt = 0
    for mi in master_items:
        inv = inv_dict.get(mi['item_name'])
        net_wt = round(inv['net_wt'] / 1000, 3) if inv else 0
        total_net_wt += net_wt
        items_with_stock.append({
            'item_name': mi['item_name'],
            'net_wt_kg': net_wt,
            'gr_wt_kg': round(inv['gr_wt'] / 1000, 3) if inv else 0,
        })
    items_with_stock.sort(key=lambda x: x['net_wt_kg'], reverse=True)

    assignment = await db.stamp_assignments.find_one({"stamp": stamp_name}, {"_id": 0})
    return {
        "stamp": stamp_name,
        "items": items_with_stock,
        "total_items": len(items_with_stock),
        "total_net_wt_kg": round(total_net_wt, 3),
        "assigned_user": assignment.get('assigned_user') if assignment else None
    }


# ==================== STAMP-USER ASSIGNMENT ====================

@api_router.get("/stamp-assignments")
async def get_stamp_assignments(current_user: dict = Depends(get_current_user)):
    """Get all stamp-to-user assignments"""
    assignments = await db.stamp_assignments.find({}, {"_id": 0}).to_list(100)
    return {"assignments": assignments}

@api_router.post("/stamp-assignments")
async def save_stamp_assignment(assignment: StampAssignment, current_user: dict = Depends(get_current_user)):
    """Assign a user to a stamp for notifications"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    await db.stamp_assignments.update_one(
        {'stamp': assignment.stamp},
        {'$set': {'stamp': assignment.stamp, 'assigned_user': assignment.assigned_user, 'updated_at': datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"success": True, "message": f"User '{assignment.assigned_user}' assigned to '{assignment.stamp}'"}

@api_router.delete("/stamp-assignments/{stamp}")
async def delete_stamp_assignment(stamp: str, current_user: dict = Depends(get_current_user)):
    """Remove stamp assignment"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    await db.stamp_assignments.delete_one({'stamp': stamp})
    return {"success": True}

# ==================== ORDER MANAGEMENT ====================

@api_router.post("/orders/create")
async def create_order(order: OrderCreate, current_user: dict = Depends(get_current_user)):
    """Create a restock order"""
    # Get buffer info
    buffer_info = await db.item_buffers.find_one({'item_name': order.item_name}, {"_id": 0})
    
    order_doc = {
        'id': str(uuid.uuid4()),
        'item_name': order.item_name,
        'quantity_kg': round(order.quantity_kg, 3),
        'supplier': order.supplier,
        'notes': order.notes,
        'status': 'ordered',
        'ordered_by': current_user['username'],
        'ordered_at': datetime.now(timezone.utc).isoformat(),
        'received_at': None,
        'verified': False,
        'stamp': buffer_info.get('stamp', '') if buffer_info else '',
        'tier': buffer_info.get('tier', '') if buffer_info else ''
    }
    
    await db.orders.insert_one(order_doc)
    
    # Notify admin
    await db.notifications.insert_one({
        'id': str(uuid.uuid4()),
        'category': 'order',
        'type': 'order_placed',
        'message': f"Order placed: {order.quantity_kg} kg of '{order.item_name}' by {current_user['username']}",
        'item_name': order.item_name,
        'target_user': 'admin',
        'read': False,
        'timestamp': datetime.now(timezone.utc).isoformat()
    })
    
    return {"success": True, "order_id": order_doc['id'], "message": "Order created"}

@api_router.get("/orders")
async def get_orders(status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get all orders (admin/manager see all, others see only their own)"""
    query = {}
    if status:
        query['status'] = status
    if current_user['role'] not in ['admin', 'manager']:
        query['ordered_by'] = current_user['username']
    orders = await db.orders.find(query, {"_id": 0}).sort("ordered_at", -1).to_list(1000)
    return {"orders": orders}

@api_router.put("/orders/{order_id}/received")
async def mark_order_received(order_id: str, current_user: dict = Depends(get_current_user)):
    """Mark order as received (admin/manager only)"""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or Manager only")
    result = await db.orders.update_one(
        {'id': order_id},
        {'$set': {
            'status': 'received',
            'received_at': datetime.now(timezone.utc).isoformat(),
            'received_by': current_user['username'],
            'verified': True
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Notify admin that order was received
    order = await db.orders.find_one({'id': order_id}, {"_id": 0})
    if order:
        await db.notifications.insert_one({
            'id': str(uuid.uuid4()),
            'category': 'order',
            'type': 'order_received',
            'message': f"Order received: {order.get('quantity_kg')} kg of '{order.get('item_name')}' by {current_user['username']}",
            'item_name': order.get('item_name'),
            'target_user': 'admin',
            'read': False,
            'timestamp': datetime.now(timezone.utc).isoformat()
        })
    
    return {"success": True, "message": "Order marked as received"}

@api_router.delete("/orders/{order_id}")
async def cancel_order(order_id: str, current_user: dict = Depends(get_current_user)):
    """Cancel/delete an order"""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or Manager only")
    result = await db.orders.delete_one({'id': order_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return {"success": True, "message": "Order cancelled"}

@api_router.get("/orders/overdue")
async def check_overdue_orders(current_user: dict = Depends(get_current_user)):
    """Check for orders that are overdue (admin/manager only)"""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Admin or Manager only")
    cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    overdue = await db.orders.find({
        'status': 'ordered',
        'ordered_at': {'$lt': cutoff}
    }, {"_id": 0}).to_list(100)
    
    # Generate notifications for overdue orders
    for order in overdue:
        if not order.get('overdue_notified'):
            days_ago = (datetime.now(timezone.utc) - datetime.fromisoformat(order['ordered_at'].replace('Z', '+00:00'))).days
            await db.notifications.insert_one({
                'id': str(uuid.uuid4()),
                'category': 'order',
                'type': 'order_overdue',
                'severity': 'warning',
                'message': f"OVERDUE: Order for {order.get('quantity_kg')} kg of '{order.get('item_name')}' placed {days_ago} days ago not received",
                'item_name': order.get('item_name'),
                'target_user': 'admin',
                'read': False,
                'timestamp': datetime.now(timezone.utc).isoformat()
            })
            await db.orders.update_one({'id': order['id']}, {'$set': {'overdue_notified': True}})
    
    return {"overdue_orders": overdue, "count": len(overdue)}

# ==================== ENHANCED NOTIFICATIONS ====================

@api_router.get("/notifications/categorized")
async def get_categorized_notifications(current_user: dict = Depends(get_current_user)):
    """Get notifications organized by category"""
    username = current_user['username']
    role = current_user['role']
    query = {"$or": [
        {"target_user": username},
        {"target_user": role},
        {"target_user": "all"},
        {"for_role": role},
    ]}
    if role == 'admin':
        query["$or"].append({"target_user": "admin"})
        query["$or"].append({"for_role": "admin"})
    
    all_notifs = await db.notifications.find(query, {"_id": 0}).sort("timestamp", -1).to_list(500)
    
    categorized = {
        'stock': [n for n in all_notifs if n.get('category') == 'stock'],
        'order': [n for n in all_notifs if n.get('category') == 'order'],
        'stamp': [n for n in all_notifs if n.get('category') == 'stamp'],
        'polythene': [n for n in all_notifs if n.get('category') == 'polythene'],
        'general': [n for n in all_notifs if n.get('category', 'general') == 'general' or not n.get('category')]
    }
    
    unread = sum(1 for n in all_notifs if not n.get('read'))
    
    return {"notifications": categorized, "total_unread": unread}

@api_router.post("/notifications/check-stock-alerts")
async def check_stock_alerts(current_user: dict = Depends(get_current_user)):
    """Check all items and generate stock deficit/excess notifications"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    buffers = await db.item_buffers.find({}, {"_id": 0}).to_list(None)
    if not buffers:
        return {"success": True, "alerts_generated": 0, "message": "No buffer data. Run categorization first."}
    
    inv_response = await get_current_inventory()
    inv_dict = {item['item_name']: item for item in inv_response['inventory']}
    inv_dict.update({item['item_name']: item for item in inv_response.get('negative_items', [])})
    
    # Get stamp assignments
    assignments = await db.stamp_assignments.find({}, {"_id": 0}).to_list(100)
    stamp_user = {a['stamp']: a['assigned_user'] for a in assignments}
    
    alerts = 0
    now = datetime.now(timezone.utc).isoformat()
    
    for buf in buffers:
        inv_item = inv_dict.get(buf['item_name'])
        current = round(inv_item['net_wt'] / 1000, 3) if inv_item else 0
        min_stock = buf.get('minimum_stock_kg', 0)
        reorder = buf.get('reorder_buffer_kg', 0)
        upper = buf.get('upper_target_kg', 0)
        stamp = buf.get('stamp', '')
        
        target = stamp_user.get(stamp, 'admin')
        
        if min_stock > 0 and current < min_stock:
            deficit = round(min_stock - current, 3)
            severity = 'critical' if current < reorder else 'warning'
            
            # Notify assigned SEE
            await db.notifications.insert_one({
                'id': str(uuid.uuid4()),
                'category': 'stock',
                'type': 'stock_deficit',
                'severity': severity,
                'message': f"LOW STOCK: '{buf['item_name']}' at {current} kg (min: {min_stock} kg, deficit: {deficit} kg)",
                'item_name': buf['item_name'],
                'stamp': stamp,
                'current_stock': current,
                'minimum_stock': min_stock,
                'deficit': deficit,
                'order_range_min': round(deficit, 3),
                'order_range_max': round(upper - current, 3) if upper > current else round(deficit, 3),
                'target_user': target,
                'read': False,
                'timestamp': now
            })
            
            # If critical, also notify admin
            if severity == 'critical' and target != 'admin':
                await db.notifications.insert_one({
                    'id': str(uuid.uuid4()),
                    'category': 'stock',
                    'type': 'stock_deficit',
                    'severity': 'critical',
                    'message': f"CRITICAL: '{buf['item_name']}' at {current} kg (below buffer!)",
                    'item_name': buf['item_name'],
                    'stamp': stamp,
                    'target_user': 'admin',
                    'read': False,
                    'timestamp': now
                })
            
            alerts += 1
        
        elif upper > 0 and current > upper:
            excess = round(current - upper, 3)
            await db.notifications.insert_one({
                'id': str(uuid.uuid4()),
                'category': 'stock',
                'type': 'stock_excess',
                'severity': 'info',
                'message': f"EXCESS: '{buf['item_name']}' at {current} kg (upper: {upper} kg, excess: {excess} kg)",
                'item_name': buf['item_name'],
                'stamp': stamp,
                'target_user': target,
                'read': False,
                'timestamp': now
            })
            alerts += 1
    
    return {"success": True, "alerts_generated": alerts}

@api_router.get("/stock-alerts/auto")
async def auto_stock_alerts(current_user: dict = Depends(get_current_user)):
    """Lightweight auto-check: returns current stock alerts for the user without generating new notifications.
    Only regenerates alerts if last check was > 30 minutes ago."""
    
    username = current_user['username']
    role = current_user['role']
    
    # Check if we need to regenerate alerts (throttle to every 30 min)
    last_check = await db.system_state.find_one({'key': 'last_stock_alert_check'}, {'_id': 0})
    now = datetime.now(timezone.utc)
    should_regenerate = True
    
    if last_check and last_check.get('timestamp'):
        try:
            last_ts = datetime.fromisoformat(last_check['timestamp'])
            if (now - last_ts).total_seconds() < 1800:  # 30 minutes
                should_regenerate = False
        except:
            pass
    
    if should_regenerate:
        # Run the stock alert check
        buffers = await db.item_buffers.find({}, {"_id": 0}).to_list(None)
        if buffers:
            inv_response = await get_current_inventory()
            inv_dict = {item['item_name']: item for item in inv_response['inventory']}
            inv_dict.update({item['item_name']: item for item in inv_response.get('negative_items', [])})
            
            assignments = await db.stamp_assignments.find({}, {"_id": 0}).to_list(100)
            stamp_user = {a['stamp']: a['assigned_user'] for a in assignments}
            
            # Clear old stock alerts (only stock category)
            await db.notifications.delete_many({'category': 'stock', 'type': {'$in': ['stock_deficit', 'stock_excess']}})
            
            now_str = now.isoformat()
            for buf in buffers:
                inv_item = inv_dict.get(buf['item_name'])
                current = round(inv_item['net_wt'] / 1000, 3) if inv_item else 0
                min_stock = buf.get('minimum_stock_kg', 0)
                reorder = buf.get('reorder_buffer_kg', 0)
                upper = buf.get('upper_target_kg', 0)
                stamp = buf.get('stamp', '')
                target = stamp_user.get(stamp, 'admin')
                
                if min_stock > 0 and current < min_stock:
                    deficit = round(min_stock - current, 3)
                    severity = 'critical' if current < reorder else 'warning'
                    
                    await db.notifications.insert_one({
                        'id': str(uuid.uuid4()), 'category': 'stock', 'type': 'stock_deficit',
                        'severity': severity,
                        'message': f"LOW STOCK: '{buf['item_name']}' at {current} kg (min: {min_stock} kg)",
                        'item_name': buf['item_name'], 'stamp': stamp,
                        'current_stock': current, 'minimum_stock': min_stock, 'deficit': deficit,
                        'order_range_min': round(deficit, 3),
                        'order_range_max': round(upper - current, 3) if upper > current else round(deficit, 3),
                        'target_user': target, 'read': False, 'timestamp': now_str
                    })
                    
                    if severity == 'critical' and target != 'admin':
                        await db.notifications.insert_one({
                            'id': str(uuid.uuid4()), 'category': 'stock', 'type': 'stock_deficit',
                            'severity': 'critical',
                            'message': f"CRITICAL: '{buf['item_name']}' at {current} kg (below buffer!)",
                            'item_name': buf['item_name'], 'stamp': stamp,
                            'target_user': 'admin', 'read': False, 'timestamp': now_str
                        })
            
            await db.system_state.update_one(
                {'key': 'last_stock_alert_check'},
                {'$set': {'key': 'last_stock_alert_check', 'timestamp': now.isoformat()}},
                upsert=True
            )
    
    # Return alerts relevant to this user
    query = {'category': 'stock', 'type': 'stock_deficit', 'read': False}
    if role != 'admin':
        query['target_user'] = username
    
    alerts = await db.notifications.find(query, {"_id": 0}).sort("severity", 1).to_list(100)
    
    return {"alerts": alerts, "count": len(alerts)}

# ==================== HISTORICAL PROFIT ANALYSIS (Aggregation-based, scales to 1M+ txns) ====================

@api_router.get("/analytics/historical-profit")
async def get_historical_profit(
    year: Optional[str] = None,
    view: str = "yearly",
    current_user: dict = Depends(get_current_user)
):
    """
    Profit analysis from historical_transactions using MongoDB aggregation.
    NEVER loads raw documents into Python — all heavy lifting done in MongoDB.
    Scales comfortably to 500k+ transactions per year.
    """
    match_filter = {}
    if year:
        match_filter["historical_year"] = year

    # 1. Load item mappings + groups for group-aware resolution
    all_mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    all_groups = await db.item_groups.find({}, {"_id": 0}).to_list(1000)
    h_mapping_dict, h_member_to_leader, _ = build_group_maps(all_groups, all_mappings)
    def resolve(name):
        return resolve_to_leader(name, h_mapping_dict, h_member_to_leader)

    # 2. Purchase basis via aggregation (grouped by item_name)
    purchase_agg = await db.historical_transactions.aggregate([
        {"$match": {**match_filter, "type": "purchase", "net_wt": {"$gt": 0.001}}},
        {"$group": {
            "_id": "$item_name",
            "fine": {"$sum": "$fine"}, "net_wt": {"$sum": "$net_wt"},
            "labor": {"$sum": "$labor"}, "total_amount": {"$sum": "$total_amount"}, "gr_wt": {"$sum": "$gr_wt"}
        }}
    ]).to_list(None)

    # Merge into master items using mappings
    purchase_basis = {}
    for doc in purchase_agg:
        master = resolve(doc["_id"])
        if master not in purchase_basis:
            purchase_basis[master] = {"fine": 0.0, "net_wt": 0.0, "labor": 0.0}
        purchase_basis[master]["fine"] += doc["fine"] or 0
        purchase_basis[master]["net_wt"] += doc["net_wt"] or 0
        purchase_basis[master]["labor"] += (doc.get("total_amount") or doc.get("labor") or 0)
    del purchase_agg

    # Compute avg tunch and labor/g per master item
    for item, d in purchase_basis.items():
        nw = d["net_wt"]
        if nw > 0.001:
            d["avg_tunch"] = (d["fine"] / nw) * 100 if d["fine"] > 0 else 0
            d["labor_per_gram"] = d["labor"] / nw
        else:
            d["avg_tunch"] = 0
            d["labor_per_gram"] = 0

    # === VIEW: YEARLY ===
    if view == "yearly":
        sale_item_agg = await db.historical_transactions.aggregate([
            {"$match": {**match_filter, "type": "sale", "net_wt": {"$gt": 0.001}}},
            {"$group": {
                "_id": "$item_name",
                "fine": {"$sum": "$fine"}, "net_wt": {"$sum": "$net_wt"},
                "labor": {"$sum": "$labor"}, "total_amount": {"$sum": "$total_amount"},
                "count": {"$sum": 1}
            }}
        ]).to_list(None)

        silver_kg = 0.0; labor_inr = 0.0; total_wt = 0.0; matched = 0
        total_sale_val = 0.0; total_sale_recs = 0
        for doc in sale_item_agg:
            master = resolve(doc["_id"])
            basis = purchase_basis.get(master)
            nw = doc["net_wt"] or 0
            total_sale_val += doc.get("total_amount") or 0
            total_sale_recs += doc["count"]
            if not basis or nw < 0.001:
                continue
            s_tunch = (doc["fine"] / nw) * 100 if (doc["fine"] or 0) > 0 else 0
            s_lpg = (doc.get("total_amount") or doc.get("labor") or 0) / nw
            silver_kg += (s_tunch - basis["avg_tunch"]) * nw / 100 / 1000
            labor_inr += (s_lpg - basis["labor_per_gram"]) * nw
            total_wt += nw / 1000
            matched += doc["count"]

        # Purchase totals
        purch_stats = await db.historical_transactions.aggregate([
            {"$match": {**match_filter, "type": "purchase"}},
            {"$group": {"_id": None, "total_amount": {"$sum": "$total_amount"}, "count": {"$sum": 1}}}
        ]).to_list(1)
        p_val = purch_stats[0]["total_amount"] if purch_stats else 0
        p_cnt = purch_stats[0]["count"] if purch_stats else 0

        return {
            "view": "yearly", "year": year,
            "silver_profit_kg": round(silver_kg, 3), "labor_profit_inr": round(labor_inr, 2),
            "total_sold_kg": round(total_wt, 3), "total_transactions": matched,
            "total_sales_value": round(total_sale_val, 2), "total_purchase_value": round(p_val, 2),
            "total_sale_records": total_sale_recs, "total_purchase_records": p_cnt,
        }

    # === VIEW: CUSTOMER ===
    if view == "customer":
        cust_agg = await db.historical_transactions.aggregate([
            {"$match": {**match_filter, "type": "sale", "net_wt": {"$gt": 0.001}}},
            {"$group": {
                "_id": {"party": "$party_name", "item": "$item_name"},
                "fine": {"$sum": "$fine"}, "net_wt": {"$sum": "$net_wt"},
                "labor": {"$sum": "$labor"}, "total_amount": {"$sum": "$total_amount"}, "count": {"$sum": 1}
            }}
        ]).to_list(None)

        cust_profit = defaultdict(lambda: {"silver": 0.0, "labor": 0.0, "wt": 0.0, "cnt": 0})
        for doc in cust_agg:
            party = doc["_id"]["party"] or "Unknown"
            master = resolve(doc["_id"]["item"])
            basis = purchase_basis.get(master)
            nw = doc["net_wt"] or 0
            if not basis or nw < 0.001:
                continue
            s_tunch = (doc["fine"] / nw) * 100 if (doc["fine"] or 0) > 0 else 0
            s_lpg = (doc.get("total_amount") or doc.get("labor") or 0) / nw
            cust_profit[party]["silver"] += (s_tunch - basis["avg_tunch"]) * nw / 100 / 1000
            cust_profit[party]["labor"] += (s_lpg - basis["labor_per_gram"]) * nw
            cust_profit[party]["wt"] += nw / 1000
            cust_profit[party]["cnt"] += doc["count"]

        rows = [{"name": k, "silver_profit_kg": round(v["silver"], 3),
                 "labor_profit_inr": round(v["labor"], 2),
                 "total_sold_kg": round(v["wt"], 3), "transactions": v["cnt"]}
                for k, v in cust_profit.items() if v["cnt"] > 0]
        rows.sort(key=lambda x: x["silver_profit_kg"], reverse=True)
        return {"view": "customer", "year": year, "data": rows, "total": len(rows)}

    # === VIEW: SUPPLIER ===
    if view == "supplier":
        # Purchase grouped by supplier + item (via aggregation)
        sup_agg = await db.historical_transactions.aggregate([
            {"$match": {**match_filter, "type": "purchase", "net_wt": {"$gt": 0.001}}},
            {"$group": {
                "_id": {"party": "$party_name", "item": "$item_name"},
                "fine": {"$sum": "$fine"}, "net_wt": {"$sum": "$net_wt"},
                "labor": {"$sum": "$labor"}
            }}
        ]).to_list(None)

        # Global sale averages per master item (via aggregation)
        sale_avg_agg = await db.historical_transactions.aggregate([
            {"$match": {**match_filter, "type": "sale", "net_wt": {"$gt": 0.001}}},
            {"$group": {
                "_id": "$item_name",
                "fine": {"$sum": "$fine"}, "net_wt": {"$sum": "$net_wt"},
                "labor": {"$sum": "$labor"}, "total_amount": {"$sum": "$total_amount"}
            }}
        ]).to_list(None)

        # Merge sale averages by master item
        item_sale_avg = {}
        sale_merged = defaultdict(lambda: {"fine": 0.0, "net_wt": 0.0, "labor": 0.0, "total_amount": 0.0})
        for doc in sale_avg_agg:
            master = resolve(doc["_id"])
            sale_merged[master]["fine"] += doc["fine"] or 0
            sale_merged[master]["net_wt"] += doc["net_wt"] or 0
            sale_merged[master]["labor"] += doc["labor"] or 0
            sale_merged[master]["total_amount"] += doc.get("total_amount") or 0
        for master, sa in sale_merged.items():
            snw = sa["net_wt"]
            if snw > 0.001:
                labour_total = sa.get("total_amount") or sa.get("labor") or 0
                item_sale_avg[master] = {
                    "avg_tunch": (sa["fine"] / snw) * 100 if sa["fine"] > 0 else 0,
                    "labor_per_gram": labour_total / snw,
                }

        # Calculate per-supplier profit
        sup_profit = defaultdict(lambda: {"silver": 0.0, "labor": 0.0, "wt": 0.0, "items": set()})
        for doc in sup_agg:
            supplier = doc["_id"]["party"] or "Unknown"
            master = resolve(doc["_id"]["item"])
            avg_sale = item_sale_avg.get(master)
            pw = doc["net_wt"] or 0
            if not avg_sale or pw < 0.001:
                continue
            p_tunch = (doc["fine"] / pw) * 100 if (doc["fine"] or 0) > 0 else 0
            p_lpg = (doc["labor"] or 0) / pw
            sup_profit[supplier]["silver"] += (avg_sale["avg_tunch"] - p_tunch) * pw / 100 / 1000
            sup_profit[supplier]["labor"] += (avg_sale["labor_per_gram"] - p_lpg) * pw
            sup_profit[supplier]["wt"] += pw / 1000
            sup_profit[supplier]["items"].add(master)

        rows = [{"name": k, "silver_profit_kg": round(v["silver"], 3),
                 "labor_profit_inr": round(v["labor"], 2),
                 "total_purchased_kg": round(v["wt"], 3), "items_count": len(v["items"])}
                for k, v in sup_profit.items() if len(v["items"]) > 0]
        rows.sort(key=lambda x: x["silver_profit_kg"], reverse=True)
        return {"view": "supplier", "year": year, "data": rows, "total": len(rows)}

    # === VIEW: ITEM ===
    if view == "item":
        item_agg = await db.historical_transactions.aggregate([
            {"$match": {**match_filter, "type": "sale", "net_wt": {"$gt": 0.001}}},
            {"$group": {
                "_id": "$item_name",
                "fine": {"$sum": "$fine"}, "net_wt": {"$sum": "$net_wt"},
                "labor": {"$sum": "$labor"}, "total_amount": {"$sum": "$total_amount"}, "count": {"$sum": 1}
            }}
        ]).to_list(None)

        # Merge by master item
        master_agg = defaultdict(lambda: {"fine": 0.0, "net_wt": 0.0, "labor": 0.0, "total_amount": 0.0, "count": 0})
        for doc in item_agg:
            master = resolve(doc["_id"])
            master_agg[master]["fine"] += doc["fine"] or 0
            master_agg[master]["net_wt"] += doc["net_wt"] or 0
            master_agg[master]["labor"] += doc["labor"] or 0
            master_agg[master]["total_amount"] += doc.get("total_amount") or 0
            master_agg[master]["count"] += doc["count"]

        rows = []
        for master, sa in master_agg.items():
            basis = purchase_basis.get(master)
            snw = sa["net_wt"]
            if not basis or snw < 0.001 or sa["count"] == 0:
                continue
            sell_tunch = (sa["fine"] / snw) * 100 if sa["fine"] > 0 else 0
            buy_tunch = basis["avg_tunch"]
            silver_kg = round((sell_tunch - buy_tunch) * snw / 100 / 1000, 3)
            s_lpg = (sa.get("total_amount") or sa.get("labor") or 0) / snw
            labor_inr = round((s_lpg - basis["labor_per_gram"]) * snw, 2)
            rows.append({"name": master, "silver_profit_kg": silver_kg, "labor_profit_inr": labor_inr,
                         "total_sold_kg": round(snw / 1000, 3), "transactions": sa["count"],
                         "avg_purchase_tunch": round(buy_tunch, 2), "avg_sale_tunch": round(sell_tunch, 2)})
        rows.sort(key=lambda x: x["silver_profit_kg"], reverse=True)
        return {"view": "item", "year": year, "data": rows, "total": len(rows)}

    # === VIEW: MONTH ===
    if view == "month":
        month_agg = await db.historical_transactions.aggregate([
            {"$match": {**match_filter, "type": "sale", "net_wt": {"$gt": 0.001}}},
            {"$addFields": {"month_key": {"$substr": ["$date", 0, 7]}}},
            {"$group": {
                "_id": {"month": "$month_key", "item": "$item_name"},
                "fine": {"$sum": "$fine"}, "net_wt": {"$sum": "$net_wt"},
                "labor": {"$sum": "$labor"}, "total_amount": {"$sum": "$total_amount"}, "count": {"$sum": 1}
            }}
        ]).to_list(None)

        month_profit = defaultdict(lambda: {"silver": 0.0, "labor": 0.0, "wt": 0.0, "cnt": 0})
        for doc in month_agg:
            mk = doc["_id"]["month"] or "Unknown"
            master = resolve(doc["_id"]["item"])
            basis = purchase_basis.get(master)
            nw = doc["net_wt"] or 0
            if not basis or nw < 0.001:
                month_profit[mk]["cnt"] += doc["count"]
                month_profit[mk]["wt"] += nw / 1000
                continue
            s_tunch = (doc["fine"] / nw) * 100 if (doc["fine"] or 0) > 0 else 0
            s_lpg = (doc.get("total_amount") or doc.get("labor") or 0) / nw
            month_profit[mk]["silver"] += (s_tunch - basis["avg_tunch"]) * nw / 100 / 1000
            month_profit[mk]["labor"] += (s_lpg - basis["labor_per_gram"]) * nw
            month_profit[mk]["wt"] += nw / 1000
            month_profit[mk]["cnt"] += doc["count"]

        MONTH_NAMES = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                      "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        rows = []
        for mk, v in sorted(month_profit.items()):
            parts = mk.split("-")
            if len(parts) == 2:
                yr_short = parts[0][2:]  # "2025" → "25"
                mi = int(parts[1]) if parts[1].isdigit() else 0
                label = f"{MONTH_NAMES[mi]} {yr_short}" if 1 <= mi <= 12 else mk
            else:
                label = mk
            rows.append({"month": mk, "month_name": label,
                         "silver_profit_kg": round(v["silver"], 3),
                         "labor_profit_inr": round(v["labor"], 2),
                         "total_sold_kg": round(v["wt"], 3), "transactions": v["cnt"]})
        return {"view": "month", "year": year, "data": rows, "total": len(rows)}

    raise HTTPException(status_code=400, detail="Invalid view. Use: customer, supplier, item, month, yearly")


# ==================== DATA VISUALIZATION ====================

@api_router.get("/analytics/visualization")
async def get_visualization_data(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    trend_granularity: Optional[str] = "auto",
    current_user: dict = Depends(get_current_user)
):
    """Get aggregated data for charts and visualizations"""
    query = {}
    if start_date and end_date:
        query['date'] = {'$gte': start_date, '$lte': end_date + ' 23:59:59'}
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(50000)
    
    # Get buffer info for tier colors
    buffers = await db.item_buffers.find({}, {"_id": 0}).to_list(None)
    tier_map = {b['item_name']: b.get('tier', 'unknown') for b in buffers}
    
    # Load mappings + groups for resolving to leaders
    all_mappings = await db.item_mappings.find({}, {"_id": 0}).to_list(None)
    mapping_dict = {m['transaction_name']: m['master_name'] for m in all_mappings}
    groups = await db.item_groups.find({}, {"_id": 0}).to_list(1000)
    member_to_leader = {}
    for g in groups:
        for member in g.get('members', []):
            member_to_leader[member] = g['group_name']
    def _resolve(name):
        master = mapping_dict.get(name, name)
        return member_to_leader.get(master, master)

    # 1. Sales by item (top 20) — resolved to leaders
    item_sales = defaultdict(lambda: {'net_wt': 0, 'amount': 0, 'count': 0})
    for t in transactions:
        if t['type'] in ['sale', 'sale_return']:
            name = _resolve(t.get('item_name', ''))
            item_sales[name]['net_wt'] += t.get('net_wt', 0)
            item_sales[name]['amount'] += t.get('total_amount', 0)
            item_sales[name]['count'] += 1
    
    sales_by_item = sorted([
        {'item_name': k, 'net_wt_kg': round(v['net_wt']/1000, 3), 'amount': round(v['amount'], 2), 'count': v['count'], 'tier': tier_map.get(k, 'unknown')}
        for k, v in item_sales.items() if v['net_wt'] > 0
    ], key=lambda x: x['net_wt_kg'], reverse=True)[:30]
    
    # 2. Sales by party (top 20)
    party_sales = defaultdict(lambda: {'net_wt': 0, 'amount': 0, 'count': 0})
    for t in transactions:
        if t['type'] in ['sale', 'sale_return']:
            party = t.get('party_name', 'Unknown')
            party_sales[party]['net_wt'] += t.get('net_wt', 0)
            party_sales[party]['amount'] += t.get('total_amount', 0)
            party_sales[party]['count'] += 1
    
    sales_by_party = sorted([
        {'party_name': k, 'net_wt_kg': round(v['net_wt']/1000, 3), 'amount': round(v['amount'], 2), 'count': v['count']}
        for k, v in party_sales.items() if v['net_wt'] > 0
    ], key=lambda x: x['net_wt_kg'], reverse=True)[:20]
    
    # 3. Purchases by supplier (top 20)
    supplier_purchases = defaultdict(lambda: {'net_wt': 0, 'amount': 0, 'count': 0})
    for t in transactions:
        if t['type'] in ['purchase', 'purchase_return']:
            party = t.get('party_name', 'Unknown')
            supplier_purchases[party]['net_wt'] += t.get('net_wt', 0)
            supplier_purchases[party]['amount'] += t.get('total_amount', 0)
            supplier_purchases[party]['count'] += 1
    
    purchases_by_supplier = sorted([
        {'party_name': k, 'net_wt_kg': round(v['net_wt']/1000, 3), 'amount': round(v['amount'], 2), 'count': v['count']}
        for k, v in supplier_purchases.items() if v['net_wt'] > 0
    ], key=lambda x: x['net_wt_kg'], reverse=True)[:20]
    
    # 4. Tier distribution
    tier_dist = defaultdict(lambda: {'count': 0, 'total_stock_kg': 0})
    for b in buffers:
        t = b.get('tier', 'unknown')
        tier_dist[t]['count'] += 1
        tier_dist[t]['total_stock_kg'] += b.get('current_stock_kg', 0)
    
    tier_distribution = [{'tier': k, 'count': v['count'], 'total_stock_kg': round(v['total_stock_kg'], 3)} for k, v in tier_dist.items()]
    
    # 5. Sales trend (daily or monthly based on granularity)
    # Auto: if date range <= 60 days use daily, else monthly
    all_sale_dates = set()
    monthly_sales = defaultdict(lambda: {'net_wt': 0, 'amount': 0})
    daily_sales = defaultdict(lambda: {'net_wt': 0, 'amount': 0})
    for t in transactions:
        if t['type'] in ['sale', 'sale_return'] and t.get('date'):
            date_str = t['date'][:10]  # YYYY-MM-DD
            month = t['date'][:7]  # YYYY-MM
            all_sale_dates.add(date_str)
            monthly_sales[month]['net_wt'] += t.get('net_wt', 0)
            monthly_sales[month]['amount'] += t.get('total_amount', 0)
            daily_sales[date_str]['net_wt'] += t.get('net_wt', 0)
            daily_sales[date_str]['amount'] += t.get('total_amount', 0)
    
    # Determine granularity
    use_daily = trend_granularity == 'daily'
    if trend_granularity == 'auto' and all_sale_dates:
        sorted_dates = sorted(all_sale_dates)
        try:
            from datetime import datetime as dt_cls
            first = dt_cls.strptime(sorted_dates[0], '%Y-%m-%d')
            last = dt_cls.strptime(sorted_dates[-1], '%Y-%m-%d')
            use_daily = (last - first).days <= 60
        except Exception:
            use_daily = len(all_sale_dates) <= 60
    
    if use_daily:
        sales_trend = sorted([
            {'label': k, 'net_wt_kg': round(v['net_wt']/1000, 3), 'amount': round(v['amount'], 2)}
            for k, v in daily_sales.items()
        ], key=lambda x: x['label'])
    else:
        sales_trend = sorted([
            {'label': k, 'net_wt_kg': round(v['net_wt']/1000, 3), 'amount': round(v['amount'], 2)}
            for k, v in monthly_sales.items()
        ], key=lambda x: x['label'])
    
    # 6. Stock health summary
    status_counts = {'red': 0, 'green': 0, 'yellow': 0}
    for b in buffers:
        s = b.get('status', 'green')
        if s in status_counts:
            status_counts[s] += 1
    
    return {
        "sales_by_item": sales_by_item,
        "sales_by_party": sales_by_party,
        "purchases_by_supplier": purchases_by_supplier,
        "tier_distribution": tier_distribution,
        "sales_trend": sales_trend,
        "trend_granularity": "daily" if use_daily else "monthly",
        "stock_health": status_counts
    }

# ==================== HISTORICAL DATA & SEASONAL ====================

# ---- Business Constants: Wholesale Silver Stock Rotation Model ----
ROTATION_CYCLE_MONTHS = 2.73  # Full stock rotation period

SEASON_PROFILES = {
    'peak': {
        'months': [10, 11, 12, 1, 4, 5],   # Diwali, weddings, Akshaya Tritiya, Sankranti
        'label': 'Peak Season (Festivals & Weddings)',
        'monthly_sales_benchmark_kg': 7000,  # 6500-8000 avg
        'target_total_stock_kg': 10500,
        'lead_time_days': 10,                # orders take longer in peak
    },
    'normal': {
        'months': [2, 3, 6],                # Transition months
        'label': 'Normal Season',
        'monthly_sales_benchmark_kg': 2500,
        'target_total_stock_kg': 8200,
        'lead_time_days': 7,
    },
    'off_season': {
        'months': [7, 8, 9],                # Monsoon / lean months
        'label': 'Off Season (Monsoon)',
        'monthly_sales_benchmark_kg': 1800,
        'target_total_stock_kg': 7500,
        'lead_time_days': 7,
    },
}

# Total stock thresholds (aggregate)
STOCK_FLOOR_KG = 7500   # Below this, sales decline
STOCK_NORMAL_KG = 8200   # Normal operating stock
STOCK_PEAK_KG = 10500    # Max effective stock (sales plateau beyond this)

def get_current_season(month=None):
    """Return the season profile key for a given month."""
    if month is None:
        month = datetime.now(timezone.utc).month
    for key, profile in SEASON_PROFILES.items():
        if month in profile['months']:
            return key
    return 'normal'  # fallback


@api_router.post("/historical/upload")
async def upload_historical_data(
    file_type: str,
    year: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload historical sales/purchase data for AI training (does NOT affect current stock)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    if file_type not in ['sale', 'purchase']:
        raise HTTPException(status_code=400, detail="file_type must be 'sale' or 'purchase'")
    
    content = await file.read()
    logger.info(f"[Historical Direct Upload] file_type={file_type}, year={year}, size={len(content)} bytes")
    
    loop = asyncio.get_event_loop()
    records = await loop.run_in_executor(_parse_executor, parse_excel_file, content, file_type)
    del content  # Free memory
    
    if not records:
        raise HTTPException(status_code=400, detail="No valid records found")
    
    batch_id = str(uuid.uuid4())
    
    # Store in historical_transactions collection (NOT transactions)
    for record in records:
        record['batch_id'] = batch_id
        record['historical_year'] = year
        record['is_historical'] = True

    hist_docs = _prepare_transactions(records, batch_id)
    await batch_insert(db.historical_transactions, hist_docs)
    
    # Verify insertion
    verify_count = await db.historical_transactions.count_documents({"batch_id": batch_id})
    logger.info(f"[Historical Direct Upload] Inserted and verified {verify_count} records for batch {batch_id}")
    
    await save_action('upload_historical', f"Uploaded {verify_count} historical {file_type} records for year {year}", user=current_user)
    
    return {
        "success": True,
        "count": verify_count,
        "year": year,
        "file_type": file_type,
        "batch_id": batch_id,
        "message": f"Uploaded {verify_count} historical {file_type} records for {year}"
    }


@api_router.get("/historical/summary")
async def get_historical_summary(current_user: dict = Depends(get_current_user)):
    """Get summary of uploaded historical data"""
    try:
        pipeline = [
            {"$match": {"historical_year": {"$ne": None, "$exists": True}}},
            {"$group": {
                "_id": {
                    "year": "$historical_year",
                    "type": {"$cond": [
                        {"$in": ["$type", ["sale", "sale_return"]]}, "sale", "purchase"
                    ]}
                },
                "count": {"$sum": 1},
                "total_net_wt": {"$sum": "$net_wt"}
            }},
            {"$sort": {"_id.year": 1, "_id.type": 1}}
        ]
        results = await db.historical_transactions.aggregate(pipeline).to_list(100)
        
        summary = {}
        for r in results:
            year = r['_id'].get('year')
            if not year:
                continue
            ttype = r['_id']['type']
            if year not in summary:
                summary[year] = {}
            total_wt = r.get('total_net_wt', 0) or 0
            summary[year][ttype] = {
                'count': r['count'],
                'total_kg': round(total_wt / 1000, 3)
            }
        
        return {"summary": summary, "years": sorted(summary.keys())}
    except Exception as e:
        # Fallback: direct count if aggregation fails
        logger.error(f"Historical summary aggregation failed: {e}")
        total = await db.historical_transactions.count_documents({})
        if total > 0:
            years = await db.historical_transactions.distinct("historical_year")
            years = [y for y in years if y]
            fallback = {}
            for yr in years:
                sale_c = await db.historical_transactions.count_documents({"historical_year": yr, "type": {"$in": ["sale", "sale_return"]}})
                purch_c = await db.historical_transactions.count_documents({"historical_year": yr, "type": {"$in": ["purchase", "purchase_return"]}})
                fallback[yr] = {}
                if sale_c:
                    fallback[yr]["sale"] = {"count": sale_c, "total_kg": 0}
                if purch_c:
                    fallback[yr]["purchase"] = {"count": purch_c, "total_kg": 0}
            return {"summary": fallback, "years": sorted(years)}
        return {"summary": {}, "years": []}


@api_router.delete("/historical/{year}")
async def delete_historical_year(year: str, current_user: dict = Depends(get_current_user)):
    """Delete historical data for a specific year"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    result = await db.historical_transactions.delete_many({"historical_year": year})
    return {"success": True, "deleted_count": result.deleted_count}


@api_router.get("/historical/debug")
async def debug_historical(current_user: dict = Depends(get_current_user)):
    """Diagnostic endpoint — shows raw DB state for historical data"""
    total = await db.historical_transactions.count_documents({})
    with_year = await db.historical_transactions.count_documents({"historical_year": {"$exists": True, "$ne": None}})
    without_year = total - with_year
    years = await db.historical_transactions.distinct("historical_year")
    types = await db.historical_transactions.distinct("type")
    
    by_year_type = {}
    for yr in (years or []):
        if not yr:
            continue
        by_year_type[yr] = {}
        for t in (types or []):
            c = await db.historical_transactions.count_documents({"historical_year": yr, "type": t})
            if c > 0:
                by_year_type[yr][t] = c
    
    # Check upload sessions
    active_sessions = await db.upload_sessions.count_documents({})
    pending_chunks = await db.upload_chunks.count_documents({})
    
    return {
        "total_records": total,
        "with_historical_year": with_year,
        "without_historical_year": without_year,
        "years": [y for y in years if y],
        "types": types,
        "breakdown": by_year_type,
        "active_upload_sessions": active_sessions,
        "pending_chunks": pending_chunks,
    }



# NOTE: Old /ai/seasonal-analysis endpoint removed — replaced by deterministic
# ML-based seasonal analytics at /api/seasonal/* routes.

@api_router.post("/ai/update-buffers-seasonal")
async def update_buffers_with_seasonal(current_user: dict = Depends(get_current_user)):
    """Update item buffers incorporating seasonal demand patterns"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Get seasonal analysis data via aggregation (memory-efficient)
    monthly_item_sales = defaultdict(lambda: defaultdict(float))
    
    for coll in [db.transactions, db.historical_transactions]:
        pipeline = [
            {"$match": {"type": {"$in": ["sale", "sale_return"]}}},
            {"$project": {"item_name": 1, "net_wt": 1, "month": {"$substr": ["$date", 5, 2]}}},
            {"$group": {"_id": {"item": "$item_name", "month": "$month"}, "total_wt": {"$sum": "$net_wt"}}},
        ]
        async for doc in coll.aggregate(pipeline):
            item = doc['_id']['item']
            try:
                month = int(doc['_id']['month'])
            except (ValueError, TypeError):
                continue
            monthly_item_sales[item][month] += abs(doc['total_wt']) / 1000
    
    now = datetime.now()
    current_month = now.month
    upcoming_months = [(current_month + i - 1) % 12 + 1 for i in range(1, 3)]
    
    updated = 0
    for item, month_data in monthly_item_sales.items():
        total = sum(month_data.values())
        if total < 0.01:
            continue
        avg = total / max(len(month_data), 1)
        
        # Calculate seasonal boost for upcoming period
        upcoming_demand = sum(month_data.get(m, avg) for m in upcoming_months)
        expected = avg * len(upcoming_months)
        boost = upcoming_demand / expected if expected > 0 else 1.0
        boost = max(min(boost, 2.0), 0.5)  # Cap between 0.5x and 2x
        
        # Update buffer with seasonal adjustment
        existing = await db.item_buffers.find_one({'item_name': item}, {"_id": 0})
        if existing:
            base_min = existing.get('minimum_stock_kg', 0)
            base_reorder = existing.get('reorder_buffer_kg', 0)
            base_upper = existing.get('upper_target_kg', 0)
            
            await db.item_buffers.update_one(
                {'item_name': item},
                {'$set': {
                    'seasonal_boost': round(boost, 2),
                    'seasonal_min_stock_kg': round(base_min * boost, 3),
                    'seasonal_reorder_buffer_kg': round(base_reorder * boost, 3),
                    'seasonal_upper_target_kg': round(base_upper * boost, 3),
                    'seasonal_updated_at': now.isoformat()
                }}
            )
            updated += 1
    
    await save_action('seasonal_buffer_update', f"Updated {updated} items with seasonal adjustments", user=current_user)
    return {"success": True, "items_updated": updated}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
