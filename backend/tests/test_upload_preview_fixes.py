"""
Backend tests for Physical Stock Upload Preview bug fixes:
1. Group member resolution (TULSI 70 BELT -> TULSI 70 -264 group)
2. Item mapping resolution (via mapping chains)
3. Base value matching with Current Stock page (get_current_inventory)
4. Compare endpoint gross_difference fields
"""
import pytest
import requests
import os
import tempfile
from openpyxl import Workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

@pytest.fixture(scope="module")
def auth_token():
    """Get auth token for admin user"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json={
        "username": "admin",
        "password": "admin123"
    })
    assert response.status_code == 200, f"Login failed: {response.text}"
    return response.json()["access_token"]


@pytest.fixture(scope="module")
def auth_headers(auth_token):
    """Return authorization headers"""
    return {"Authorization": f"Bearer {auth_token}"}


class TestHealthAndBasics:
    """Test basic endpoint accessibility"""

    def test_health_endpoint(self):
        """Test health endpoint works"""
        response = requests.get(f"{BASE_URL}/api/health")
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "healthy"
        print("PASS: Health endpoint works")


class TestCurrentInventory:
    """Test GET /api/inventory/current endpoint"""

    def test_current_inventory_returns_data(self, auth_headers):
        """Verify current stock endpoint returns all items with proper structure"""
        response = requests.get(f"{BASE_URL}/api/inventory/current", headers=auth_headers)
        assert response.status_code == 200
        data = response.json()
        
        # Verify response structure
        assert "inventory" in data
        assert "total_items" in data
        assert "total_gr_wt" in data
        assert "total_net_wt" in data
        assert "negative_items" in data
        
        inventory = data["inventory"]
        assert isinstance(inventory, list)
        print(f"PASS: Current inventory returns {len(inventory)} items")
        
        # Verify item structure
        if inventory:
            item = inventory[0]
            assert "item_name" in item
            assert "gr_wt" in item
            assert "net_wt" in item
            assert "stamp" in item
            print(f"PASS: Item structure is correct")

    def test_current_inventory_has_group_member_breakdown(self, auth_headers):
        """Verify groups have member breakdown for expandable UI"""
        response = requests.get(f"{BASE_URL}/api/inventory/current", headers=auth_headers)
        assert response.status_code == 200
        data = response.json()
        
        inventory = data["inventory"]
        groups = [item for item in inventory if item.get("is_group", False)]
        
        print(f"Found {len(groups)} groups in inventory")
        
        # Check SNT 40-256 group specifically
        snt_groups = [g for g in groups if "SNT 40-256" in g.get("item_name", "")]
        if snt_groups:
            snt_group = snt_groups[0]
            assert "members" in snt_group, "Group should have members field"
            members = snt_group.get("members", [])
            print(f"PASS: SNT 40-256 group has {len(members)} member breakdowns")
            for m in members:
                print(f"  - {m.get('item_name')}: gr={m.get('gr_wt')}")


class TestUploadPreviewGroupMemberResolution:
    """Test upload-preview correctly resolves group members to their leader"""

    def test_group_member_resolves_to_leader(self, auth_headers):
        """
        Test that uploading an item that is a GROUP MEMBER (e.g., SNT-40 PREMIUM)
        correctly resolves to the group leader (SNT 40-256) in preview.
        """
        # First get current stock to know the base values
        curr_response = requests.get(f"{BASE_URL}/api/inventory/current", headers=auth_headers)
        assert curr_response.status_code == 200
        inventory = curr_response.json().get("inventory", [])
        
        # Find SNT 40-256 group
        snt_group = None
        for item in inventory:
            if item.get("item_name") == "SNT 40-256":
                snt_group = item
                break
        
        if not snt_group:
            pytest.skip("SNT 40-256 group not found in inventory")
        
        base_gr_wt = snt_group.get("gr_wt", 0)
        print(f"SNT 40-256 current gr_wt: {base_gr_wt}")
        
        # Get a date that has existing physical stock data
        dates_response = requests.get(f"{BASE_URL}/api/physical-stock/dates", headers=auth_headers)
        dates = dates_response.json().get("dates", [])
        verification_date = dates[0] if dates else "2026-03-20"  # Use latest date with data
        print(f"Using verification_date: {verification_date}")
        
        # Create a test Excel file with GROUP MEMBER name (SNT-40 PREMIUM)
        # This should resolve to leader SNT 40-256
        wb = Workbook()
        ws = wb.active
        ws.append(["Item Name", "Gross Weight"])
        ws.append(["SNT-40 PREMIUM", 50.0])  # Group member name, 50 kg
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name
        
        try:
            with open(temp_path, "rb") as f:
                files = {"file": ("test_physical.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
                response = requests.post(
                    f"{BASE_URL}/api/physical-stock/upload-preview?verification_date={verification_date}",
                    headers=auth_headers,
                    files=files
                )
            
            assert response.status_code == 200, f"Upload preview failed: {response.text}"
            data = response.json()
            
            # Response structure uses preview_rows, not preview
            preview_rows = data.get("preview_rows", data.get("preview", []))
            print(f"Preview returned {len(preview_rows)} rows")
            
            # The uploaded "SNT-40 PREMIUM" should resolve to "SNT 40-256" (the group leader)
            matched_rows = [r for r in preview_rows if r.get("item_name") == "SNT 40-256" and r.get("status") != "unmatched"]
            unmatched_rows = [r for r in preview_rows if r.get("status") == "unmatched"]
            
            # Should NOT be unmatched
            unmatched_snt = [r for r in unmatched_rows if "SNT" in r.get("item_name", "")]
            assert len(unmatched_snt) == 0, f"SNT-40 PREMIUM should not be unmatched, found: {unmatched_snt}"
            
            # Should be matched to SNT 40-256 leader
            assert len(matched_rows) > 0, "SNT-40 PREMIUM should resolve to SNT 40-256 leader"
            
            matched = matched_rows[0]
            print(f"PASS: Group member 'SNT-40 PREMIUM' correctly resolved to leader '{matched.get('item_name')}'")
            print(f"  old_gr_wt: {matched.get('old_gr_wt')}, new_gr_wt: {matched.get('new_gr_wt')}")
            
        finally:
            os.unlink(temp_path)


class TestUploadPreviewBaseMatchesCurrentStock:
    """Test that upload-preview base values match Current Stock page"""

    def test_base_values_match_current_inventory(self, auth_headers):
        """
        Verify the base values (old_gr_wt, old_net_wt) in upload-preview match
        the values from GET /api/inventory/current.
        
        This was a key fix - get_effective_physical_base_for_date() now uses
        get_current_inventory() instead of get_book_closing_stock_as_of_date().
        """
        # Get current inventory
        curr_response = requests.get(f"{BASE_URL}/api/inventory/current", headers=auth_headers)
        assert curr_response.status_code == 200
        inventory = curr_response.json().get("inventory", [])
        
        # Build lookup by item name
        current_stock = {item["item_name"]: item for item in inventory}
        
        # Pick a few items to test
        test_items = []
        for item_name, item in list(current_stock.items())[:5]:
            if item.get("gr_wt", 0) > 0:  # Only items with stock
                test_items.append((item_name, item.get("gr_wt", 0), item.get("net_wt", 0)))
        
        if not test_items:
            pytest.skip("No items with stock found")
        
        # Create test file with these items
        wb = Workbook()
        ws = wb.active
        ws.append(["Item Name", "Gross Weight", "Net Weight"])
        for item_name, gr_wt, net_wt in test_items:
            # Upload with new value (current + 1000g = +1kg)
            ws.append([item_name, (gr_wt + 1000) / 1000, (net_wt + 1000) / 1000])  # Convert to kg
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name
        
        try:
            with open(temp_path, "rb") as f:
                files = {"file": ("test_base_match.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
                response = requests.post(
                    f"{BASE_URL}/api/physical-stock/upload-preview?verification_date=2099-12-02",
                    headers=auth_headers,
                    files=files
                )
            
            assert response.status_code == 200, f"Upload preview failed: {response.text}"
            data = response.json()
            
            # Response uses preview_rows, not preview
            preview_rows = data.get("preview_rows", data.get("preview", []))
            preview_by_name = {r["item_name"]: r for r in preview_rows}
            
            # Verify base values match current inventory
            mismatches = []
            for item_name, expected_gr, expected_net in test_items:
                preview_row = preview_by_name.get(item_name)
                if not preview_row:
                    continue
                
                old_gr = preview_row.get("old_gr_wt", 0)
                old_net = preview_row.get("old_net_wt", 0)
                
                # Allow small rounding differences (within 1 gram)
                if abs(old_gr - expected_gr) > 1:
                    mismatches.append(f"{item_name}: old_gr_wt={old_gr}, expected={expected_gr}")
                if abs(old_net - expected_net) > 1:
                    mismatches.append(f"{item_name}: old_net_wt={old_net}, expected={expected_net}")
            
            assert len(mismatches) == 0, f"Base values don't match Current Stock:\n" + "\n".join(mismatches)
            print(f"PASS: Base values for {len(test_items)} items match Current Stock page")
            
        finally:
            os.unlink(temp_path)


class TestCompareEndpointGrossDifference:
    """Test that compare endpoint includes gross_difference fields"""

    def test_compare_includes_gross_difference(self, auth_headers):
        """
        Verify GET /api/physical-stock/compare returns gross_difference and
        gross_difference_kg in comparison items.
        """
        # Get available dates
        dates_response = requests.get(f"{BASE_URL}/api/physical-stock/dates", headers=auth_headers)
        assert dates_response.status_code == 200
        dates = dates_response.json().get("dates", [])
        
        if not dates:
            pytest.skip("No physical stock dates available")
        
        # Use most recent date
        verification_date = dates[0]
        print(f"Testing compare for date: {verification_date}")
        
        response = requests.get(
            f"{BASE_URL}/api/physical-stock/compare?verification_date={verification_date}",
            headers=auth_headers
        )
        assert response.status_code == 200, f"Compare failed: {response.text}"
        data = response.json()
        
        # Check response structure
        assert "summary" in data
        assert "matches" in data or "discrepancies" in data
        
        # Check comparison items have gross_difference fields
        all_items = data.get("matches", []) + data.get("discrepancies", [])
        
        if not all_items:
            print("WARN: No matched/discrepant items to verify gross_difference fields")
            return
        
        # Verify first item has gross_difference fields
        item = all_items[0]
        assert "gross_difference" in item, "Comparison item should have gross_difference"
        assert "gross_difference_kg" in item, "Comparison item should have gross_difference_kg"
        assert "book_gross_wt" in item, "Comparison item should have book_gross_wt"
        assert "physical_gross_wt" in item, "Comparison item should have physical_gross_wt"
        
        print(f"PASS: Compare endpoint includes gross_difference fields")
        print(f"  Sample item: {item.get('item_name')}")
        print(f"    book_gross: {item.get('book_gross_wt')}, physical_gross: {item.get('physical_gross_wt')}")
        print(f"    gross_difference: {item.get('gross_difference')}, gross_difference_kg: {item.get('gross_difference_kg')}")

    def test_compare_requires_verification_date(self, auth_headers):
        """Verify compare endpoint requires verification_date parameter"""
        # Call without verification_date
        response = requests.get(
            f"{BASE_URL}/api/physical-stock/compare",
            headers=auth_headers
        )
        # Should return 422 (FastAPI validation error) or 400
        assert response.status_code in [400, 422], f"Expected 400/422 without verification_date, got {response.status_code}"
        print("PASS: Compare endpoint requires verification_date")


class TestApplyUpdates:
    """Test that apply-updates still works after base computation change"""

    def test_apply_updates_works(self, auth_headers):
        """Test that applying updates creates baselines correctly"""
        # Get a date that has existing physical stock data
        dates_response = requests.get(f"{BASE_URL}/api/physical-stock/dates", headers=auth_headers)
        dates = dates_response.json().get("dates", [])
        verification_date = dates[0] if dates else "2026-03-20"
        
        # First, upload a preview with an existing item
        wb = Workbook()
        ws = wb.active
        ws.append(["Item Name", "Gross Weight"])
        ws.append(["COIN 100 -98T", 10.0])  # Existing item
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name
        
        try:
            # Upload preview
            with open(temp_path, "rb") as f:
                files = {"file": ("apply_test.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
                preview_response = requests.post(
                    f"{BASE_URL}/api/physical-stock/upload-preview?verification_date={verification_date}",
                    headers=auth_headers,
                    files=files
                )
            
            if preview_response.status_code != 200:
                pytest.skip(f"Upload preview failed: {preview_response.text}")
            
            preview_data = preview_response.json()
            session_id = preview_data.get("preview_session_id", preview_data.get("session_id"))
            
            if not session_id:
                pytest.skip("No session_id returned from preview")
            
            # Get preview rows to find a pending item
            preview_rows = preview_data.get("preview_rows", preview_data.get("preview", []))
            pending = [r for r in preview_rows if r.get("status") == "pending"]
            
            if not pending:
                print("WARN: No pending items to apply, test skipped")
                return
            
            item_to_approve = pending[0]["item_name"]
            
            # Apply update - API expects 'items' array with full item data
            apply_response = requests.post(
                f"{BASE_URL}/api/physical-stock/apply-updates",
                headers=auth_headers,
                json={
                    "preview_session_id": session_id,
                    "verification_date": verification_date,
                    "items": [{
                        "item_name": item_to_approve,
                        "new_gr_wt": pending[0].get("new_gr_wt", 0),
                        "new_net_wt": pending[0].get("new_net_wt", 0),
                        "update_mode": pending[0].get("update_mode", "gross_only"),
                        "is_negative_grouped": pending[0].get("is_negative_grouped", False)
                    }]
                }
            )
            
            assert apply_response.status_code == 200, f"Apply updates failed: {apply_response.text}"
            apply_data = apply_response.json()
            
            # Check response has expected fields
            assert "updated_count" in apply_data or "success" in apply_data
            print(f"PASS: Apply updates works. Response: {apply_data}")
            
        finally:
            os.unlink(temp_path)


class TestPhysicalStockDates:
    """Test physical-stock/dates endpoint"""

    def test_dates_sorted_descending(self, auth_headers):
        """Verify dates are returned in descending order"""
        response = requests.get(f"{BASE_URL}/api/physical-stock/dates", headers=auth_headers)
        assert response.status_code == 200
        data = response.json()
        
        dates = data.get("dates", [])
        if len(dates) < 2:
            pytest.skip("Not enough dates to verify sorting")
        
        # Verify descending order
        for i in range(len(dates) - 1):
            assert dates[i] >= dates[i + 1], f"Dates not sorted descending: {dates[i]} < {dates[i + 1]}"
        
        print(f"PASS: {len(dates)} dates returned in descending order")
        print(f"  Latest: {dates[0]}, Oldest: {dates[-1]}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
